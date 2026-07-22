#!/usr/bin/env python3
"""
Monitor de Imóveis — Algarve  v4.0
====================================
Novidades v4:
  1. Autenticação (login com password)
  2. Extração de detalhes completos (área, ano, GPS, descrição)
  3. Alertas de scraper com falha persistente (3+ rondas)
  4. Comparação lado a lado (até 3 imóveis)
  5. Exportar favoritos para Excel
  6. Mapa interativo (Leaflet + OpenStreetMap)
  7. Estimativa de custos (IMT, Imposto de Selo, Registo, etc.)
  8. Histórico de visitas com notas
"""

import os, re, time, json, random, logging, schedule, threading
from collections import deque, defaultdict as _dd, OrderedDict
from urllib.parse import urljoin, urlparse as _urlparse
import hashlib, secrets, functools
import requests
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# curl_cffi — impersonates real browsers at TLS level (very hard to detect)
try:
    from curl_cffi import requests as cffi_requests
    CURL_CFFI_AVAILABLE = True
except ImportError:
    CURL_CFFI_AVAILABLE = False

# cloudscraper — bypasses Cloudflare JS challenges without browser
try:
    import cloudscraper
    _cloudscraper = cloudscraper.create_scraper(
        browser={"browser":"chrome","platform":"windows","mobile":False})
    CLOUDSCRAPER_AVAILABLE = True
except ImportError:
    CLOUDSCRAPER_AVAILABLE = False
    _cloudscraper = None
import psycopg
from psycopg import rows as psycopg_rows
from datetime import datetime, timedelta
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from bs4 import BeautifulSoup

def safe_soup(html, fonte="?"):
    """
    Parse HTML com validação completa.
    Nunca chama BeautifulSoup com dados inválidos —
    assim o MarkupResemblesLocatorWarning nunca aparece.
    """
    if not html:
        return None
    if not isinstance(html, str):
        logging.getLogger("algarve-monitor").debug(f"safe_soup [{fonte}]: tipo inválido {type(html)}")
        return None
    stripped = html.strip()
    # URL em vez de HTML
    if stripped.startswith(("http://","https://")) and len(stripped) < 500:
        logging.getLogger("algarve-monitor").warning(f"safe_soup [{fonte}]: recebido URL em vez de HTML: {stripped[:80]}")
        return None
    # Caminho de ficheiro
    if stripped.startswith(("/","./","../")) and "\n" not in stripped and len(stripped) < 300:
        logging.getLogger("algarve-monitor").warning(f"safe_soup [{fonte}]: parece caminho de ficheiro: {stripped[:80]}")
        return None
    # Muito curto para ser HTML útil
    if len(stripped) < 50:
        return None
    # Sem tags HTML
    if "<html" not in stripped.lower() and "<body" not in stripped.lower() and "<div" not in stripped.lower():
        logging.getLogger("algarve-monitor").debug(f"safe_soup [{fonte}]: sem tags HTML ({len(stripped)} chars)")
        return None
    return BeautifulSoup(html, "html.parser")
from flask import (Flask, render_template_string, jsonify, request,
                   session, redirect, url_for, make_response)
# Playwright replaces Selenium for better JS rendering
# Patchright > Playwright para anti-deteção
# Camoufox disponível para sites muito protegidos (usado separadamente)
try:
    from patchright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout
    PLAYWRIGHT_AVAILABLE = True
    PLAYWRIGHT_ENGINE = "patchright"
except ImportError:
    try:
        from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout
        PLAYWRIGHT_AVAILABLE = True
        PLAYWRIGHT_ENGINE = "playwright"
    except ImportError:
        PLAYWRIGHT_AVAILABLE = False
        PLAYWRIGHT_ENGINE = None

try:
    from camoufox.sync_api import Camoufox
    CAMOUFOX_AVAILABLE = True
except ImportError:
    CAMOUFOX_AVAILABLE = False

logging.basicConfig(level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s", datefmt="%d/%m/%Y %H:%M:%S")
log = logging.getLogger("algarve-monitor")

# In-memory log buffer para /api/logs
_log_buffer = deque(maxlen=500)

class _BufferHandler(logging.Handler):
    def emit(self, record):
        from datetime import datetime as _dtnow
        _log_buffer.append({
            "ts": _dtnow.now().strftime("%d/%m/%Y %H:%M:%S"),
            "level": record.levelname,
            "msg": record.getMessage()
        })
_buf = _BufferHandler()
log.addHandler(_buf)

# ============================================================
# CONFIG
# ============================================================
EMAIL_REMETENTE    = os.getenv("EMAIL_REMETENTE",    "o_teu_email@gmail.com")
EMAIL_PASSWORD     = os.getenv("EMAIL_PASSWORD",     "a_tua_app_password")
EMAIL_DESTINATARIO = os.getenv("EMAIL_DESTINATARIO", EMAIL_REMETENTE)
INTERVALO_HORAS    = int(os.getenv("INTERVALO_HORAS", "24"))
DATABASE_URL       = os.getenv("DATABASE_URL", "")
TELEGRAM_BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN", "")
def _load_multi_keys(base_name):
    """Carrega KEY, KEY_1, KEY_2, ... — devolve lista de keys disponíveis."""
    keys = []
    base = os.getenv(base_name, "")
    if base: keys.append(base)
    for i in range(1, 6):  # suporta até _5
        k = os.getenv(f"{base_name}_{i}", "")
        if k and k not in keys: keys.append(k)
    return keys

# Multi-key: suporta SCRAPERAPI_KEY, SCRAPERAPI_KEY_1, SCRAPERAPI_KEY_2, ...
_SCRAPERAPI_KEYS  = _load_multi_keys("SCRAPERAPI_KEY")
_SCRAPINGBEE_KEYS = _load_multi_keys("SCRAPINGBEE_KEY")
_SCRAPEDO_KEYS    = _load_multi_keys("SCRAPEDO_KEY")
_key_index = {}  # provider -> índice da key activa

def _get_active_key(provider):
    """Devolve a key activa para o provider (rotação quando esgota)."""
    keys_map = {
        "scraperapi": _SCRAPERAPI_KEYS,
        "scrapingbee": _SCRAPINGBEE_KEYS,
        "scrapedo": _SCRAPEDO_KEYS,
    }
    keys = keys_map.get(provider, [])
    if not keys: return ""
    idx = _key_index.get(provider, 0)
    return keys[idx % len(keys)]

def _rotate_key(provider):
    """Roda para a próxima key quando a actual esgota. True se há mais keys."""
    keys_map = {
        "scraperapi": _SCRAPERAPI_KEYS,
        "scrapingbee": _SCRAPINGBEE_KEYS,
        "scrapedo": _SCRAPEDO_KEYS,
    }
    keys = keys_map.get(provider, [])
    if not keys: return False
    idx = _key_index.get(provider, 0) + 1
    if idx < len(keys):
        _key_index[provider] = idx
        log.warning(f"  🔄 {provider}: a rodar para key #{idx+1}/{len(keys)}")
        try: _save_provider_state()   # persiste índice entre restarts
        except Exception: pass
        return True
    return False  # todas as keys esgotadas

SCRAPERAPI_KEY     = _SCRAPERAPI_KEYS[0] if _SCRAPERAPI_KEYS else ""
ZENROWS_KEY        = os.getenv("ZENROWS_KEY", "")
SCRAPINGBEE_KEY    = _SCRAPINGBEE_KEYS[0] if _SCRAPINGBEE_KEYS else ""
CRAWLBASE_KEY      = os.getenv("CRAWLBASE_KEY", "")
BROWSERLESS_TOKEN  = os.getenv("BROWSERLESS_TOKEN", "")
SCRAPEDO_KEY       = _SCRAPEDO_KEYS[0] if _SCRAPEDO_KEYS else ""
SCRAPINGANT_KEY    = os.getenv("SCRAPINGANT_KEY", "")
WEBSHARE_KEY       = os.getenv("WEBSHARE_KEY", "")

# Dicionário central de API keys — adicionar novos providers aqui
API_KEYS = {
    "scrapingant":  SCRAPINGANT_KEY,
    "scraperapi":   SCRAPERAPI_KEY,
    "zenrows":      ZENROWS_KEY,
    "crawlbase":    CRAWLBASE_KEY,
    "scrapedo":     SCRAPEDO_KEY,
    "scrapingbee":  SCRAPINGBEE_KEY,
    "webshare":     WEBSHARE_KEY,
}
PORT               = int(os.getenv("PORT", "8080"))
GOOGLE_MAPS_KEY    = os.getenv("GOOGLE_MAPS_KEY", "")    # opcional para geocoding
VAPID_PUBLIC_KEY   = os.getenv("VAPID_PUBLIC_KEY", "")
VAPID_PRIVATE_KEY  = os.getenv("VAPID_PRIVATE_KEY", "")
VAPID_EMAIL        = os.getenv("VAPID_EMAIL", EMAIL_REMETENTE)

# Autenticação
DASHBOARD_USERNAME = os.getenv("DASHBOARD_USERNAME", "admin")
_DASHBOARD_PASSWORD_RAW = os.getenv("DASHBOARD_PASSWORD", "")
if not _DASHBOARD_PASSWORD_RAW:
    _DASHBOARD_PASSWORD_RAW = secrets.token_urlsafe(16)
    log.warning(f"⚠️  DASHBOARD_PASSWORD não definida — password temporária: {_DASHBOARD_PASSWORD_RAW}")
DASHBOARD_PASSWORD = _DASHBOARD_PASSWORD_RAW
SECRET_KEY         = os.getenv("SECRET_KEY", secrets.token_hex(32))

PERFIS = [
    {
        "nome":          "Sotavento — T2+",
        "email":         os.getenv("PERFIL_1_EMAIL", EMAIL_DESTINATARIO),
        "telegram_chat": os.getenv("PERFIL_1_TELEGRAM_CHAT", ""),
        # 5M por omissão = URLs de pesquisa dos sites devolvem o inventário
        # inteiro; o filtro fino é feito na aplicação (slider do dashboard)
        "preco_max":     int(os.getenv("PERFIL_1_PRECO_MAX",  "5000000")),
        "quartos_min":   int(os.getenv("PERFIL_1_QUARTOS_MIN", "2")),
        "tipos":         ["apartamentos", "moradias-e-vivendas"],
        "zonas":         ["faro", "tavira", "olhao",
                          "vila-real-de-santo-antonio", "castro-marim"],
    },
    # {
    #     "nome":          "Lagos — Moradia T3+ até 400k",
    #     "email":         os.getenv("PERFIL_2_EMAIL", "outro@gmail.com"),
    #     "telegram_chat": os.getenv("PERFIL_2_TELEGRAM_CHAT", ""),
    #     "preco_max":     int(os.getenv("PERFIL_2_PRECO_MAX",  "400000")),
    #     "quartos_min":   int(os.getenv("PERFIL_2_QUARTOS_MIN", "3")),
    #     "tipos":         ["moradias-e-vivendas"],
    #     "zonas":         ["lagos", "portimao", "silves"],
    # },
]

TODAS_AS_ZONAS = {
    "faro": "Faro", "tavira": "Tavira", "olhao": "Olhão",
    "vila-real-de-santo-antonio": "VRSA", "castro-marim": "Castro Marim",
    "lagos": "Lagos", "portimao": "Portimão", "silves": "Silves",
    "aljezur": "Aljezur", "albufeira": "Albufeira",
    "lagoa-algarve": "Lagoa", "loul": "Loulé", "monchique": "Monchique",
}

ZONA_SCORE = {
    "Tavira": 10, "Faro": 9, "Olhão": 9, "VRSA": 8, "Castro Marim": 7,
    "Lagos": 8, "Portimão": 7,
}

USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 Chrome/123.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:125.0) Gecko/20100101 Firefox/125.0",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 Chrome/122.0.0.0 Safari/537.36",
]

def random_headers():
    return {"User-Agent": random.choice(USER_AGENTS),
            "Accept-Language": "pt-PT,pt;q=0.9,en;q=0.8",
            "Accept": "text/html,application/xhtml+xml,*/*;q=0.8"}

_proxy_counter = 0
_scraper_metrics  = {}  # métricas da última ronda por scraper

# ── Free Proxy Pool (Webshare + ProxyScrape) ─────────────────
_free_proxy_pool   = []   # ["ip:port", ...]
_free_proxy_bad    = set() # proxies que falharam
_free_proxy_ts     = 0     # último refresh
_FREE_PROXY_TTL    = 1800  # refresh a cada 30 min

def _validar_proxy(proxy_addr, test_url="http://httpbin.org/ip", timeout=3):
    """Testa proxy com curl_cffi (mais rápido e difícil de bloquear)."""
    px = {"http": f"http://{proxy_addr}", "https": f"http://{proxy_addr}"}
    try:
        if CURL_CFFI_AVAILABLE:
            r = cffi_requests.get(test_url, timeout=timeout, impersonate="chrome120",
                                  proxies=px, verify=False)
        else:
            r = requests.get(test_url, timeout=timeout, proxies=px,
                             headers={"User-Agent":"Mozilla/5.0"}, verify=False)
        return r.status_code == 200
    except Exception:
        return False

def _refresh_free_proxies():
    """Atualiza lista de proxies gratuitos de múltiplas fontes."""
    global _free_proxy_pool, _free_proxy_ts
    if time.time() - _free_proxy_ts < _FREE_PROXY_TTL:
        return

    pool = []

    # 1. ProxyScrape — elite proxies
    try:
        r = requests.get(
            "https://api.proxyscrape.com/v2/?request=displayproxies"
            "&protocol=http&timeout=5000&country=all&ssl=all&anonymity=elite",
            timeout=15)
        proxies = [p.strip() for p in r.text.splitlines()
                   if p.strip() and ':' in p and len(p.strip()) < 30]
        pool.extend(proxies[:40])
        log.info(f"  ProxyScrape: {len(proxies)} proxies")
    except Exception as e:
        log.debug(f"  ProxyScrape: {e}")

    # 2. Webshare (free tier)
    if WEBSHARE_KEY:
        try:
            r = requests.get("https://proxy.webshare.io/api/v2/proxy/list/"
                "?mode=direct&page=1&page_size=25",
                headers={"Authorization": f"Token {WEBSHARE_KEY}"}, timeout=15)
            ws = [f"{p['proxy_address']}:{p['port']}"
                  for p in r.json().get("results", []) if p.get("valid")]
            pool.extend(ws)
            log.info(f"  Webshare: {len(ws)} proxies")
        except Exception as e:
            log.debug(f"  Webshare: {e}")

    # 3. Geonode Free API
    try:
        r = requests.get(
            "https://proxylist.geonode.com/api/proxy-list"
            "?limit=50&page=1&sort_by=lastChecked&sort_type=desc"
            "&protocols=http&anonymityLevel=elite",
            timeout=10)
        gn = [f"{p['ip']}:{p['port']}" for p in r.json().get("data", [])]
        pool.extend(gn)
        log.info(f"  Geonode: {len(gn)} proxies")
    except Exception as e:
        log.debug(f"  Geonode: {e}")

    # 4. Monosans proxy list (GitHub raw)
    try:
        r = requests.get(
            "https://raw.githubusercontent.com/monosans/proxy-list/main/proxies/http.txt",
            timeout=10)
        ms = [p.strip() for p in r.text.splitlines()
              if p.strip() and ':' in p and len(p.strip()) < 30]
        pool.extend(ms[:50])
        log.info(f"  Monosans: {len(ms)} proxies")
    except Exception as e:
        log.debug(f"  Monosans: {e}")

    # 5. TheSpeedX proxy list (GitHub raw)
    try:
        r = requests.get(
            "https://raw.githubusercontent.com/TheSpeedX/SOCKS-List/master/http.txt",
            timeout=10)
        sx = [p.strip() for p in r.text.splitlines()
              if p.strip() and ':' in p and len(p.strip()) < 30]
        pool.extend(sx[:50])
        log.info(f"  TheSpeedX: {len(sx)} proxies")
    except Exception as e:
        log.debug(f"  TheSpeedX: {e}")

    # Remove proxies conhecidamente maus e duplicados
    pool = list(dict.fromkeys(p for p in pool if p not in _free_proxy_bad))
    random.shuffle(pool)

    # Valida amostra (max 20) para garantir qualidade mínima
    if len(pool) > 10:
        amostra = pool[:20]
        validos = [p for p in amostra if _validar_proxy(p)]
        taxa = len(validos) / len(amostra) * 100
        log.info(f"  Proxy validation: {len(validos)}/{len(amostra)} OK ({taxa:.0f}%)")

    _free_proxy_pool = pool
    _free_proxy_ts   = time.time()
    log.info(f"  Free proxy pool: {len(pool)} proxies de 5 fontes")

def _try_free_proxy(url, timeout=15):
    """Tenta obter URL via pool de proxies gratuitos. Retorna response ou None."""
    _refresh_free_proxies()
    if not _free_proxy_pool:
        return None

    # Tenta até 3 proxies diferentes
    tried = 0
    for proxy_addr in list(_free_proxy_pool):
        if proxy_addr in _free_proxy_bad or tried >= 3:
            break
        tried += 1
        try:
            t0 = time.time()
            proxies_dict = {"http": f"http://{proxy_addr}",
                            "https": f"http://{proxy_addr}"}
            if CURL_CFFI_AVAILABLE:
                r = cffi_requests.get(url, timeout=timeout, impersonate="chrome120",
                    proxies=proxies_dict, headers=random_headers(), verify=False)
            else:
                r = requests.get(url, timeout=timeout,
                    proxies=proxies_dict, headers=random_headers(), verify=False)
            ms = int((time.time()-t0)*1000)
            if r.status_code == 200 and len(r.text) > 500:
                _record_provider("direto", True, ms)
                return r
            else:
                _free_proxy_bad.add(proxy_addr)
        except Exception:
            _free_proxy_bad.add(proxy_addr)
    return None

# ── Provider metrics ──────────────────────────────────────────
from datetime import date as _date  # _dt já importado como datetime
_dt = datetime  # alias para retrocompatibilidade

# Carrega configuração de providers (providers.json ou defaults)
def _load_provider_config():
    import json as _json
    defaults = {
        "scrapingant": {"priority":1,"monthly_limit":10000,"cost_weight":0.1,"js":True},
        "scraperapi":  {"priority":2,"monthly_limit":5000, "cost_weight":0.2,"js":True},
        "zenrows":     {"priority":3,"monthly_limit":1000, "cost_weight":0.3,"js":True},
        "crawlbase":   {"priority":4,"monthly_limit":1000, "cost_weight":0.3,"js":True},
        "scrapedo":    {"priority":5,"monthly_limit":1000, "cost_weight":0.2,"js":True},
        "scrapingbee": {"priority":6,"monthly_limit":1000, "cost_weight":0.3,"js":True},
        "webshare":    {"priority":7,"monthly_limit":99999,"cost_weight":0.0,"js":False},
        "proxyscrape": {"priority":8,"monthly_limit":99999,"cost_weight":0.0,"js":False},
        "playwright":  {"priority":99,"monthly_limit":99999,"cost_weight":0.0,"js":True},
        "direto":      {"priority":0,"monthly_limit":99999,"cost_weight":0.0,"js":False},
    }
    cfg_path = os.path.join(os.path.dirname(__file__), "providers.json")
    if os.path.exists(cfg_path):
        try:
            with open(cfg_path) as f:
                data = _json.load(f)
            defaults.update(data.get("providers", {}))
            log.info(f"providers.json carregado: {list(data.get('providers',{}).keys())}")
        except Exception as e:
            log.warning(f"providers.json erro: {e} — a usar defaults")
    return defaults

PROVIDER_CONFIG = _load_provider_config()
_PROVIDERS = list(PROVIDER_CONFIG.keys())

# Global stats per provider
_provider_data = {p: {
    "requests_today": 0, "success_today": 0,
    "latencies": [], "last_error": None,
    "exhausted_at": None, "cooldown_until": None,
    "consecutive_failures": 0,
    "circuit_open": False,
    "circuit_open_until": None,
    "circuit_half_open": False,
} for p in _PROVIDERS}

# Semáforos de concorrência por provider (respeita max_parallel do providers.json)
_provider_semaphores = {
    p: threading.Semaphore(PROVIDER_CONFIG.get(p, {}).get("max_parallel", 5))
    for p in _PROVIDERS
}

# Per-domain stats: domain -> provider -> {success, total, latencies}
_domain_stats = _dd(lambda: _dd(lambda: {"success":0,"total":0,"latencies":[]}))

# HTTP response cache (url → (timestamp, html))
_http_cache     = OrderedDict()
_http_cache_ttl = 1800  # 30 minutos

_proxy_stats    = {}
_proxy_exhausted = {}
_proxy_success  = {}
_proxy_fail     = {}
_proxy_time     = {}
_proxy_cooldown = {}
_session_exhausted = set()  # providers esgotados NESTA sessão — nunca tentados
_cache_lock     = threading.Lock()
_stats_lock     = threading.Lock()
_exhausted_lock = threading.Lock()

TIMEOUT_PROXY_DEFAULT = 60
TIMEOUT_PROXY_FAST    = 30
TIMEOUT_DIRETO        = 15
TIMEOUT_PREFLIGHT     = 8
PROXY_COOLDOWN_SECONDS = 300

CIRCUIT_FAIL_THRESHOLD = 5
CIRCUIT_OPEN_SECONDS   = 1800  # 30 min
CIRCUIT_HALF_OPEN_WAIT = 60    # 1 min antes de testar

# Thresholds por provider (sobrepõe os globais)
CIRCUIT_THRESHOLDS = {
    "scrapingant": {"fail": 1, "open_secs": 3600},  # 1 falha → 1h cooldown
    "zenrows":     {"fail": 3, "open_secs": 1800},
    "scrapingbee": {"fail": 3, "open_secs": 1800},
}

# Patterns que indicam quota MENSAL esgotada → cooldown até próximo mês
EXHAUSTED_PATTERNS = [
    ("zenrows",     ["exhausted the API Credits", "upgrade your subscription"]),
    ("scrapingbee", ["AUTH004", "reached its usage limit"]),
    ("scrapedo",    ["Monthly request limit exceeded"]),
    ("crawlbase",   ["Token is invalid", "exhausted the free tier", "Add Billing"]),
    ("scraperapi",  ["out of API credits"]),
    ("scrapingant", ["insufficient credits", "quota exceeded", "monthly limit",
                     "exhausted the API Credits", "monthly cycle"]),
    # NOTA: "concurrency limit" é temporário — NÃO marca como esgotado mensal
]

# Patterns que indicam erro TEMPORÁRIO → circuit breaker (30 min)
TRANSIENT_PATTERNS = [
    ("scrapingant", ["concurrency limit", "Free user concurrency"]),
    ("zenrows",     ["rate limit", "too many requests", "429"]),
    ("scraperapi",  ["rate limit", "429"]),
]

def _record_provider(provider, success, latency_ms=0, error=None, domain=None):
    """Regista resultado de um pedido — actualiza stats globais e por domínio."""
    d = _provider_data.get(provider)
    if not d: return

    with _stats_lock:
        d["requests_today"] += 1
    if success:
        with _stats_lock:
            d["success_today"] += 1
            d["consecutive_failures"] = 0  # reset circuit breaker
        if d.get("circuit_half_open"):
            # Sucesso no teste — fecha o circuit breaker
            d["circuit_open"] = False
            d["circuit_half_open"] = False
            d["circuit_open_until"] = None
            log.info(f"  ✅ Circuit breaker fechado: {provider}")
    else:
        with _stats_lock:
            d["consecutive_failures"] = d.get("consecutive_failures", 0) + 1
            if error: d["last_error"] = error
        thresholds = CIRCUIT_THRESHOLDS.get(provider, {})
        fail_limit = thresholds.get("fail", CIRCUIT_FAIL_THRESHOLD)
        open_secs  = thresholds.get("open_secs", CIRCUIT_OPEN_SECONDS)
        if d["consecutive_failures"] >= fail_limit and not d.get("circuit_open"):
            until = _dt.now().timestamp() + open_secs
            d["circuit_open"] = True
            d["circuit_open_until"] = until
            d["circuit_half_open"] = False
            pause_min = open_secs // 60
            log.warning(f"  ⚡ Circuit breaker ABERTO: {provider} "
                        f"({d['consecutive_failures']} falhas) — pausa {pause_min}min")

    with _stats_lock:
        if latency_ms: d["latencies"] = (d["latencies"] + [latency_ms])[-100:]
        _proxy_stats[provider] = _proxy_stats.get(provider, 0) + (1 if success else 0)

    # Per-domain stats
    if domain:
        with _stats_lock:
            ds = _domain_stats[domain][provider]
            ds["total"] += 1
            if success: ds["success"] += 1
            if latency_ms: ds["latencies"] = (ds["latencies"] + [latency_ms])[-20:]
            if hasattr(_record_provider, '_last_items'):
                ds["items_total"] = ds.get("items_total", 0) + _record_provider._last_items

def _circuit_allows(provider):
    """Verifica se o circuit breaker permite usar este provider."""
    d = _provider_data.get(provider, {})
    if not d.get("circuit_open"): return True
    until = d.get("circuit_open_until", 0)
    now   = _dt.now().timestamp()
    if now >= until:
        # Passa para half-open: testa 1 pedido
        d["circuit_half_open"] = True
        d["circuit_open"] = False
        log.info(f"  ⚡ Circuit breaker HALF-OPEN: {provider} — a testar 1 pedido")
        return True
    # Ainda aberto
    remaining = int((until - now) / 60)
    log.debug(f"  ⚡ Circuit breaker aberto: {provider} — {remaining}min restantes")
    return False

def _domain_provider_score(domain, provider):
    """Score de um provider para um domínio específico (com confiança)."""
    ds  = _domain_stats[domain][provider]
    total   = ds["total"]
    success = ds["success"]
    lats    = ds["latencies"]
    cfg     = PROVIDER_CONFIG.get(provider, {})

    # Confiança: mais amostras = mais confiança (até 1.0 com 50+ amostras)
    confidence = min(1.0, total / 50)

    if total < 3:
        # Pouca história — usa score global com confiança baixa
        d_global = _provider_data.get(provider, {})
        req_g = d_global.get("requests_today", 0)
        suc_g = d_global.get("success_today", 0)
        sr = (suc_g / req_g) if req_g >= 3 else 0.75
        base_priority = 1 - (cfg.get("priority", 5) / 10)
        return (sr * 0.5 + base_priority * 0.5), 0.1, total

    sr      = success / total
    avg_lat = sum(lats)/len(lats) if lats else 3000
    speed   = min(1.0, 2000 / avg_lat)
    cost    = cfg.get("cost_weight", 0.2)
    headroom = max(0.1, 1 - ds["total"] / max(cfg.get("monthly_limit",1000), 1))
    # Items per request (normalized: 10+ items = perfect)
    items_total = ds.get("items_total", 0)
    ipr   = min(1.0, (items_total / max(total, 1)) / 10)

    score = (sr * 0.40 + ipr * 0.30 + speed * 0.20 - cost * 0.10) * headroom
    return score, confidence, total

def _is_exhausted(provider):
    """Verifica cooldown mensal OU circuit breaker aberto."""
    # Circuit breaker
    if not _circuit_allows(provider):
        return True
    # Cooldown mensal (créditos esgotados)
    d = _provider_data.get(provider, {})
    until = d.get("cooldown_until")
    if not until: return False
    if _date.today() >= until:
        d["cooldown_until"] = None; d["exhausted_at"] = None
        if provider in _proxy_exhausted: del _proxy_exhausted[provider]
        _proxy_cooldown.pop(provider, None)
        _session_exhausted.discard(provider)
        log.info(f"  ✅ {provider} resetou — novo ciclo")
        _save_provider_state()  # actualiza ficheiro após reset
        return False
    return True

_STATE_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "providers_state.json")

def _save_provider_state():
    """Guarda estado dos providers em disco — persiste entre restarts."""
    state = {}
    for prov in _PROVIDERS:
        d = _provider_data.get(prov, {})
        if d.get("cooldown_until"):
            state[prov] = {
                "disabled_until": str(d["cooldown_until"]),
                "reason": d.get("last_error", "unknown"),
                "exhausted_at": d.get("exhausted_at", ""),
            }
    state["_key_index"] = dict(_key_index)  # índice das keys activas
    try:
        with open(_STATE_FILE, "w") as f:
            json.dump(state, f, indent=2, default=str)
    except Exception as e:
        log.debug(f"_save_provider_state: {e}")

def _load_provider_state():
    """Carrega estado dos providers do disco — restaura após restart."""
    if not os.path.exists(_STATE_FILE):
        return
    try:
        with open(_STATE_FILE) as f:
            state = json.load(f)
        # Restaura índice das keys activas (rotação sobrevive a restarts)
        _key_index.update(state.pop("_key_index", {}) or {})
        restored = 0
        for prov, info in state.items():
            until_str = info.get("disabled_until", "")
            if not until_str: continue
            try:
                until = _date.fromisoformat(until_str[:10])
                if _date.today() >= until:
                    continue
                d = _provider_data.get(prov)
                if d:
                    d["cooldown_until"] = until
                    d["exhausted_at"]   = info.get("exhausted_at", "")
                    d["last_error"]     = info.get("reason", "")
                    _proxy_cooldown[prov] = _dt(2099, 1, 1).timestamp()
                    _session_exhausted.add(prov)
                    restored += 1
            except Exception as e:
                log.debug(f"_load_provider_state {prov}: {e}")
        if restored:
            log.info(f"  📂 Estado restaurado: {restored} provider(s) ainda esgotado(s)")
    except Exception as e:
        log.debug(f"_load_provider_state: {e}")

def _mark_exhausted(provider, msg=""):
    # Antes de marcar esgotado, tenta rodar para a próxima key
    if provider in ("scraperapi", "scrapingbee", "scrapedo"):
        if _rotate_key(provider):
            return  # nova key activa — provider continua disponível
    d = _provider_data.get(provider, {})
    if d.get("cooldown_until"): return
    today = _date.today()
    until = _date(today.year + (1 if today.month==12 else 0),
                  today.month % 12 + 1, 1)
    d["exhausted_at"]   = _dt.now().isoformat()
    d["cooldown_until"] = until
    d["last_error"]     = "quota_exceeded"
    _proxy_exhausted[provider] = _dt.now()
    _proxy_cooldown[provider] = _dt(2099, 1, 1).timestamp()
    _session_exhausted.add(provider)  # nunca tentado novamente nesta sessão
    log.warning(f"  🔴 {provider} ESGOTADO — cooldown até {until} | {msg[:60]}")
    _save_provider_state()  # persiste em disco

def provider_status_dict():
    out = {}
    for p in _PROVIDERS:
        d    = _provider_data.get(p, {})
        cfg  = PROVIDER_CONFIG.get(p, {})
        lats = d.get("latencies", [])
        req  = d.get("requests_today", 0)
        suc  = d.get("success_today", 0)
        total_items = sum(ds[p].get("items_total",0) for ds in _domain_stats.values() if p in ds)
        total_reqs  = sum(ds[p].get("total",0)       for ds in _domain_stats.values() if p in ds)
        ipr  = round(total_items / max(total_reqs, 1), 1)
        avg_lat = round(sum(lats)/len(lats)) if lats else 0
        score = round((suc/req*0.6 + min(1,2000/max(avg_lat,1))*0.3 - cfg.get("cost_weight",0.2)*0.1)*100
                      ) if req >= 3 else 0
        out[p] = {
            "requests":             req,
            "success_rate":         round(suc/req*100,1) if req else 0,
            "avg_latency_ms":       avg_lat,
            "items_per_request":    ipr,
            "score":                score,
            "circuit_open":         d.get("circuit_open", False),
            "consecutive_failures": d.get("consecutive_failures", 0),
            "cooldown_until":       str(d.get("cooldown_until","")) or None,
            "last_error":           d.get("last_error"),
            "exhausted":            _is_exhausted(p),
            "enabled":              cfg.get("enabled", True),
            "monthly_limit":        cfg.get("monthly_limit"),
            "cost_weight":          cfg.get("cost_weight"),
        }
    return out

def cache_get(url):
    with _cache_lock:
        if url in _http_cache:
            ts, data = _http_cache[url]
            if time.time() - ts < _http_cache_ttl:
                return data
            del _http_cache[url]
    return None

def cache_set(url, data):
    with _cache_lock:
        if url in _http_cache:
            _http_cache.move_to_end(url)
        _http_cache[url] = (time.time(), data)
        while len(_http_cache) > 200:
            _http_cache.popitem(last=False)

def registar_proxy_resultado(proxy, sucesso, tempo_ms):
    """Regista resultado de um pedido para o proxy manager."""
    _proxy_stats[proxy] = _proxy_stats.get(proxy, 0) + 1
    if sucesso:
        _proxy_success[proxy] = _proxy_success.get(proxy, 0) + 1
    else:
        _proxy_fail[proxy] = _proxy_fail.get(proxy, 0) + 1
        # Cooldown se falhar 3+ vezes seguidas
        falhas = _proxy_fail.get(proxy, 0)
        sucessos = _proxy_success.get(proxy, 0)
        total = _proxy_stats.get(proxy, 1)
        taxa = sucessos / total if total > 0 else 0
        if taxa < 0.3 and total >= 5:
            _proxy_cooldown[proxy] = time.time() + PROXY_COOLDOWN_SECONDS
            log.warning(f"⏸ Proxy {proxy} em cooldown ({taxa*100:.0f}% sucesso)")
    # Tempo médio
    if proxy not in _proxy_time:
        _proxy_time[proxy] = []
    _proxy_time[proxy].append(tempo_ms)
    if len(_proxy_time[proxy]) > 20:
        _proxy_time[proxy] = _proxy_time[proxy][-20:]

def proxied_get(url, render=True, _recursion=0, **kwargs):
    """Rotação automática de providers: ScraperAPI→ZenRows→ScrapingBee→Crawlbase→Scrape.do→direto."""
    if _recursion > 2:
        _r=requests.models.Response(); _r.status_code=0; _r._content=b""; return _r
    global _proxy_counter
    # 0. Tenta direto com stack anti-deteção progressiva
    t0 = time.time()

    # 0a. curl_cffi — impersonates Chrome TLS fingerprint (mais difícil de detectar)
    if CURL_CFFI_AVAILABLE:
        try:
            r0 = cffi_requests.get(url, timeout=8, impersonate="chrome120",
                                   headers=random_headers(), verify=False)
            if r0.status_code == 200 and len(r0.text) > 1000:
                _record_provider("direto", True, int((time.time()-t0)*1000))
                return r0
        except Exception:
            pass

    # 0b. cloudscraper — bypass Cloudflare JS challenges
    if CLOUDSCRAPER_AVAILABLE and _cloudscraper:
        try:
            r0 = _cloudscraper.get(url, timeout=10, headers=random_headers())
            if r0.status_code == 200 and len(r0.text) > 1000:
                _record_provider("direto", True, int((time.time()-t0)*1000))
                return r0
        except Exception:
            pass

    # 0c. requests normal (fallback)
    try:
        r0 = requests.get(url, timeout=8, headers=random_headers(), verify=False)
        if r0.status_code == 200 and len(r0.text) > 1000:
            _record_provider("direto", True, int((time.time()-t0)*1000))
            return r0
    except Exception:
        pass

    # 1. Free proxies (ProxyScrape + Webshare) — só para sites sem JS
    if not render:
        free_r = _try_free_proxy(url, timeout=12)
        if free_r and len(free_r.text) > 1000:
            return free_r

    # 2-8. Proxies pagos por ordem de generosidade
    proxies = []
    # Ordena por prioridade do providers.json (menor = primeiro)
    available = [
        (PROVIDER_CONFIG.get(p,{}).get("priority",99), p)
        for p, key in [
            ("scrapingant",  SCRAPINGANT_KEY),
            ("scraperapi",   SCRAPERAPI_KEY),
            ("zenrows",      ZENROWS_KEY),
            ("crawlbase",    CRAWLBASE_KEY),    # antes da ScrapingBee (melhor taxa)
            ("scrapedo",     SCRAPEDO_KEY),
            ("scrapingbee",  SCRAPINGBEE_KEY),
        ] if key and PROVIDER_CONFIG.get(p,{}).get("enabled",True)
    ]
    available.sort()
    proxies = [p for _, p in available]

    if not proxies:
        _proxy_stats["direto"] = _proxy_stats.get("direto", 0) + 1
        return requests.get(url, headers=random_headers(), timeout=TIMEOUT_DIRETO)

    # Check cache first
    cached = cache_get(url)
    if cached:
        log.debug(f"Cache hit: {url[:60]}")
        return cached

    # Filtra providers esgotados (quota mensal) e em cooldown
    proxies_ativos_now = [p for p in proxies
                          if p not in _session_exhausted
                          and not _is_exhausted(p)
                          and time.time() > _proxy_cooldown.get(p, 0)]
    if not proxies_ativos_now:
        _fb = [p for p in proxies if p not in _session_exhausted and not _is_exhausted(p)]
        if _fb:
            proxies_ativos_now = sorted(_fb, key=lambda p: _proxy_cooldown.get(p,0))[:1]
        else:
            try:
                return requests.get(url, headers=random_headers(), timeout=15, verify=False)
            except:
                _r=requests.models.Response(); _r.status_code=0; _r._content=b""; return _r

    # Score por domínio × provider (com confiança)
    _domain = _urlparse(url).netloc.replace("www.","")

    def provider_score(p):
        score, conf, samples = _domain_provider_score(_domain, p)
        cfg   = PROVIDER_CONFIG.get(p, {})
        prior = max(0, 1 - cfg.get("priority", 10) / 10)
        # Low confidence → fall back to priority; high confidence → use learned score
        return score * conf + prior * (1 - conf)

    proxies_ativos_now.sort(key=provider_score, reverse=True)
    proxy = proxies_ativos_now[0]

    if log.isEnabledFor(logging.DEBUG):
        scores = {}
        for p in proxies_ativos_now:
            s, c, n = _domain_provider_score(_domain, p)
            scores[p] = f"{s:.2f}(c={c:.1f},n={n})"
        log.debug(f"Scores [{_domain}]: {scores}")
    _proxy_counter += 1
    score, conf, n = _domain_provider_score(_domain, proxy)
    log.debug(f"Proxy: {proxy} [{_proxy_counter}] render={render} score={score:.2f}(conf={conf:.1f},n={n})")


    # _session_exhausted: set O(1) — check antes de qualquer trabalho
    if proxy in _session_exhausted:
        r = requests.models.Response(); r.status_code = 0; r._content = b""
        return r

    t0 = time.time()
    try:
        # Re-verifica se proxy ainda está válido (pode ter sido marcado entre tentativas)
        if _is_exhausted(proxy):
            log.debug(f"  {proxy} ficou esgotado — a tentar direto")
            raise Exception(f"{proxy} exhausted during run")
        if proxy == "scraperapi":
            rp = "&render=true&wait=3000" if render else ""
            _key = _get_active_key("scraperapi") or SCRAPERAPI_KEY
            api = f"http://api.scraperapi.com?api_key={_key}&url={requests.utils.quote(url)}{rp}"
            result = requests.get(api, timeout=60, headers=random_headers())

        elif proxy == "zenrows":
            with _provider_semaphores.get("zenrows", threading.Semaphore(3)):
                params = {
                    "url": url, "apikey": ZENROWS_KEY,
                    "js_render": "true" if render else "false",
                    "premium_proxy": "true",
                }
                result = requests.get("https://api.zenrows.com/v1/", params=params, timeout=60)

        elif proxy == "scrapingbee":
            params = {
                "api_key": _get_active_key("scrapingbee") or SCRAPINGBEE_KEY, "url": url,
                "render_js": "true" if render else "false",
                "premium_proxy": "true",
            }
            result = requests.get("https://app.scrapingbee.com/api/v1/", params=params, timeout=60)

        elif proxy == "crawlbase":
            # Crawlbase — suporta JS com &ajax_wait=true
            ajax = "&ajax_wait=true&page_wait=3000" if render else ""
            api = f"https://api.crawlbase.com/?token={CRAWLBASE_KEY}&url={requests.utils.quote(url)}{ajax}"
            result = requests.get(api, timeout=60, headers=random_headers())

        elif proxy == "scrapingant":
            # ScrapingAnt — só usa se for o único provider disponível
            # (free tier: 1 pedido simultâneo, timeout agressivo)
            outros_disponiveis = [p for p in proxies_ativos_now
                                  if p != "scrapingant" and not _is_exhausted(p)]
            if outros_disponiveis:
                # Há outros providers — salta ScrapingAnt para evitar fila
                r = requests.models.Response()
                r.status_code = 503
                r._content = b'{"skip": "others_available"}'
                return r
            # Só chega aqui se ScrapingAnt for o único disponível
            acquired = _provider_semaphores["scrapingant"].acquire(timeout=1)
            if not acquired:
                r = requests.models.Response()
                r.status_code = 503
                r._content = b'{"skip": "busy"}'
                return r
            try:
                params = {
                    "url": url, "x-api-key": SCRAPINGANT_KEY,
                    "browser": "true" if render else "false",
                }
                result = requests.get("https://api.scrapingant.com/v2/general",
                    params=params, timeout=5)
            finally:
                _provider_semaphores["scrapingant"].release()

        elif proxy == "scrapedo":
            # Scrape.do — suporta JS com &render=true
            rp = "&render=true" if render else ""
            _key = _get_active_key("scrapedo") or SCRAPEDO_KEY
            api = f"https://api.scrape.do?token={_key}&url={requests.utils.quote(url)}{rp}"
            result = requests.get(api, timeout=60, headers=random_headers())

        ms = int((time.time()-t0)*1000)
        if hasattr(result,'text') and result.text and len(result.text) < 500:
            msg = result.text[:200]
            # Fallback genérico: mensagens inequívocas de quota marcam o proxy ACTUAL,
            # independentemente do provider a que o padrão está associado
            _generic_quota = ["exhausted the api credits", "upgrade your subscription",
                              "monthly request limit", "out of api credits"]
            if any(g in msg.lower() for g in _generic_quota):
                log.warning(f"    proxy resposta bloqueada ({len(result.text)} chars) [{proxy}]: {msg[:100]}")
                _record_provider(proxy, False, ms, error="quota", domain=_domain)
                _mark_exhausted(proxy, msg)
                _next = [p for p in proxies if p != proxy
                         and p not in _session_exhausted and not _is_exhausted(p)]
                if _next:
                    return proxied_get(url, render=render, _recursion=_recursion+1)
            for _pv, _pts in EXHAUSTED_PATTERNS:
                if _pv == proxy and any(_p.lower() in msg.lower() for _p in _pts):
                    log.warning(f"    proxy resposta bloqueada ({len(result.text)} chars): {msg[:100]}")
                    _record_provider(proxy, False, ms, error="quota", domain=_domain)
                    _mark_exhausted(proxy, msg)
                    _next = [p for p in proxies if p != proxy
                             and p not in _session_exhausted and not _is_exhausted(p)]
                    if _next:
                        return proxied_get(url, render=render, _recursion=_recursion+1)
                    break
        _record_provider(proxy, True, ms, domain=_domain)
        registar_proxy_resultado(proxy, True, ms)
        if result and hasattr(result, 'text') and len(result.text) > 1000:
            cache_set(url, result)
        return result
    except Exception as e:
        _record_provider(proxy, False, int((time.time()-t0)*1000),
                        error=str(e), domain=_domain)
        registar_proxy_resultado(proxy, False, int((time.time()-t0)*1000))
        log.warning(f"Proxy {proxy} falhou: {e} — a tentar direto")

    # Fallback direto
    _proxy_stats["direto"] = _proxy_stats.get("direto", 0) + 1
    try:
        r = requests.get(url, headers=random_headers(), timeout=15)
        if r.status_code != 200:
            log.debug(f"Direto {url[:50]}: HTTP {r.status_code}")
        return r
    except Exception as e:
        log.debug(f"Direto falhou {url[:50]}: {e}")
        r = requests.models.Response()
        r.status_code = 0
        r._content = b""
        return r

def get_proxy_stats():
    """Estatísticas de uso de cada proxy."""
    total = sum(_proxy_stats.values())
    return {k: {"pedidos": v, "pct": round(v/total*100) if total else 0}
            for k, v in _proxy_stats.items() if v > 0}

def proxies_disponiveis():
    """Número de proxies configurados e habilitados."""
    total = 0
    for prov, key in API_KEYS.items():
        if not key: continue
        if not PROVIDER_CONFIG.get(prov, {}).get("enabled", True): continue
        total += 1
    return total

# ============================================================
# BASE DE DADOS
# ============================================================

# Pool de conexões: 42 call-sites abriam uma ligação TCP nova cada um.
# Todos usam "with get_db() as conn:" — o pool.connection() é um context
# manager compatível, portanto a troca não exige mudar nenhum call-site.
try:
    from psycopg_pool import ConnectionPool
    _PSYCOPG_POOL_OK = True
except ImportError:
    _PSYCOPG_POOL_OK = False

_db_pool = None

def get_db():
    global _db_pool
    if not DATABASE_URL: raise RuntimeError("DATABASE_URL não definida.")
    if _PSYCOPG_POOL_OK:
        if _db_pool is None:
            _db_pool = ConnectionPool(DATABASE_URL, min_size=1, max_size=5,
                                      kwargs={"autocommit": False})
        return _db_pool.connection()
    return psycopg.connect(DATABASE_URL)   # fallback sem pool

def init_db():
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                CREATE TABLE IF NOT EXISTS price_history (
                id              SERIAL PRIMARY KEY,
                imovel_id       TEXT NOT NULL,
                preco           INTEGER NOT NULL,
                registado_em    TIMESTAMPTZ DEFAULT NOW()
            );
            CREATE INDEX IF NOT EXISTS idx_ph_imovel ON price_history(imovel_id);

            -- Availability tracking
            ALTER TABLE imoveis
                ADD COLUMN IF NOT EXISTS first_seen  TIMESTAMPTZ DEFAULT NOW(),
                ADD COLUMN IF NOT EXISTS last_seen   TIMESTAMPTZ DEFAULT NOW(),
                ADD COLUMN IF NOT EXISTS dias_online INTEGER DEFAULT 0,
                ADD COLUMN IF NOT EXISTS preco_inicial INTEGER;

            CREATE TABLE IF NOT EXISTS scraper_stats (
                nome            TEXT PRIMARY KEY,
                total_runs      INTEGER DEFAULT 0,
                success_runs    INTEGER DEFAULT 0,
                consecutive_zeros INTEGER DEFAULT 0,
                last_items      INTEGER DEFAULT 0,
                last_run_ts     TIMESTAMPTZ,
                broken          BOOLEAN DEFAULT FALSE,
                broken_since    TIMESTAMPTZ,
                priority_boost  REAL DEFAULT 0.0
            );
            CREATE TABLE IF NOT EXISTS imoveis (
                    id              TEXT PRIMARY KEY,
                    perfil_nome     TEXT NOT NULL,
                    titulo          TEXT,
                    preco           TEXT,
                    preco_valor     INTEGER,
                    area_m2         INTEGER,
                    preco_m2        INTEGER,
                    quartos         INTEGER,
                    ano_construcao  INTEGER,
                    descricao       TEXT,
                    lat             DOUBLE PRECISION,
                    lng             DOUBLE PRECISION,
                    morada          TEXT,
                    link            TEXT,
                    fonte           TEXT,
                    zona            TEXT,
                    imagem_url      TEXT,
                    imagens         TEXT[],
                    score           INTEGER DEFAULT 0,
                    disponivel      BOOLEAN DEFAULT TRUE,
                    estado          TEXT DEFAULT 'novo',
                    nota            TEXT DEFAULT '',
                    criado_em       TIMESTAMP DEFAULT NOW(),
                    atualizado_em   TIMESTAMP DEFAULT NOW(),
                    removido_em     TIMESTAMP,
                    reativado_em    TIMESTAMP,
                    favorito        BOOLEAN DEFAULT FALSE,
                    detalhes_extra  JSONB DEFAULT '{}'
                );
                CREATE TABLE IF NOT EXISTS historico_precos (
                    id           SERIAL PRIMARY KEY,
                    imovel_id    TEXT REFERENCES imoveis(id),
                    preco_antigo TEXT,
                    preco_novo   TEXT,
                    alterado_em  TIMESTAMP DEFAULT NOW()
                );
                CREATE TABLE IF NOT EXISTS scraper_logs (
                    id           SERIAL PRIMARY KEY,
                    fonte        TEXT,
                    perfil_nome  TEXT,
                    total        INTEGER,
                    novos        INTEGER,
                    paginas      INTEGER DEFAULT 1,
                    erros        TEXT,
                    executado_em TIMESTAMP DEFAULT NOW()
                );
                CREATE TABLE IF NOT EXISTS push_subscriptions (
                    id        SERIAL PRIMARY KEY,
                    endpoint  TEXT UNIQUE,
                    p256dh    TEXT,
                    auth      TEXT,
                    criado_em TIMESTAMP DEFAULT NOW()
                );
                CREATE TABLE IF NOT EXISTS mercado_snapshot (
                    id           SERIAL PRIMARY KEY,
                    zona         TEXT,
                    perfil_nome  TEXT,
                    preco_medio  INTEGER,
                    total_ativos INTEGER,
                    data         DATE DEFAULT CURRENT_DATE,
                    UNIQUE(zona, perfil_nome, data)
                );
                CREATE TABLE IF NOT EXISTS visitas (
                    id          SERIAL PRIMARY KEY,
                    imovel_id   TEXT REFERENCES imoveis(id),
                    data_visita DATE NOT NULL,
                    nota        TEXT DEFAULT '',
                    avaliacao   INTEGER CHECK(avaliacao BETWEEN 1 AND 5),
                    criado_em   TIMESTAMP DEFAULT NOW()
                );
                CREATE TABLE IF NOT EXISTS perfis_config (
                    id           SERIAL PRIMARY KEY,
                    nome         TEXT NOT NULL,
                    email        TEXT NOT NULL,
                    ativo        BOOLEAN DEFAULT TRUE,
                    preco_max    INTEGER DEFAULT 200000,
                    quartos_min  INTEGER DEFAULT 2,
                    tipos        TEXT[] DEFAULT ARRAY['apartamentos','moradias-e-vivendas'],
                    zonas        TEXT[] DEFAULT ARRAY['faro','tavira','olhao','vila-real-de-santo-antonio','castro-marim'],
                    criado_em    TIMESTAMP DEFAULT NOW(),
                    atualizado_em TIMESTAMP DEFAULT NOW()
                );
            """)
        conn.commit()
    log.info("Base de dados v4 inicializada.")
    try: migrar_precos_colados()
    except Exception as e: log.warning(f"migração de preços saltada: {e}")

def _load_broken_scrapers():
    """Carrega scrapers marcados como broken da BD."""
    broken = set()
    try:
        with get_db() as conn:
            rows = conn.execute(
                "SELECT nome FROM scraper_stats WHERE broken=TRUE").fetchall()
            broken = {r["nome"] for r in rows}
            if broken:
                log.info(f"  Scrapers em pausa (broken): {broken}")
    except Exception as e:
        log.debug(f"load_broken_scrapers: {e}")
    return broken

_broken_scrapers = set()  # carregado em verificar()

def _update_scraper_stat(nome, items):
    """Actualiza métricas persistentes de um scraper após cada ronda."""
    try:
        with get_db() as conn:
            existing = conn.execute(
                "SELECT * FROM scraper_stats WHERE nome=%s", (nome,)).fetchone()
            if not existing:
                conn.execute("""INSERT INTO scraper_stats
                    (nome,total_runs,success_runs,consecutive_zeros,last_items,last_run_ts)
                    VALUES (%s,1,%s,%s,%s,NOW())""",
                    (nome, 1 if items>0 else 0, 0 if items>0 else 1, items))
            else:
                consec = 0 if items > 0 else (existing["consecutive_zeros"] or 0) + 1
                broken = consec >= 10  # 10 rondas seguidas a 0 = broken
                broken_since = existing["broken_since"]
                if broken and not existing["broken"]:
                    broken_since = "NOW()"
                    log.warning(f"  🔴 {nome} marcado BROKEN ({consec} rondas a zero)")
                elif not broken:
                    broken_since = None
                conn.execute("""UPDATE scraper_stats SET
                    total_runs=total_runs+1,
                    success_runs=success_runs+%s,
                    consecutive_zeros=%s,
                    last_items=%s,
                    last_run_ts=NOW(),
                    broken=%s,
                    broken_since=COALESCE(%s::TIMESTAMPTZ, broken_since)
                    WHERE nome=%s""",
                    (1 if items>0 else 0, consec, items,
                     broken, "NOW()" if broken and not existing["broken"] else None,
                     nome))
    except Exception as e:
        log.debug(f"_update_scraper_stat {nome}: {e}")

def get_perfis_db():
    """Lê perfis da base de dados. Se não houver, usa os do código."""
    try:
        with get_db() as conn:
            with conn.cursor(row_factory=psycopg_rows.dict_row) as cur:
                cur.execute("SELECT * FROM perfis_config WHERE ativo=TRUE ORDER BY id")
                rows = cur.fetchall()
                if rows:
                    return [{
                        "nome": r["nome"], "email": r["email"],
                        "preco_max": r["preco_max"], "quartos_min": r["quartos_min"],
                        "tipos": list(r["tipos"]), "zonas": list(r["zonas"]),
                        "telegram_chat": "",
                        "_db_id": r["id"],
                    } for r in rows]
    except Exception as e:
        log.warning(f"get_perfis_db: {e}")
    return PERFIS  # fallback para perfis do código

def get_todos_perfis_db():
    """Lê todos os perfis (ativos e inativos)."""
    try:
        with get_db() as conn:
            with conn.cursor(row_factory=psycopg_rows.dict_row) as cur:
                cur.execute("SELECT * FROM perfis_config ORDER BY id")
                return [dict(r) for r in cur.fetchall()]
    except Exception as e:
        log.warning(f"get_todos_perfis_db: {e}")
        return []

def criar_perfil_db(nome, email, preco_max, quartos_min, tipos, zonas):
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                INSERT INTO perfis_config (nome, email, preco_max, quartos_min, tipos, zonas)
                VALUES (%s, %s, %s, %s, %s, %s) RETURNING id
            """, (nome, email, preco_max, quartos_min, tipos, zonas))
            new_id = cur.fetchone()[0]
        conn.commit()
    return new_id

def atualizar_perfil_db(pid, nome, email, preco_max, quartos_min, tipos, zonas, ativo):
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                UPDATE perfis_config SET nome=%s, email=%s, preco_max=%s,
                quartos_min=%s, tipos=%s, zonas=%s, ativo=%s, atualizado_em=NOW()
                WHERE id=%s
            """, (nome, email, preco_max, quartos_min, tipos, zonas, ativo, pid))
        conn.commit()

def apagar_perfil_db(pid):
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM perfis_config WHERE id=%s", (pid,))
        conn.commit()

def sincronizar_perfis_iniciais():
    """Se não há perfis na DB, copia os do código."""
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT COUNT(*) FROM perfis_config")
                count = cur.fetchone()[0]
        if count == 0:
            for p in PERFIS:
                criar_perfil_db(p["nome"], p["email"], p["preco_max"],
                    p["quartos_min"], p["tipos"], p["zonas"])
            log.info(f"Perfis iniciais copiados para a DB ({len(PERFIS)})")
    except Exception as e:
        log.warning(f"sincronizar_perfis_iniciais: {e}")

def geocodificar_existentes():
    """Geocodifica imóveis que ainda não têm coordenadas GPS."""
    try:
        with get_db() as conn:
            with conn.cursor(row_factory=psycopg_rows.dict_row) as cur:
                cur.execute("""
                    SELECT id, titulo, zona FROM imoveis
                    WHERE lat IS NULL AND disponivel=TRUE
                    LIMIT 50
                """)
                sem_coords = cur.fetchall()

        if not sem_coords:
            log.info("Geocoding: todos os imóveis já têm coordenadas.")
            return

        log.info(f"A geocodificar {len(sem_coords)} imóveis...")
        atualizados = 0
        for im in sem_coords:
            lat, lng = geocodificar(im["titulo"], im["zona"])
            if lat and lng:
                with get_db() as conn:
                    with conn.cursor() as cur:
                        cur.execute(
                            "UPDATE imoveis SET lat=%s, lng=%s WHERE id=%s",
                            (lat, lng, im["id"]))
                    conn.commit()
                atualizados += 1

        log.info(f"Geocoding concluído: {atualizados}/{len(sem_coords)} atualizados")
    except Exception as e:
        log.error(f"geocodificar_existentes: {e}")

def extrair_preco_valor(p):
    r"""Extrai o PRIMEIRO preço plausível de uma string.

    Antes: re.sub(r"[^\d]","",p) juntava todos os dígitos — um card com
    "930.000 €950.000 €" virava 930_000_950_000 (biliões). Também partia
    decimais: "1.250.000,50 €" -> 125000050.
    Agora: procura grupos com separador de milhares (. ou espaço) e aceita
    o primeiro valor numa gama plausível para imóveis.
    """
    if not p: return None
    s = p.replace("\u00a0", " ").replace("\u202f", " ")   # NBSP / thin space
    for m in re.finditer(r"\d{1,3}(?:\.\d{3})+|\d{1,3}(?: \d{3})+|\d+", s):
        v = int(re.sub(r"\D", "", m.group()))
        if 1000 <= v <= 50_000_000:
            return v
    return None

def extrair_area(t):
    if not t: return None
    m = re.search(r"(\d+)\s*m[²2]", t, re.I)
    return int(m.group(1)) if m else None

def extrair_quartos(t):
    if not t: return None
    m = re.search(r"[Tt](\d)", t)
    if m: return int(m.group(1))
    m = re.search(r"(\d)\s*quarto", t, re.I)
    return int(m.group(1)) if m else None

# Cache de geocoding para não repetir pedidos
_geocoding_cache = {}

def geocodificar(titulo, zona):
    """
    Converte zona/título em coordenadas GPS usando Nominatim (OpenStreetMap).
    Completamente gratuito, sem API key.
    """
    if not zona: return None, None

    # Mapeamento direto de zonas conhecidas (mais rápido e fiável)
    COORDS_ZONAS = {
        "Faro":         (37.0193, -7.9304),
        "Tavira":       (37.1241, -7.6507),
        "Olhão":        (37.0290, -7.8418),
        "VRSA":         (37.1940, -7.4148),
        "Castro Marim": (37.2147, -7.4440),
        "Lagos":        (37.1019, -8.6752),
        "Portimão":     (37.1358, -8.5380),
        "Silves":       (37.1897, -8.4388),
        "Albufeira":    (37.0882, -8.2506),
        "Loulé":        (37.1435, -8.0240),
        "VRSA/Castro Marim": (37.2040, -7.4300),
        "Algarve Sotavento": (37.1000, -7.7000),
        "Algarve":      (37.0902, -8.0902),
    }

    if zona in COORDS_ZONAS:
        lat, lng = COORDS_ZONAS[zona]
        # Adiciona pequena variação aleatória para não sobrepor pins no mesmo ponto
        import random
        lat += random.uniform(-0.02, 0.02)
        lng += random.uniform(-0.02, 0.02)
        return round(lat, 6), round(lng, 6)

    # Tenta geocodificar via Nominatim se zona não está no mapa
    cache_key = zona.lower().strip()  # url+render+provider handled in proxied_get
    if cache_key in _geocoding_cache:
        return _geocoding_cache[cache_key]

    try:
        url = f"https://nominatim.openstreetmap.org/search?q={requests.utils.quote(zona+', Algarve, Portugal')}&format=json&limit=1"
        r = requests.get(url, headers={"User-Agent":"AlgarveMonitor/1.0"}, timeout=5)
        data = r.json()
        if data:
            lat = float(data[0]["lat"])
            lng = float(data[0]["lon"])
            _geocoding_cache[cache_key] = (lat, lng)
            time.sleep(1)  # respeitar rate limit Nominatim
            return lat, lng
    except Exception as e:
        log.debug(f"Geocoding {zona}: {e}")

    _geocoding_cache[cache_key] = (None, None)
    return None, None

# Palavras que aumentam o score
PALAVRAS_POSITIVAS = {
    "piscina": 8, "pool": 8, "vista mar": 10, "sea view": 10,
    "garagem": 4, "garage": 4, "elevador": 3, "lift": 3,
    "renovado": 5, "renovada": 5, "renovated": 5, "novo": 3, "nova": 3,
    "jardim": 3, "garden": 3, "terraço": 3, "terrace": 3,
    "ar condicionado": 2, "ac": 2, "solar": 2, "eficiência": 2,
    "praia": 5, "beach": 5, "centro": 3, "center": 3,
    "moradia": 2, "villa": 2, "quinta": 3,
}

def calcular_score(item, perfil):
    score = 0
    preco   = item.get("preco_valor")
    area    = item.get("area_m2")
    zona    = item.get("zona","")
    quartos = item.get("quartos")
    desc    = (item.get("descricao") or "").lower()
    titulo  = (item.get("titulo") or "").lower()
    texto   = titulo + " " + desc

    # Preço (0-30 pts)
    if preco and preco > 0:
        score += max(0, min(30, int((1-(preco/perfil["preco_max"]))*30)))

    # Preço/m² (0-20 pts)
    if preco and area and area > 0:
        ratio = 1 - min(1, (preco/area - 800) / 2200)
        score += max(0, min(20, int(ratio*20)))

    # Zona (0-15 pts)
    score += int((ZONA_SCORE.get(zona,5)/10)*15)

    # Quartos (0-10 pts)
    if quartos:
        score += min(10, (quartos - perfil["quartos_min"] + 1) * 3)

    # Características positivas (0-20 pts)
    bonus = 0
    for palavra, pts in PALAVRAS_POSITIVAS.items():
        if palavra in texto:
            bonus += pts
    score += min(20, bonus)

    # Baixa de preço recente (+5 pts)
    if item.get("preco_antigo"):
        score += 5

    # Penalização por anúncio antigo (>30 dias)
    criado = item.get("criado_em","")
    if criado:
        try:
            from datetime import datetime
            dias = (datetime.now() - datetime.fromisoformat(criado[:19])).days
            if dias > 30: score -= min(10, dias // 30 * 2)
        except: pass

    return max(0, min(100, score))

def imovel_existe(imovel_id):
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT preco, disponivel FROM imoveis WHERE id=%s", (imovel_id,))
            row = cur.fetchone()
            return (row[0], row[1]) if row else (None, None)

def _int_seguro(v, lo, hi):
    """int dentro de [lo, hi]; fora → None. Protege colunas INTEGER do Postgres:
    a 20/07 um valor colado da página de detalhe deu NumericValueOutOfRange
    em guardar_imovel e matou a ronda inteira."""
    try:
        v = int(v)
    except (TypeError, ValueError):
        return None
    return v if lo <= v <= hi else None

def guardar_imovel(item, perfil_nome, score):
    pv  = _int_seguro(item.get("preco_valor") or extrair_preco_valor(item.get("preco")),
                      1_000, 50_000_000)
    a   = _int_seguro(item.get("area_m2"), 5, 200_000)
    pm2 = _int_seguro(int(pv/a), 1, 1_000_000) if pv and a else None
    quartos = _int_seguro(item.get("quartos"), 0, 20)
    ano     = _int_seguro(item.get("ano_construcao"), 1500, 2100)

    # Geocodifica se não tem coordenadas
    if not item.get("lat") and not item.get("lng"):
        zona = item.get("zona","")
        titulo = item.get("titulo","")
        lat, lng = geocodificar(titulo, zona)
        if lat and lng:
            item["lat"] = lat
            item["lng"] = lng
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                INSERT INTO imoveis
                    (id,perfil_nome,titulo,preco,preco_valor,area_m2,preco_m2,
                     quartos,ano_construcao,descricao,lat,lng,morada,
                     link,fonte,zona,imagem_url,imagens,score,disponivel,atualizado_em,detalhes_extra)
                VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,TRUE,NOW(),%s)
                ON CONFLICT(id) DO UPDATE SET
                    preco=EXCLUDED.preco, preco_valor=EXCLUDED.preco_valor,
                    area_m2=COALESCE(EXCLUDED.area_m2,imoveis.area_m2),
                    preco_m2=COALESCE(EXCLUDED.preco_m2,imoveis.preco_m2),
                    quartos=COALESCE(EXCLUDED.quartos,imoveis.quartos),
                    ano_construcao=COALESCE(EXCLUDED.ano_construcao,imoveis.ano_construcao),
                    descricao=COALESCE(EXCLUDED.descricao,imoveis.descricao),
                    lat=COALESCE(EXCLUDED.lat,imoveis.lat),
                    lng=COALESCE(EXCLUDED.lng,imoveis.lng),
                    morada=COALESCE(EXCLUDED.morada,imoveis.morada),
                    imagens=COALESCE(EXCLUDED.imagens,imoveis.imagens),
                    score=EXCLUDED.score, disponivel=TRUE, atualizado_em=NOW(),
                    detalhes_extra=EXCLUDED.detalhes_extra
            """, (item["id"],perfil_nome,item.get("titulo"),item.get("preco"),pv,a,pm2,
                  quartos,ano,item.get("descricao"),
                  item.get("lat"),item.get("lng"),item.get("morada"),
                  item.get("link"),item.get("fonte"),item.get("zona"),
                  item.get("imagem_url"),item.get("imagens",[]),score,
                  json.dumps(item.get("detalhes_extra",{}))))
        conn.commit()

def marcar_removidos(ids_vistos, perfil_nome, total_encontrados=0, fontes_ok=None):
    """Marca como removidos os imóveis que desapareceram do site de origem.

    fontes_ok: fontes que devolveram ≥1 item NESTA ronda. Um imóvel só é
    marcado removido se a sua fonte respondeu — quando um site inteiro falha
    (ex.: SuperCasa bloqueada com 403), o inventário dele NÃO desaparece do
    mercado; ficava tudo "Removido" e na ronda seguinte "Reativado" (o
    carrossel de 315 removidos / 145 reativados visível no dashboard)."""
    if 0 < total_encontrados < 10:
        log.warning(f"  Só {total_encontrados} items — a saltar marcar_removidos")
        return []
    removidos = []
    saltados_fonte_falhada = 0
    with get_db() as conn:
        with conn.cursor(row_factory=psycopg_rows.dict_row) as cur:
            cur.execute("""
                SELECT id,titulo,link,zona,fonte FROM imoveis
                WHERE perfil_nome=%s AND disponivel=TRUE
                AND atualizado_em < NOW()-INTERVAL '12 hours'
            """, (perfil_nome,))
            for c in cur.fetchall():
                if fontes_ok is not None and c.get("fonte") and c["fonte"] not in fontes_ok:
                    saltados_fonte_falhada += 1
                    continue   # a fonte falhou a ronda — não é o imóvel que saiu
                if c["id"] not in ids_vistos:
                    cur.execute("UPDATE imoveis SET disponivel=FALSE,removido_em=NOW() WHERE id=%s",(c["id"],))
                    removidos.append(dict(c))
        conn.commit()
    if saltados_fonte_falhada:
        log.info(f"  marcar_removidos: {saltados_fonte_falhada} imóveis poupados (fonte falhou a ronda)")
    return removidos

def migrar_precos_colados():
    """Aplica _preco_bonito aos preços já guardados na BD.
    Corre uma vez ao arranque; imóveis novos são normalizados no fazer_item.
    Alvo: 'comprar85.000 €89.000 €' → '85.000 €'."""
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    SELECT id, preco FROM imoveis
                    WHERE preco LIKE 'comprar%' OR preco ~ '€.*€'
                """)
                mudou = 0; total = 0
                updates = []
                for row in cur.fetchall():
                    total += 1
                    novo = _preco_bonito(row[1])
                    if novo != row[1]:
                        updates.append((novo, row[0]))
                        mudou += 1
                if updates:
                    cur.executemany("UPDATE imoveis SET preco=%s WHERE id=%s", updates)
                    conn.commit()
                log.info(f"  ✨ Migração de preços: {mudou}/{total} normalizados")
    except Exception as e:
        log.warning(f"migrar_precos_colados: {e}")

def marcar_reativado(imovel_id):
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("UPDATE imoveis SET disponivel=TRUE,reativado_em=NOW(),removido_em=NULL WHERE id=%s",(imovel_id,))
        conn.commit()

def registar_mudanca_preco(imovel_id, p_ant, p_nov):
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("INSERT INTO historico_precos(imovel_id,preco_antigo,preco_novo) VALUES(%s,%s,%s)",
                        (imovel_id,p_ant,p_nov))
        conn.commit()

def registar_log_scraper(fonte, perfil_nome, total, novos, paginas=1, erros="", tempo_ms=0):
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                INSERT INTO scraper_logs(fonte,perfil_nome,total,novos,paginas,erros)
                VALUES(%s,%s,%s,%s,%s,%s)
            """, (fonte,perfil_nome,total,novos,paginas,erros))
        conn.commit()

def verificar_scrapers_com_falha():
    """Envia email se um scraper falhou 3+ rondas consecutivas."""
    with get_db() as conn:
        with conn.cursor(row_factory=psycopg_rows.dict_row) as cur:
            cur.execute("""
                SELECT fonte, perfil_nome,
                       COUNT(*) FILTER (WHERE erros!='' AND erros IS NOT NULL) as falhas,
                       COUNT(*) FILTER (WHERE total=0) as zero_resultados
                FROM scraper_logs
                WHERE executado_em > NOW()-INTERVAL '24 hours'
                GROUP BY fonte, perfil_nome
                HAVING COUNT(*) FILTER (WHERE erros!='' OR total=0) >= 3
            """)
            scrapers_com_falha = cur.fetchall()
    if scrapers_com_falha:
        linhas = "\n".join([
            f"• {r['fonte']} ({r['perfil_nome']}): {r['falhas']} erros, {r['zero_resultados']} sem resultados"
            for r in scrapers_com_falha])
        assunto = f"⚠️ {len(scrapers_com_falha)} scraper(s) com falhas — Monitor Algarve"
        msg = MIMEMultipart("alternative")
        msg["Subject"] = assunto
        msg["From"] = EMAIL_REMETENTE
        msg["To"]   = EMAIL_DESTINATARIO
        msg.attach(MIMEText(f"<pre>{linhas}</pre>","html"))
        try:
            _html_body = f"<pre>{linhas}</pre>"
            ok, motivo = _send_via_resend(assunto, _html_body, EMAIL_DESTINATARIO)
            if ok: log.info("✉  Email enviado via Resend")
            else: log.error(f"Email alerta falha: {motivo}")
            log.warning(f"⚠️  Alerta de scrapers com falha enviado.")
        except Exception as e:
            log.error(f"Email alerta falha: {e}")

def atualizar_snapshot_mercado(perfil_nome):
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT zona, AVG(preco_valor)::INTEGER, COUNT(*)
                FROM imoveis WHERE perfil_nome=%s AND disponivel=TRUE AND preco_valor IS NOT NULL
                GROUP BY zona
            """, (perfil_nome,))
            for zona, pm, total in cur.fetchall():
                cur.execute("""
                    INSERT INTO mercado_snapshot(zona,perfil_nome,preco_medio,total_ativos)
                    VALUES(%s,%s,%s,%s)
                    ON CONFLICT(zona,perfil_nome,data) DO UPDATE SET
                        preco_medio=EXCLUDED.preco_medio, total_ativos=EXCLUDED.total_ativos
                """, (zona,perfil_nome,pm,total))
        conn.commit()

def get_imoveis(perfil_nome=None, limite=300):
    with get_db() as conn:
        with conn.cursor(row_factory=psycopg_rows.dict_row) as cur:
            q = "SELECT * FROM imoveis"
            p = []
            if perfil_nome: q+=" WHERE perfil_nome=%s"; p.append(perfil_nome)
            q+=" ORDER BY score DESC, criado_em DESC LIMIT %s"; p.append(limite)
            cur.execute(q,p)
            rows = cur.fetchall()
            import datetime as _dtmod
            result = []
            for r in rows:
                d = dict(r)
                # Serializa QUALQUER coluna de data/hora (não só as 4 conhecidas).
                # Colunas novas como first_seen/last_seen ficavam como datetime e
                # rebentavam o jsonify → 500 em /api/imoveis (mas não em /api/stats).
                for k, v in list(d.items()):
                    if isinstance(v, (_dtmod.datetime, _dtmod.date)):
                        d[k] = v.isoformat()
                result.append(d)
            return result

def calcular_saude_scraper(taxa_sucesso, tempo_medio_ms, media_items, ultima_falha_dias):
    """
    Calcula pontuação de saúde de um scraper (0-100).
    taxa_sucesso: 0-100 (%)
    tempo_medio_ms: milliseconds
    media_items: média de items por ronda
    ultima_falha_dias: dias desde a última falha (None = nunca falhou)
    """
    score = 0
    # Taxa de sucesso (0-50 pts)
    score += min(50, taxa_sucesso * 0.5)
    # Tempo médio (0-20 pts) — penaliza > 10s
    if tempo_medio_ms:
        t_score = max(0, 20 - (tempo_medio_ms / 1000))
        score += min(20, t_score)
    else:
        score += 10  # neutral se não há dados
    # Items médios (0-20 pts)
    if media_items:
        score += min(20, media_items * 0.5)
    # Última falha (0-10 pts)
    if ultima_falha_dias is None:
        score += 10  # nunca falhou
    elif ultima_falha_dias > 30:
        score += 8
    elif ultima_falha_dias > 7:
        score += 5
    elif ultima_falha_dias > 1:
        score += 2
    else:
        score += 0  # falhou hoje

    score = max(0, min(100, int(score)))
    if score >= 80: label, emoji = "Excelente", "🟢"
    elif score >= 60: label, emoji = "Bom", "🟢"
    elif score >= 40: label, emoji = "Instável", "🟠"
    else: label, emoji = "Problemático", "🔴"
    return score, label, emoji

def get_stats():
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT COUNT(*) FROM imoveis WHERE disponivel=TRUE"); total=cur.fetchone()[0]
            cur.execute("SELECT COUNT(*) FROM imoveis WHERE criado_em>NOW()-INTERVAL '24 hours'"); ult=cur.fetchone()[0]
            try:
                cur.execute("SELECT COUNT(*) FROM historico_precos"); baixas=cur.fetchone()[0]
            except Exception:
                conn.rollback(); baixas=0
            cur.execute("SELECT COUNT(*) FROM imoveis WHERE disponivel=FALSE"); rem=cur.fetchone()[0]
            cur.execute("SELECT COUNT(*) FROM imoveis WHERE reativado_em IS NOT NULL"); reat=cur.fetchone()[0]
            cur.execute("SELECT COUNT(*) FROM imoveis WHERE favorito=TRUE"); favs=cur.fetchone()[0]
            return {"total":total,"ultimas_24h":ult,"baixas_preco":baixas,
                    "removidos":rem,"reativados":reat,"favoritos":favs}

def get_dados_mercado():
    try:
        with get_db() as conn:
            with conn.cursor(row_factory=psycopg_rows.dict_row) as cur:
                cur.execute("""
                    SELECT zona, AVG(preco_valor)::INTEGER preco_medio,
                           COUNT(*) total, AVG(preco_m2)::INTEGER preco_m2_medio
                    FROM imoveis WHERE disponivel=TRUE AND preco_valor IS NOT NULL
                    GROUP BY zona ORDER BY preco_medio ASC
                """)
                por_zona = [dict(r) for r in cur.fetchall()]
                cur.execute("""
                    SELECT data::TEXT, zona, preco_medio, total_ativos
                    FROM mercado_snapshot WHERE data>CURRENT_DATE-INTERVAL '90 days'
                    ORDER BY data ASC
                """)
                evolucao = [dict(r) for r in cur.fetchall()]
                cur.execute("""
                    SELECT
                        COUNT(*) FILTER (WHERE preco_valor<100000) ate_100k,
                        COUNT(*) FILTER (WHERE preco_valor BETWEEN 100000 AND 150000) ate_150k,
                        COUNT(*) FILTER (WHERE preco_valor BETWEEN 150000 AND 200000) ate_200k,
                        COUNT(*) FILTER (WHERE preco_valor BETWEEN 200000 AND 300000) ate_300k,
                        COUNT(*) FILTER (WHERE preco_valor>300000) acima_300k
                    FROM imoveis WHERE disponivel=TRUE AND preco_valor IS NOT NULL
                """)
                row = cur.fetchone()
                if row:
                    # dict_row: aceder por row[0] dava KeyError: 0 — era o
                    # erro "get_dados_mercado: 0" nos logs
                    dist = {"<100k": row.get("ate_100k") or 0,
                            "100-150k": row.get("ate_150k") or 0,
                            "150-200k": row.get("ate_200k") or 0,
                            "200-300k": row.get("ate_300k") or 0,
                            ">300k":   row.get("acima_300k") or 0}
                else:
                    dist = {"<100k":0,"100-150k":0,"150-200k":0,"200-300k":0,">300k":0}
                return {"por_zona": por_zona, "evolucao": evolucao, "distribuicao": dist}
    except Exception as e:
        log.error(f"get_dados_mercado: {e}")
        return {"por_zona":[], "evolucao":[], "distribuicao":{"<100k":0,"100-150k":0,"150-200k":0,"200-300k":0,">300k":0}}

def get_visitas(imovel_id=None):
    with get_db() as conn:
        with conn.cursor(row_factory=psycopg_rows.dict_row) as cur:
            if imovel_id:
                cur.execute("SELECT * FROM visitas WHERE imovel_id=%s ORDER BY data_visita DESC",(imovel_id,))
            else:
                cur.execute("""
                    SELECT v.*, i.titulo, i.zona, i.preco
                    FROM visitas v JOIN imoveis i ON v.imovel_id=i.id
                    ORDER BY v.data_visita DESC LIMIT 50
                """)
            rows = cur.fetchall()
            result = []
            for r in rows:
                d = dict(r)
                for k in ["data_visita","criado_em"]:
                    if d.get(k): d[k] = str(d[k])
                result.append(d)
            return result

# ============================================================
# EXTRAÇÃO DE DETALHES (entra em cada anúncio)
# ============================================================

def extrair_detalhes_idealista(link):
    """Extrai detalhes completos de um anúncio do Idealista."""
    try:
        r = proxied_get(link); time.sleep(1)
        soup = safe_soup(r.text, "detail")
        if not soup: return {}
        detalhes = {}
        # Área
        for item in soup.select(".details-property_features li, .feature"):
            txt = item.get_text(strip=True)
            if m := re.search(r"(\d{1,6})\s*m[²2]",txt,re.I): detalhes["area_m2"]=_int_seguro(m.group(1), 5, 200_000)
            # Preço da página de detalhe — ancorado no € (o parser genérico
            # apanharia refs/ids; ancorar no símbolo evita falsos positivos)
            if m := re.search(r"(\\d{1,3}(?:[\\.\\s\\u00a0]\\d{3})+|\\d{4,8})\\s*€|€\\s*(\\d{1,3}(?:[\\.\\s\\u00a0]\\d{3})+|\\d{4,8})", txt):
                _pv = extrair_preco_valor((m.group(1) or m.group(2)) + " €")
                if _pv: detalhes["preco"] = f"{_pv:,}".replace(",", ".") + " €"
            if (m := re.search(r"(\d{4})",txt)) and "constru" in txt.lower(): detalhes["ano_construcao"]=_int_seguro(m.group(1), 1500, 2100)
        # Descrição
        desc = soup.select_one(".comment .description, #description")
        if desc: detalhes["descricao"] = desc.get_text(strip=True)[:500]
        # Imagens
        imgs = [img.get("src") or img.get("data-src")
                for img in soup.select(".detail-image-gallery img, .multimedia-slider img")
                if img.get("src") or img.get("data-src")]
        detalhes["imagens"] = imgs[:8]
        # GPS (some listings have it)
        if m := re.search(r'"latitude":\s*([\d.]+).*?"longitude":\s*([\d.]+)', r.text, re.S):
            detalhes["lat"] = float(m.group(1)); detalhes["lng"] = float(m.group(2))
        return detalhes
    except Exception as e:
        log.debug(f"Detalhe Idealista: {e}"); return {}

def extrair_detalhes_imovirtual(link):
    """Extrai detalhes de um anúncio do Imovirtual."""
    try:
        r = proxied_get(link); time.sleep(1)
        soup = safe_soup(r.text, "detail")
        if not soup: return {}
        detalhes = {}
        for item in soup.select("[aria-label], .listing-item-info"):
            txt = item.get_text(strip=True)
            if m := re.search(r"(\d{1,6})\s*m[²2]",txt,re.I): detalhes["area_m2"]=_int_seguro(m.group(1), 5, 200_000)
            # Preço da página de detalhe — ancorado no € (o parser genérico
            # apanharia refs/ids; ancorar no símbolo evita falsos positivos)
            if m := re.search(r"(\\d{1,3}(?:[\\.\\s\\u00a0]\\d{3})+|\\d{4,8})\\s*€|€\\s*(\\d{1,3}(?:[\\.\\s\\u00a0]\\d{3})+|\\d{4,8})", txt):
                _pv = extrair_preco_valor((m.group(1) or m.group(2)) + " €")
                if _pv: detalhes["preco"] = f"{_pv:,}".replace(",", ".") + " €"
        desc = soup.select_one("[data-cy='advert-description']")
        if desc: detalhes["descricao"] = desc.get_text(strip=True)[:500]
        imgs = [img.get("src") for img in soup.select("img[data-cy='gallery-image']") if img.get("src")]
        detalhes["imagens"] = imgs[:8]
        if m := re.search(r'"lat":\s*([\d.]+).*?"lng":\s*([\d.]+)', r.text, re.S):
            detalhes["lat"] = float(m.group(1)); detalhes["lng"] = float(m.group(2))
        return detalhes
    except Exception as e:
        log.debug(f"Detalhe Imovirtual: {e}"); return {}

def extrair_detalhes_generico(link):
    """Extração genérica para imobiliárias locais."""
    try:
        r = proxied_get(link); time.sleep(1)
        soup = safe_soup(r.text, "detail")
        if not soup: return {}
        detalhes = {}
        texto_completo = soup.get_text(" ")
        if m := re.search(r"(\d{2,3})\s*m[²2]", texto_completo, re.I):
            detalhes["area_m2"] = _int_seguro(m.group(1), 5, 200_000)
        if m := re.search(r"(19[5-9]\d|20[0-2]\d)", texto_completo):
            detalhes["ano_construcao"] = _int_seguro(m.group(1), 1500, 2100)
        desc_el = soup.select_one(".description,.descricao,.detail-description,#description,[class*='desc']")
        if desc_el: detalhes["descricao"] = desc_el.get_text(strip=True)[:500]
        imgs = []
        for img in soup.select("img"):
            src = img.get("src") or img.get("data-src","")
            if src and any(ext in src.lower() for ext in [".jpg",".jpeg",".png",".webp"]):
                if not any(x in src.lower() for x in ["logo","icon","avatar","placeholder"]):
                    imgs.append(src)
        detalhes["imagens"] = imgs[:8]
        return detalhes
    except Exception as e:
        log.debug(f"Detalhe genérico: {e}"); return {}

def enriquecer_com_detalhes(item):
    """Entra no anúncio e extrai detalhes completos."""
    link = item.get("link","")
    if "idealista" in link:       extra = extrair_detalhes_idealista(link)
    elif "imovirtual" in link:    extra = extrair_detalhes_imovirtual(link)
    else:                         extra = extrair_detalhes_generico(link)
    # Preço: "N/D" conta como vazio — sem isto, o preço da página de
    # detalhe nunca substituía o N/D dos itens do fallback de links
    if extra.get("preco") and (item.get("preco") or "N/D") in ("N/D", "", None):
        item["preco"] = extra["preco"]
        item["preco_valor"] = extrair_preco_valor(extra["preco"])
    for k,v in extra.items():
        if v and not item.get(k): item[k] = v
    if extra.get("imagens") and not item.get("imagem_url"):
        item["imagem_url"] = extra["imagens"][0]
    return item

# ============================================================
# DEDUPLICAÇÃO
# ============================================================

def extrair_referencia(titulo):
    """Extrai referência do imóvel do título (ex: REF: 12345, T/123456)."""
    if not titulo: return None
    m = re.search(r'(?:ref[:\s#.]+|t/)(\w{4,})', titulo, re.I)
    return m.group(1).upper() if m else None

def similaridade_titulo(t1, t2):
    """Calcula similaridade entre dois títulos (0-1)."""
    if not t1 or not t2: return 0
    t1, t2 = t1.lower().strip(), t2.lower().strip()
    if t1 == t2: return 1.0
    w1 = set(t1.split()); w2 = set(t2.split())
    if not w1 or not w2: return 0
    return len(w1 & w2) / max(len(w1), len(w2))

def preco_similar(p1, p2, tol=0.03):
    if not p1 or not p2: return False
    return abs(p1 - p2) / max(p1, p2) < tol

def area_similar(a1, a2, tol=0.05):
    if not a1 or not a2: return False
    return abs(a1 - a2) / max(a1, a2) < tol

def gerar_chave_dedup(item):
    """
    Gera chave de deduplicação. Usa referência se disponível,
    senão usa zona+preço+quartos+área.
    """
    zona  = (item.get("zona") or "").lower().strip()
    preco = round((item.get("preco_valor") or 0)/1000)*1000
    qts   = item.get("quartos") or 0
    area  = round((item.get("area_m2") or 0)/5)*5

    # Se tem referência, usa como chave primária
    ref = extrair_referencia(item.get("titulo",""))
    if ref and zona:
        return hashlib.md5(f"ref|{zona}|{ref}".encode()).hexdigest()

    # Fallback: zona+preço+quartos+área
    return hashlib.md5(f"{zona}|{preco}|{qts}|{area}".encode()).hexdigest()

def deduplicar(items):
    """Deduplicação melhorada: chave + similaridade de título + preço/área."""
    por_chave = {}; sem_chave = []
    for item in items:
        if not item.get("preco_valor"):
            item["preco_valor"] = extrair_preco_valor(item.get("preco"))
        if not item.get("area_m2"):
            item["area_m2"] = extrair_area(item.get("titulo"))
        if not item.get("quartos"):
            item["quartos"] = extrair_quartos(item.get("titulo"))
        if not item.get("preco_valor"):
            sem_chave.append(item); continue
        chave = gerar_chave_dedup(item)
        if chave not in por_chave:
            por_chave[chave] = item
        else:
            ex = por_chave[chave]
            # Merge fontes
            fontes = ex.get("_fontes", [ex["fonte"]])
            fontes.append(item["fonte"])
            ex["_fontes"] = list(set(fontes))
            ex["fonte"] = ", ".join(sorted(ex["_fontes"]))
            # Manter o que tem mais info
            if len(item.get("titulo","")) > len(ex.get("titulo","")):
                item["_fontes"] = ex["_fontes"]
                item["fonte"] = ex["fonte"]
                por_chave[chave] = item
            elif item.get("imagem_url") and not ex.get("imagem_url"):
                ex["imagem_url"] = item["imagem_url"]

    # Segunda passagem: dedup por similaridade de título + preço similar
    lista = list(por_chave.values())
    duplicados = set()
    for i, a in enumerate(lista):
        if i in duplicados: continue
        for j, b in enumerate(lista[i+1:], i+1):
            if j in duplicados: continue
            if (a.get("zona") == b.get("zona")
                    and preco_similar(a.get("preco_valor"), b.get("preco_valor"))
                    and area_similar(a.get("area_m2"), b.get("area_m2"))
                    and similaridade_titulo(a.get("titulo"), b.get("titulo")) > 0.6):
                duplicados.add(j)
                # Merge fontes
                fontes = set(a.get("_fontes", [a["fonte"]]) + b.get("_fontes", [b["fonte"]]))
                a["_fontes"] = list(fontes)
                a["fonte"] = ", ".join(sorted(fontes))

    resultado = [item for i, item in enumerate(lista) if i not in duplicados] + sem_chave
    removidos = len(items) - len(resultado)
    log.info(f"  Dedup: {len(items)} → {len(resultado)} ({removidos} duplicados removidos)")
    return resultado

# ============================================================
# SELENIUM
# ============================================================

_driver = None

def get_driver():
    """Selenium driver — apenas para fallback. Playwright é o método principal."""
    global _driver
    try:
        from selenium import webdriver
        from selenium.webdriver.chrome.options import Options
        from selenium.webdriver.chrome.service import Service
    except ImportError:
        log.warning("Selenium não instalado — usar Playwright")
        return None
    if _driver is not None:
        try: _ = _driver.title; return _driver
        except: _driver = None
    import glob, subprocess
    opts = Options()
    for a in ["--headless=new","--no-sandbox","--disable-dev-shm-usage",
              "--disable-gpu","--disable-setuid-sandbox","--single-process",
              "--window-size=1920,1080","--disable-blink-features=AutomationControlled",
              "--disable-extensions","--no-first-run","--disable-default-apps"]:
        opts.add_argument(a)
    opts.add_argument(f"user-agent={random.choice(USER_AGENTS)}")
    opts.add_experimental_option("excludeSwitches",["enable-automation"])
    opts.add_experimental_option("useAutomationExtension",False)

    import subprocess
    # Check env vars first (set by Dockerfile)
    chromium_bin = os.getenv("CHROME_BIN")
    chromedriver_bin = os.getenv("CHROMEDRIVER_BIN")

    # Auto-detect if not set
    if not chromium_bin or not os.path.exists(chromium_bin):
        chromium_bin = None
        for path in ["/usr/bin/chromium", "/usr/bin/chromium-browser",
                     "/usr/bin/google-chrome", "/usr/bin/google-chrome-stable"]:
            if os.path.exists(path):
                chromium_bin = path; break
        if not chromium_bin:
            for pattern in ["/nix/store/*/bin/chromium","/nix/store/*/bin/chromium-browser"]:
                matches = glob.glob(pattern)
                if matches: chromium_bin = matches[0]; break
        if not chromium_bin:
            try:
                result = subprocess.run(["which","chromium"], capture_output=True, text=True)
                if result.returncode == 0: chromium_bin = result.stdout.strip()
            except: pass

    if not chromedriver_bin or not os.path.exists(chromedriver_bin):
        chromedriver_bin = None
        for path in ["/usr/bin/chromedriver", "/usr/bin/chromium-driver"]:
            if os.path.exists(path):
                chromedriver_bin = path; break
        if not chromedriver_bin:
            for pattern in ["/nix/store/*/bin/chromedriver"]:
                matches = glob.glob(pattern)
                if matches: chromedriver_bin = matches[0]; break

    if chromium_bin:
        opts.binary_location = chromium_bin
        log.info(f"Chromium: {chromium_bin}")
    else:
        log.warning("Chromium não encontrado!")

    if chromedriver_bin:
        svc = Service(chromedriver_bin)
        log.info(f"ChromeDriver: {chromedriver_bin}")
    else:
        log.warning("ChromeDriver não encontrado, a usar webdriver-manager")

    _driver = webdriver.Chrome(service=svc, options=opts)
    _driver.set_page_load_timeout(30)
    return _driver

def quit_driver():
    global _driver
    if _driver:
        try: _driver.quit()
        except: pass
        _driver=None

def selenium_get(url, wait_sel=None, wait_s=5):
    """Usa Playwright se disponível, senão fallback para requests."""
    if PLAYWRIGHT_AVAILABLE:
        try:
            with sync_playwright() as p:
                browser = p.chromium.launch(
                    headless=True,
                    executable_path="/usr/bin/chromium",
                    args=["--no-sandbox","--disable-dev-shm-usage","--disable-gpu",
                          "--disable-blink-features=AutomationControlled"]
                )
                ctx = browser.new_context(
                    user_agent=random.choice(USER_AGENTS),
                    viewport={"width":1920,"height":1080}
                )
                page = ctx.new_page()
                page.goto(url, wait_until="networkidle", timeout=30000)
                if wait_sel:
                    try: page.wait_for_selector(wait_sel, timeout=wait_s*1000)
                    except: pass
                html = page.content()
                browser.close()
                return html
        except Exception as e:
            log.warning(f"Playwright: {e} — a usar requests")

    # Fallback para proxied_get
    try:
        r = proxied_get(url, render=True)
        return r.text
    except Exception as e:
        log.warning(f"selenium_get fallback: {e}")
        return ""

def com_retry(fn,n=3,w=5):
    for i in range(n):
        try: return fn()
        except Exception as e:
            if i<n-1: log.warning(f"  T{i+1} falhou: {e}. Em {w}s..."); time.sleep(w)
            else: raise

# ============================================================
# HELPERS / SCRAPERS
# ============================================================

MAX_PAGINAS = 2  # max 2 páginas por zona

# ── Títulos-lixo devolvidos pelos sites ────────────────────
# Muitos sites devolvem breadcrumbs ("Faro>Faro>Faro"), contadores
# ("Imóveis encontrados:4835") ou placeholders ("Não Aplicável") no lugar do
# título. Como o URL do anúncio traz um slug descritivo, derivamos daí.
_TITULOS_A_REPARAR = re.compile(
    r"^(n[ãa]o\s+aplic[áa]vel|sem\s+t[íi]tulo|s/\s*t[íi]tulo)$"
    r"|im[óo]veis\s+encontrados"
    r"|^[^>]{1,30}(>[^>]{1,30}){2,4}$"          # breadcrumb: A>B>C (mín. 2 '>')
    r"|^\{\{", re.I)

def _titulo_do_slug(url):
    """Deriva um título legível do slug do URL do anúncio."""
    if not url: return None
    m = re.search(r"/(?:imovel|imoveis|propriedade|detalhes-do-imovel|property)/([a-z0-9\-]{8,})", url, re.I)
    if not m:
        m = re.search(r"/((?:comprar|venda)-[a-z0-9\-]{8,})\.html", url, re.I)
    if not m: return None
    slug = re.sub(r"-[0-9a-f]{6,}(-[0-9a-f]+)*$", "", m.group(1), flags=re.I)  # tira UUID final
    slug = re.sub(r"^(comprar|venda)-", "", slug, flags=re.I)
    t = slug.replace("-", " ").strip()
    if len(t) < 8: return None
    return (t[:1].upper() + t[1:])[:120]

def _melhor_titulo(titulo, link):
    """Se o título for lixo/genérico, tenta reconstruí-lo a partir do URL."""
    t = (titulo or "").strip()
    # Também repara títulos que são o nome da agência ("KW Portugal", "LNHouse"):
    # se o URL tiver slug útil, o imóvel é recuperado; senão fica igual (rejeitado
    # pelo validar como antes — sem regressão).
    e_agencia = t.lower() in TITULOS_GENERICOS
    if not t or len(t) < 5 or e_agencia or _TITULOS_A_REPARAR.search(t):
        return _titulo_do_slug(link) or t or "Sem título"
    return t

def _preco_bonito(preco_raw):
    """Normaliza a string de preço para exibição.
    Casa SAPO devolve "comprar85.000 €89.000 €" (CTA + preço colado + preço
    antigo colado); o dashboard mostrava tudo em bruto. Extrai o 1º valor
    plausível e formata "85.000 €". Preserva "Sob Consulta"/"N/D"/etc.
    Aceita cêntimos opcionais ("1.250.000,50 €") sem os incluir no output."""
    if not preco_raw:
        return "N/D"
    s = str(preco_raw).replace("\u00a0"," ").replace("\u202f"," ")
    _RE = re.compile(
        r"(\d{1,3}(?:[\.\s]\d{3})+(?:[,\.]\d{1,2})?|\d{4,8}(?:[,\.]\d{1,2})?)\s*€"
        r"|€\s*(\d{1,3}(?:[\.\s]\d{3})+(?:[,\.]\d{1,2})?|\d{4,8}(?:[,\.]\d{1,2})?)")
    m = _RE.search(s)
    if m:
        raw = m.group(1) or m.group(2)
        val = re.sub(r"\D","", re.sub(r"[,\.]\d{1,2}$", "", raw))
        try: n = int(val)
        except ValueError: return preco_raw
        if 1000 <= n <= 50_000_000:
            return f"{n:,}".replace(",", ".") + " €"
    return preco_raw   # "Sob Consulta", "N/D", texto sem € — preserva

def fazer_item(link,titulo,preco,fonte,zona,img=None):
    pv=extrair_preco_valor(preco)
    titulo=_melhor_titulo(titulo, link)   # repara breadcrumbs/contadores/placeholders
    preco=_preco_bonito(preco)            # tira "comprar" e preço antigo colado
    return {"id":link,"titulo":titulo,"preco":preco or "N/D",
            "preco_valor":pv,"link":link,"fonte":fonte,"zona":zona,"imagem_url":img,
            "area_m2":extrair_area(titulo),"quartos":extrair_quartos(titulo)}

# Palavras que indicam arrendamento — exclui estes imóveis
# Detecção de arrendamento por PALAVRA INTEIRA (\b) — o antigo match por
# substring com "mes" descartava imóveis reais em MESsines; "renda" apanharia
# "arrendatário" ok mas também nada de Messines. Casos reais nos logs 20/07.
PALAVRAS_ARRENDAMENTO = [   # mantido para retrocompatibilidade de imports
    "arrendar","arrendamento","arrendado","aluguer","alugar","aluga-se",
    "renda","rent","rental","mensal","month","monthly","trespasse",
]
ARRENDAMENTO_RE = re.compile(
    r"\b(arrendar|arrendamento|arrendado|aluguer|alugar|aluga-se|renda|"
    r"rent|rental|rentals|mensal|monthly|trespasse|trespass)\b"
    r"|€\s*/\s*m[eê]s|por\s+m[eê]s|/m[eê]s\b", re.I)
PALAVRAS_VENDA = ["venda","comprar","compra","vende-se","para venda","sale","sell"]

# Domínios que não são imóveis (redes sociais, etc.)
DOMINIOS_EXCLUIR = [
    "linkedin.com","facebook.com","instagram.com","twitter.com",
    "youtube.com","tiktok.com","whatsapp.com","t.me","telegram.me",
    "google.com","maps.google","mailto:","tel:","javascript:",
]

# ── Links institucionais / navegação que NÃO são imóveis ────
# 35% da BD eram links de rodapé, selectores de idioma e páginas
# institucionais. TITULOS_GENERICOS só fazia correspondência exacta.
TITULOS_LIXO_RE = re.compile(r"""
    ^(pt|en|es|fr|de|it|nl)-\s                 # selector de idioma "es- Español"
  | ^\{\{                                       # template Angular por renderizar
  | ^vendido$ | ^reservado$
  | pol[íi]tica\s+(de|privacidade)
  | termos\s+e\s+condi | condi[çc][õo]es\s+(gerais|de\s+utiliza)
  | livro\s+(de\s+)?reclama | canal\s+de\s+den[úu]ncias
  | resolu[çc][ãa]o\s+(alternativa|de\s+lit)
  | ^franchising$ | sistema\s+de\s+franchising | modelo\s+de\s+neg[óo]cio
  | ^quem\s+somos$ | ^sobre\s+n[óo]s$ | ^contact(e-nos|os|s|o\s+geral)?$
  | ^servi[çc]os$ | ^favoritos$ | ^not[íi]cias$ | ^imprensa$
  | ^equipas?$ | ^consultores$ | ^vantagens$ | ^candidatura
  | ^recrutamento$ | ^market\s+centers$ | ^empreendimentos$
  | ^(comprar|vender)$ | ^im[óo]veis$ | ^properties$ | ^visita\s+virtual$
  | ^apresenta[çc][ãa]o$ | ^homepage$ | ^kw\s+luxury$
  | o\s+que\s+procura | ajude-nos\s+a\s+melhorar | responsabilidade\s+social
  | gabinete\s+de\s+imprensa | casos\s+de\s+sucesso | [áa]reas\s+de\s+trabalho
  | ^ser\s+era$ | an[úu]ncio\s+gratuito | acelerador\s+digital
  | garantia\s+era | casa\s+a\s+estrear | casa\s+nova,?\s+vida\s+nova
  | portugal\s+sweet\s+home | at[ée]\s+100%?\s+financiamento
""", re.I | re.X)

URLS_LIXO_RE = re.compile(r"""
    /pol[íi]tica | /politica- | /privacy | /cookies
  | /termos | condicoes-gerais | livroreclamacoes | whistleblower
  | /franchising | /recrutamento | /candidatura
  | /quem-somos | /sobre-nos | /sobre$ | /about
  | /contact(o|os|s|e-nos)?/?$
  | /noticias | /press/?$ | /imprensa
  | /equipas?/?$ | /equipa/?$ | /consultores/?$
  | /marketcenters | /market-centers
  | /vantagens | /o-que-procura | /era-portugal/ | /trabalhar-na-era
  | /visita-virtual | /campanhas/
  | \#$                                         # âncora vazia
  | [?&]page=\d+                                # página de LISTAGEM, não anúncio
""", re.I | re.X)

# Títulos genéricos que indicam que não é um imóvel real
TITULOS_GENERICOS = [
    "kw portugal","era imobiliária","lnhouse","algarvila","villas tavira",
    "casas do sotavento","garvetur","sortami","imocusto","engel","völkers",
    "remax","re/max","century 21","coldwell","keller williams",
    "concelhos","naturezas","sobre nós","contactos","homepage",
    "facebook","instagram","linkedin","twitter","youtube",
]

def validar(item, perfil, _log_descarte=False):
    """Valida imóvel contra filtros do perfil. _log_descarte=True loga o motivo."""
    titulo    = (item.get("titulo") or "").lower().strip()
    preco_str = (item.get("preco") or "").lower()
    link      = (item.get("link") or "").lower()
    pv        = item.get("preco_valor")
    def _rej(motivo):
        if _log_descarte:
            log.info(f"    DESCARTADO [{motivo}] — {item.get('fonte','?')} | {item.get('titulo','')[:40]}")
        return False

    if any(d in link for d in DOMINIOS_EXCLUIR):
        return _rej("domínio excluído")

    if titulo in TITULOS_GENERICOS or len(titulo) < 5:
        return _rej(f"título genérico/curto: '{titulo[:25]}'")

    # Links institucionais / navegação (política, contactos, franchising, ...)
    if TITULOS_LIXO_RE.search(titulo):
        return _rej("título institucional/navegação")
    if URLS_LIXO_RE.search(link):
        return _rej("URL institucional ou página de listagem")

    # Título = nome da fonte: só rejeita se também não houver preço
    # (sites sem título usam o nome da agência; com preço real, o imóvel é válido)
    fonte = (item.get("fonte") or "").lower()
    if (titulo == fonte or titulo.replace(" ","") == fonte.replace(" ","")) and not pv:
        return _rej("título = nome da fonte, sem preço")

    _m = ARRENDAMENTO_RE.search(titulo) or ARRENDAMENTO_RE.search(preco_str)
    if _m:
        return _rej(f"arrendamento: '{_m.group(0)}'")

    if any(x in link for x in ["/arrendar/","/arrendamento/","/alugar/","/rent/"]):
        return _rej("URL de arrendamento")

    # DECISÃO 22/07 (pedido do utilizador): o scraper NÃO filtra por preço.
    # Recolhe TUDO e a aplicação filtra (slider já suporta "sem limite").
    # Rendas continuam bloqueadas pelo ARRENDAMENTO_RE e pelos URLs /arrendar.
    # perfil["preco_max"] continua a ser usado nos URLs de pesquisa dos sites
    # como optimização, mas nunca para descartar um imóvel já extraído.

    return True

def paginar_requests(url_tpl, parse_fn):
    todos=[]; pag=1
    for pag in range(1,MAX_PAGINAS+1):
        url=url_tpl.format(page=pag)
        def _f(u=url):
            r=proxied_get(u); return parse_fn(r.text)
        try:
            items=com_retry(_f)
            if not items: break
            todos.extend(items); time.sleep(random.uniform(1.5,3))
        except Exception as e: log.error(f"Pag {pag}: {e}"); break
    return todos,pag

MAX_PROXY_TIMEOUT = 20  # segundos — evita que um scraper bloqueie 112s

def paginar_scraperapi(url_tpl, parse_fn):
    """Usa rotação de proxies com fallback para Playwright."""
    todos=[]; pag=1
    for pag in range(1,MAX_PAGINAS+1):
        url=url_tpl.format(page=pag)
        def _fetch(u=url):
            r=proxied_get(u, render=True)
            # Detecta e regista respostas bloqueadas
            if r.status_code == 503 and b'"skip"' in r.content:
                return []  # ScrapingAnt busy — silencioso, não é erro
            if len(r.text) < 500:
                msg = r.text[:150]
                # Loga resposta bloqueada (proxy identificado pelo padrão da mensagem)
                log.warning(f"    proxy resposta bloqueada ({len(r.text)} chars): {msg}")
                # Verifica se é quota mensal esgotada
                # NOTA: não atribuir por padrão aqui — o proxied_get já marca o
                # provider correcto (o que serviu a resposta). Match por padrão neste
                # ponto mis-atribuía respostas do ScrapingAnt ao ZenRows.
                if any(g in msg.lower() for g in ["exhausted", "quota", "limit exceeded",
                                                   "out of api credits"]):
                    return []  # resposta de quota — proxied_get já marcou quem devia
                else:
                    # Verifica se é erro temporário (concorrência/rate limit)
                    for prov, patterns in TRANSIENT_PATTERNS:
                        if any(p.lower() in msg.lower() for p in patterns):
                            d = _provider_data.get(prov, {})
                            d["consecutive_failures"] = d.get("consecutive_failures",0) + 1
                            if d["consecutive_failures"] >= 3:
                                d["circuit_open"] = True
                                d["circuit_open_until"] = _dt.now().timestamp() + 300  # 5 min
                                log.warning(f"  ⚡ {prov} em pausa 5min (concorrência)")
                            break
            return parse_fn(r.text)
        try:
            items=com_retry(_fetch)
            log.info(f"    proxy pag {pag}: {len(items)} items")
            if not items: break
            todos.extend(items); time.sleep(random.uniform(1,2))
        except Exception as e:
            log.error(f"  proxy pag {pag}: {e}"); break
    return todos,pag

def paginar_playwright(url_tpl, parse_fn, nome="?"):
    """Usa Playwright para scraping — fallback quando proxies falham."""
    if not PLAYWRIGHT_AVAILABLE: return [], 0
    todos=[]; pag=1
    for pag in range(1, min(3, MAX_PAGINAS)+1):  # max 3 páginas com Playwright
        url = url_tpl.format(page=pag)
        with _playwright_semaphore:
            try:
                with sync_playwright() as p:
                    browser = p.chromium.launch(
                        headless=True, executable_path="/usr/bin/chromium",
                        args=["--no-sandbox","--disable-dev-shm-usage","--disable-gpu",
                              "--disable-blink-features=AutomationControlled"])
                    ctx = browser.new_context(
                        user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
                        viewport={"width":1920,"height":1080}, locale="pt-PT")
                    page = ctx.new_page()
                    page.goto(url, wait_until="domcontentloaded", timeout=20000)
                    page.wait_for_timeout(3000)
                    html = page.content()
                    browser.close()
                items = parse_fn(html)
                log.info(f"    playwright pag {pag}: {len(items)} items")
                if not items: break
                todos.extend(items)
                time.sleep(2)
            except Exception as e:
                log.error(f"  playwright pag {pag} {nome}: {e}")
                break
    return todos, pag

def paginar_selenium(url_tpl,parse_fn):
    todos=[]; pag=1
    for pag in range(1,MAX_PAGINAS+1):
        html=selenium_get(url_tpl.format(page=pag),wait_sel="article,.property,.listing,h2,h3",wait_s=5)
        items=parse_fn(html)
        if not items: break
        todos.extend(items); time.sleep(random.uniform(2,4))
    return todos,pag

def _parse_generic(html,base,fonte,zona):
    soup=safe_soup(html); 
    if not soup: return []
    items=[]
    for sel in [".property","article",".imovel","[class*='property-card']"]:
        found=soup.select(sel)
        for it in found:
            lt=it.select_one("a"); pt=it.select_one(".price,.preco,[class*='price']")
            tt=it.select_one("h2,h3,.title,[class*='title']"); img=it.select_one("img")
            if not lt: continue
            href=lt.get("href","")
            link=href if href.startswith("http") else base.rstrip("/")+"/"+href.lstrip("/")
            if link==base or not href: continue
            items.append(fazer_item(link,
                tt.get_text(strip=True) if tt else fonte,
                pt.get_text(strip=True) if pt else "N/D",
                fonte,zona,
                img.get("src") or img.get("data-src") if img else None))
        if found: break
    return items

def scrape_idealista(perfil):
    res=[]; pags=0
    for ts in perfil["tipos"]:
        tl={"apartamentos":"Apartamento","moradias-e-vivendas":"Moradia"}.get(ts,ts)
        for zs in perfil["zonas"]:
            zl=TODAS_AS_ZONAS.get(zs,zs)
            # Idealista PT URL format: /comprar-casas/{zona}/com-{tipo},t{q}/
            # Idealista PT: com-apartamentos ou com-moradias (sem "e-vivendas")
            tipo_slug = "apartamentos" if "apart" in ts else "moradias"
            filtro_tipo = f"com-{tipo_slug}"
            tpl=(f"https://www.idealista.pt/comprar-casas/{zs}/{filtro_tipo}/"
                 f"?preco-max={perfil['preco_max']}&quartos-min={perfil['quartos_min']}&pagina={{page}}")
            def parse(html,tl=tl,zl=zl):
                its=[]
                soup=safe_soup(html); 
                if not soup: return its
                for it in soup.select("article.item"):
                    lt=it.select_one("a.item-link"); pt=it.select_one(".item-price")
                    tt=it.select_one(".item-title"); img=it.select_one("img")
                    if not lt: continue
                    its.append(fazer_item("https://www.idealista.pt"+lt.get("href",""),
                        tt.get_text(strip=True) if tt else tl,
                        pt.get_text(strip=True) if pt else "N/D",
                        "Idealista",zl,img.get("src") or img.get("data-src") if img else None))
                return its
            items,p=paginar_scraperapi(tpl,parse)
            if not items:
                # Portais grandes bloqueiam sempre Playwright de datacenter — não tentar
                _no_pw = ["idealista","imovirtual","casa.sapo","supercasa","sapo.pt"]
                if not any(x in tpl.lower() for x in _no_pw):
                    log.info(f"    Proxy 0 items — a tentar Playwright ({tl} {zl})...")
                    items,p=paginar_playwright(tpl,parse,f"{tl} {zl}")
            res.extend(items); pags+=p
    return res,pags

def scrape_imovirtual(perfil):
    res=[]; pags=0
    for ts in perfil["tipos"]:
        tl={"apartamentos":"Apartamento","moradias-e-vivendas":"Moradia"}.get(ts,ts)
        tiv="apartamento" if "apart" in ts else "moradia"
        for zs in perfil["zonas"]:
            zl=TODAS_AS_ZONAS.get(zs,zs)
            # Imovirtual URL format (apartamento/moradia before zona)
            tpl=(f"https://www.imovirtual.com/comprar/{tiv}/{zs}/"
                 f"?priceMax={perfil['preco_max']}&roomsMin={perfil['quartos_min']}&nrAdsPerPage=36&page={{page}}")
            def parse(html,tl=tl,zl=zl):
                its=[]
                soup=safe_soup(html); 
                if not soup: return its
                for it in soup.select("article[data-cy='listing-item']"):
                    lt=it.select_one("a"); pt=it.select_one("[data-cy='listing-item-price']")
                    tt=it.select_one("[data-cy='listing-item-title']"); img=it.select_one("img")
                    if not lt: continue
                    its.append(fazer_item("https://www.imovirtual.com"+lt.get("href",""),
                        tt.get_text(strip=True) if tt else tl,
                        pt.get_text(strip=True) if pt else "N/D",
                        "Imovirtual",zl,img.get("src") or img.get("data-src") if img else None))
                return its
            items,p=paginar_scraperapi(tpl,parse)
            if not items:
                # Portais grandes bloqueiam sempre Playwright de datacenter — não tentar
                _no_pw = ["idealista","imovirtual","casa.sapo","supercasa","sapo.pt"]
                if not any(x in tpl.lower() for x in _no_pw):
                    log.info(f"    Proxy 0 items — a tentar Playwright ({tl} {zl})...")
                    items,p=paginar_playwright(tpl,parse,f"{tl} {zl}")
            res.extend(items); pags+=p
    return res,pags

def scrape_casasapo(perfil):
    res=[]; pags=0
    for ts in perfil["tipos"]:
        tl={"apartamentos":"Apartamento","moradias-e-vivendas":"Moradia"}.get(ts,ts)
        tsk="apartamentos" if "apart" in ts else "moradias"
        for zs in perfil["zonas"]:
            zl=TODAS_AS_ZONAS.get(zs,zs)
            tips=",".join([f"T{i}" for i in range(perfil["quartos_min"],7)])
            tpl=(f"https://casa.sapo.pt/comprar-{tsk}/{zs}/"
                 f"?precomax={perfil['preco_max']}&tipologia={tips}&pn={{page}}")
            def parse(html,tl=tl,zl=zl):
                its=[]
                soup=safe_soup(html); 
                if not soup: return its
                for it in soup.select(".property-info-content,.searchResultProperty"):
                    lt=it.select_one("a"); pt=it.select_one(".property-price,.price")
                    tt=it.select_one(".property-title,h2"); img=it.select_one("img")
                    if not lt: continue
                    href=lt.get("href","")
                    link=href if href.startswith("http") else "https://casa.sapo.pt"+href
                    its.append(fazer_item(link,tt.get_text(strip=True) if tt else tl,
                        pt.get_text(strip=True) if pt else "N/D","Casa SAPO",zl,
                        img.get("src") if img else None))
                return its
            items,p=paginar_scraperapi(tpl,parse)
            if not items:
                # Portais grandes bloqueiam sempre Playwright de datacenter — não tentar
                _no_pw = ["idealista","imovirtual","casa.sapo","supercasa","sapo.pt"]
                if not any(x in tpl.lower() for x in _no_pw):
                    log.info(f"    Proxy 0 items — a tentar Playwright ({tl} {zl})...")
                    items,p=paginar_playwright(tpl,parse,f"{tl} {zl}")
            res.extend(items); pags+=p
    return res,pags

def scrape_supercasa(perfil):
    res=[]; pags=0
    for ts in perfil["tipos"]:
        tl={"apartamentos":"Apartamento","moradias-e-vivendas":"Moradia"}.get(ts,ts)
        tsc="apartamentos" if "apart" in ts else "moradias"
        for zs in perfil["zonas"]:
            zl=TODAS_AS_ZONAS.get(zs,zs)
            tips=",".join([f"T{i}" for i in range(perfil["quartos_min"],6)])
            # SuperCasa URL: /comprar-casas/{zona}/com-apartamentos ou com-moradias
            tipo_sc2 = "apartamentos" if "apart" in ts else "moradias"
            tpl=(f"https://supercasa.pt/comprar-casas/{zs}/com-{tipo_sc2}/"
                 f"?preco-max={perfil['preco_max']}&quartos-min={perfil['quartos_min']}&pagina={{page}}")
            def parse(html,tl=tl,zl=zl):
                its=[]
                soup=safe_soup(html); 
                if not soup: return its
                for it in soup.select("[data-id],.property-item,article"):
                    lt=it.select_one("a"); pt=it.select_one(".price,[class*='price']")
                    tt=it.select_one("h2,h3,[class*='title']"); img=it.select_one("img")
                    if not lt: continue
                    href=lt.get("href","")
                    link=href if href.startswith("http") else "https://supercasa.pt"+href
                    its.append(fazer_item(link,tt.get_text(strip=True) if tt else tl,
                        pt.get_text(strip=True) if pt else "N/D","SuperCasa",zl,
                        img.get("src") if img else None))
                return its
            items,p=paginar_scraperapi(tpl,parse)
            if not items:
                # Portais grandes bloqueiam sempre Playwright de datacenter — não tentar
                _no_pw = ["idealista","imovirtual","casa.sapo","supercasa","sapo.pt"]
                if not any(x in tpl.lower() for x in _no_pw):
                    log.info(f"    Proxy 0 items — a tentar Playwright ({tl} {zl})...")
                    items,p=paginar_playwright(tpl,parse,f"{tl} {zl}")
            res.extend(items); pags+=p
    return res,pags

def _api_scrape(url_tpl, base, fonte, zona, extra_sels=None):
    """Usa ScraperAPI com render=true para sites com JavaScript."""
    def parse(html):
        soup = safe_soup(html)
        if not soup: return []
        items = []
        # Seletores genéricos + específicos por fonte
        sels = [".property",".card","article",".imovel",
                ".listing-item","[class*='property-card']",
                "[class*='card-anchor']","[class*='listing']","li[class]"]
        if extra_sels:
            sels = extra_sels + sels
        for sel in sels:
            found = soup.select(sel)
            for it in found:
                lt = it.select_one("a")
                pt = it.select_one("[class*='price'],[class*='preco'],.price,.preco")
                tt = it.select_one("h2,h3,h4,[class*='title'],[class*='name']")
                img = it.select_one("img")
                if not lt or not lt.get("href"): continue
                href = lt.get("href","")
                link = href if href.startswith("http") else base.rstrip("/")+"/"+href.lstrip("/")
                if link == base: continue
                titulo = tt.get_text(strip=True) if tt else fonte
                preco  = pt.get_text(strip=True) if pt else "N/D"
                # Filtrar entradas sem preço nem título útil
                if len(titulo) < 5 and preco == "N/D": continue
                imagem = img.get("src") or img.get("data-src") if img else None
                items.append(fazer_item(link, titulo, preco, fonte, zona, imagem))
            if items: break
        return items

    todos = []; pag = 1
    for pag in range(1, MAX_PAGINAS+1):
        url = url_tpl.format(page=pag)
        def _fetch(u=url):
            r = proxied_get(u, render=True)
            return parse(r.text)
        try:
            items = com_retry(_fetch)
            log.info(f"    api_scrape pag {pag}: {len(items)} items")
            if not items: break
            todos.extend(items)
            time.sleep(random.uniform(1,2))
        except Exception as e:
            log.error(f"  api_scrape pag {pag}: {e}"); break
    return todos, pag

def _sel_scrape(url_tpl,base,fonte,zona):
    """Mantido para compatibilidade — usa _api_scrape."""
    return _api_scrape(url_tpl, base, fonte, zona)

def scrape_casasdosotavento(p):
    res=[]; pags=0
    for zs in p["zonas"]:
        zl=TODAS_AS_ZONAS.get(zs,zs)
        url=(f"https://www.casasdosotavento.pt/imoveis/venda/"
             f"?concelho={zs}&preco_max={p['preco_max']}&quartos_min={p['quartos_min']}&page={{page}}")
        try: its,pg=_api_scrape(url,"https://www.casasdosotavento.pt","Casas do Sotavento",zl); res.extend(its); pags+=pg
        except Exception as e: log.error(f"CasasSotavento/{zs}: {e}")
        time.sleep(random.uniform(2,4))
    return res,pags

def scrape_algarvila(p):
    try: return _api_scrape("https://www.algarvila.com/en-gb/properties?page={page}","https://www.algarvila.com","AlgarVila","Tavira")
    except Exception as e: log.error(f"AlgarVila: {e}"); return [],0

def scrape_villastavira(p):
    try: return _api_scrape("https://www.villastavira.pt/imoveis?page={page}","https://www.villastavira.pt","Villas Tavira","Tavira")
    except Exception as e: log.error(f"VillasTavira: {e}"); return [],0

def scrape_imocusto(p):
    try: return _sel_scrape("https://www.imocusto.pt/imoveis/venda?page={page}","https://www.imocusto.pt","Imocusto","VRSA/Castro Marim")
    except Exception as e: log.error(f"Imocusto: {e}"); return [],0

def scrape_lnhouse(p):
    """LNHouse: SPA — o HTTP só devolvia lnhouse.pt/# (navegação). Playwright."""
    try:
        return scrape_playwright_html("LNHouse", "https://www.lnhouse.pt/venda",
                                      "a[href*='/imovel'], a[href*='/detalhe']",
                                      "VRSA/Castro Marim", p)
    except Exception as e: log.error(f"LNHouse: {e}"); return [],0

def scrape_engelvoelkers(p):
    res=[]; pags=0
    for zs in p["zonas"]:
        zl=TODAS_AS_ZONAS.get(zs,zs)
        url=(f"https://www.engelvoelkers.com/pt/en/search/?q=&pageSize=20&adType=BUY"
             f"&realEstateType=APARTMENT,HOUSE&country=PRT&city={zs}"
             f"&priceMax={p['preco_max']}&roomsMin={p['quartos_min']}&page={{page}}")
        try:
            its,pg=_api_scrape(url,"https://www.engelvoelkers.com","Engel & Völkers",zl,
                extra_sels=["[class*='property-card']","[class*='PropertyCard']","[class*='ev-property']"])
            res.extend(its); pags+=pg
        except Exception as e: log.error(f"E&V/{zs}: {e}")
        time.sleep(random.uniform(1,2))
    return res,pags

def scrape_era(p):
    res=[]; pags=0
    for zs in p["zonas"]:
        zl=TODAS_AS_ZONAS.get(zs,zs)
        url=(f"https://www.era.pt/comprar/imoveis/{zs}/?preco_max={p['preco_max']}"
             f"&quartos_min={p['quartos_min']}&tipologia=apartamento,moradia&page={{page}}")
        try:
            its,pg=_api_scrape(url,"https://www.era.pt","ERA Imobiliária",zl,
                extra_sels=[".card-anchor",".card",".filter--results .card"])
            res.extend(its); pags+=pg
        except Exception as e: log.error(f"ERA/{zs}: {e}")
        time.sleep(random.uniform(1,2))
    return res,pags

def scrape_remax(p):
    res=[]; pags=0
    for zs in p["zonas"]:
        zl=TODAS_AS_ZONAS.get(zs,zs)
        url=(f"https://www.remax.pt/comprar/{zs}/?pricemax={p['preco_max']}"
             f"&rooms={p['quartos_min']}&type=apartamento,moradia&page={{page}}")
        try:
            its,pg=_api_scrape(url,"https://www.remax.pt","RE/MAX",zl,
                extra_sels=["[class*='listing-card']","[class*='property-card']","a[href*='/imoveis/']"])
            res.extend(its); pags+=pg
        except Exception as e: log.error(f"REMAX/{zs}: {e}")
        time.sleep(random.uniform(1,2))
    return res,pags

def scrape_kwportugal(p):
    """KW Portugal: site novo é uma SPA React (classes Tailwind '!items-baseline'
    'shadow-card' — nada semântico). Nenhum a[href*='/imovel'] existe.
    Evidência 21/07: 616KB renderizados, 0 cards. Mas o fallback por keywords
    de URL ainda encontra links reais (88 resultados nessa mesma ronda), então
    deixamos passar por lá sem forçar seletor específico."""
    res=[]; pags=0
    for zs in p["zonas"][:2]:
        zl=TODAS_AS_ZONAS.get(zs,zs)
        url=f"https://www.kwportugal.pt/pt/pesquisa/?localizacao={zs}&tipo=comprar&priceMax={p['preco_max']}"
        try:
            # Seletor 'a' engloba tudo → soup.select devolve todas as âncoras;
            # o fallback por keywords (imovel/property/...) filtra as boas.
            its,pg=scrape_playwright_html("KW Portugal", url, "a[href]", zl, p)
            res.extend(its); pags+=pg
        except Exception as e: log.error(f"KW/{zs}: {e}")
    return res,pags

# ============================================================
# SCRAPER GENÉRICO PARA TODAS AS IMOBILIÁRIAS
# ============================================================

def scrape_generico(nome, urls, perfil, seletores_extra=None):
    """
    Scraper genérico para imobiliárias — tenta ScraperAPI com render=true.
    urls: lista de URLs a verificar (uma por zona/tipo)
    """
    resultados = []
    _pw_fallbacks = 0   # máx. 2 por site para respeitar o timeout de 120s/scraper
    for url in urls:
        def _fetch(u=url):
            r = proxied_get(u, render=True)
            return _api_scrape_html(r.text, u, nome, seletores_extra)
        items = []
        try:
            items = com_retry(_fetch, n=2)
            _raw = len(items)
            items = [i for i in items if validar(i, perfil, _log_descarte=True)]
            if _raw and not items:
                log.info(f"    [DIAG {nome}] {_raw} itens extraídos, todos descartados (motivos acima)")
        except Exception as e:
            log.error(f"  {nome}: {e}")

        # ── Fallback de renderização JS ─────────────────────────────
        # Estes sites dependiam do render=true do ScraperAPI. Com os renderers
        # pagos esgotados (16/07), o HTML chega sem os anúncios (SPA) → 0 cards.
        # Ex.: Inside-Villas dava 24 até 13/07, zero desde que as keys acabaram.
        # O Browserless continua disponível — renderiza uma vez por URL falhado.
        if not items and PLAYWRIGHT_AVAILABLE and _pw_fallbacks < 2:
            _pw_fallbacks += 1
            try:
                with _playwright_semaphore:   # Browserless free: 1 sessão de cada vez
                    html_pw = _pw_open_page(nome, url, "a[href]")
                if html_pw and len(html_pw) > 3000:
                    items = _api_scrape_html(html_pw, url, nome, seletores_extra)
                    items = [i for i in items if validar(i, perfil)]
                    if items:
                        log.info(f"    {nome}: {len(items)} items via render JS (fallback)")
            except Exception as e:
                log.debug(f"  {nome} fallback JS: {e}")

        log.info(f"    {nome}: {len(items)} items de {url[:60]}")
        resultados.extend(items)
        time.sleep(random.uniform(2, 4))
    return resultados, 1

def _apanhar_link(card):
    """Escolhe a âncora com href útil num card.
    Ordem: o próprio card se for <a> (plataforma EGO: Sunpoint/Algarve Unique),
    senão a 1ª âncora interior com href real (ignora '#', javascript:, mailto:
    — as setas de slideshow/favoritos vinham primeiro e o '#' era rejeitado
    pelo URLS_LIXO_RE), senão a âncora-pai que envolve o card."""
    def _util(h):
        h = (h or "").strip().lower()
        return bool(h) and h != "#" and not h.startswith(("javascript", "mailto", "tel:"))
    if card.name == "a" and _util(card.get("href")):
        return card
    for a in card.select("a[href]"):
        if _util(a.get("href")):
            return a
    pai = card.find_parent("a", href=True)
    return pai if (pai is not None and _util(pai.get("href"))) else None

def _api_scrape_html(html, base_url, fonte, extra_sels=None):
    """Extrai imóveis de HTML usando seletores CSS."""
    if not html or len(html) < 200:
        return []
    if "<html" not in html.lower() and "<div" not in html.lower():
        log.debug(f"  {fonte}: resposta inválida (não é HTML)")
        return []
    soup = safe_soup(html)
    if not soup: return []
    items = []
    sels = (extra_sels or []) + [
        "article", ".property", ".card", ".imovel", ".listing-item",
        "[class*='property-card']", "[class*='listing-card']",
        "[class*='result-item']", "[class*='property-item']",
    ]
    zona = "Algarve"
    # Tentar determinar zona pelo URL
    for slug, label in TODAS_AS_ZONAS.items():
        if slug in base_url.lower():
            zona = label; break

    for sel in sels:
        found = soup.select(sel)
        for it in found:
            lt  = _apanhar_link(it)
            pt  = it.select_one("[class*='price'],[class*='preco'],.price,.preco")
            tt  = it.select_one("h1,h2,h3,h4,[class*='title'],[class*='name']")
            img = it.select_one("img[src]")
            if not lt: continue
            href = lt.get("href","")
            link = href if href.startswith("http") else urljoin(base_url, href)
            titulo = tt.get_text(strip=True) if tt else ""   # vazio → slug repara
            preco  = pt.get_text(strip=True) if pt else "N/D"
            imagem = (img.get("src") or img.get("data-src","")) if img else None
            item = fazer_item(link, titulo, preco, fonte, zona, imagem)
            items.append(item)
        if items: break

    # ── Fallback por keywords de URL ─────────────────────────
    # Os 14 sites SPA renderizados pelo Browserless traziam 100-700KB de HTML
    # mas 0 itens: os seletores genéricos de cards não batem com o markup deles.
    # Os links de anúncio, esses, seguem padrões universais — extraímos por aí.
    if not items:
        keywords = ["imovel","imoveis/","property","properties/","propriedade",
                    "detalhes","detalhe/","for-sale","listing","ficha","/ref",
                    "apartamento","moradia","villa"]
        seen = set()
        for a in soup.select("a[href]"):
            href = (a.get("href") or "").strip()
            if not href or len(href) < 15: continue
            if not any(k in href.lower() for k in keywords): continue
            link = href if href.startswith("http") else urljoin(base_url, href)
            if link in seen: continue
            seen.add(link)
            titulo = (a.get("title") or a.get_text(strip=True) or "")[:120]
            item = fazer_item(link, titulo, "N/D", fonte, zona, None)
            items.append(item)
            if len(seen) >= 60: break   # tecto de segurança
        if items:
            log.info(f"    {fonte}: {len(items)} links por keywords de URL (validar filtra)")

    # ── Último recurso: URLs embebidos em payloads JS/JSON ──
    if not items:
        _PAYLOAD_RE = re.compile(
            r"""["'](/(?:imovel|imoveis|property|properties|propriedade|detalhe)/[a-z0-9\-_/]{8,120})["']""", re.I)
        _seen_pl = set()
        for _m in _PAYLOAD_RE.finditer(html):
            _u = urljoin(base_url, _m.group(1))
            if _u in _seen_pl: continue
            _seen_pl.add(_u)
            items.append(fazer_item(_u, "", "N/D", fonte, zona, None))
            if len(_seen_pl) >= 60: break
        if items:
            log.info(f"    {fonte}: {len(items)} itens de payloads JS (regex)")

    # ── DIAG: HTML grande, zero itens → mostrar os padrões de href reais ──
    if not items and len(html) > 20000:
        from collections import Counter
        from urllib.parse import urlparse as _up
        segs = Counter()
        for a in soup.select("a[href]")[:400]:
            h = (a.get("href") or "").strip()
            if not h or len(h) < 2: continue
            path = _up(h).path if h.startswith("http") else h   # absolutos → só o path
            partes = [s for s in path.split("/") if s]
            if partes: segs[partes[0][:25]] += 1
        top = segs.most_common(8)
        log.info(f"    [DIAG {fonte}] 0 itens de {len(html)//1024}KB. Padrões de href: {top}")
    return items

# ── REDES NACIONAIS/INTERNACIONAIS ───────────────────────

def scrape_coldwell(p):
    urls=[f"https://www.coldwellbanker.pt/imoveis?transacao=compra&distrito=faro&quartos_min={p['quartos_min']}"]
    return scrape_generico("Coldwell Banker", urls, p)

def scrape_sothebys(p):
    """Portugal Sotheby's International Realty.

    Domínio correcto: sothebysrealtypt.com (o antigo sothebysrealty.pt tem
    certificado SSL inválido — ERR_CERT_COMMON_NAME_INVALID em 21/07 e daí
    os timeouts crónicos que motivaram a desactivação). URL confirmado em
    produção via pesquisa web: `.../properties/sale/-/faro/lagoa-algarve`."""
    res = []
    for u in ["https://www.sothebysrealtypt.com/properties/sale/-/faro",
              "https://www.sothebysrealtypt.com/properties/sale"]:
        try:
            # DIAG 21/07 revelou hrefs: [('properties', 41), ...] — o path é
            # /properties/ (plural, sem "y"), não /property/. Já dava 7 items
            # via payload-regex mesmo com seletor errado; com o seletor certo
            # deve subir para dezenas.
            its,_ = scrape_playwright_html("Sotheby's", u,
                        "a[href*='/properties/']",
                        "Algarve", p)
            res.extend(its)
            if len(res) >= 5: break
        except Exception as e:
            log.error(f"  Sotheby's: {e}")
    return res, 1

def scrape_iad(p):
    urls=["https://www.iadportugal.pt/comprar",
          f"https://www.iadportugal.pt/comprar/moradia?prix_max={p['preco_max']}"]
    return scrape_generico("IAD Portugal", urls, p)

def scrape_fineandcountry(p):
    urls=[f"https://www.fineandcountry.com/pt/imoveis-para-venda/algarve?max_price={p['preco_max']}"]
    return scrape_generico("Fine & Country", urls, p)

def scrape_century21(p):
    urls=[f"https://www.century21.pt/imoveis/?local=faro&tipo=comprar&preco_max={p['preco_max']}&quartos={p['quartos_min']}"]
    return scrape_generico("Century 21", urls, p)

def scrape_chavanova(p):
    """Chave Nova — Playwright (connection error via requests)."""
    return scrape_playwright_html(
        "Chave Nova",
        "https://www.chavanova.pt/imoveis?distrito=faro&tipo=venda",
        "[class*='property'],[class*='listing'],article,.card",
        "Algarve", p
    )

def scrape_arcada(p):
    urls=[f"https://www.arcada.com.pt/imoveis?zona=algarve&tipo=venda&preco_max={p['preco_max']}"]
    return scrape_generico("Arcada Imobiliária", urls, p)

# ── ALGARVE TODA A REGIÃO ────────────────────────────────

def scrape_villaskey(p):
    """Villas Key (Lagos/Alvor/Lagoa, Sotavento e Barlavento).

    O domínio antigo villaskey.com está expirado (ERR_NAME_NOT_RESOLVED em
    21/07). O actual — confirmado por pesquisa web 21/07 e por URLs
    partilhados pelo utilizador — é villaskeyproperty.com. Inventário
    substancial (156 imóveis no idealista/pro/villaskey, e a listagem raiz
    tem paginação até pag=22)."""
    res = []
    # DIAG 21/07: /imoveis/ e /pt-pt/imoveis devolveram 4,426 chars com
    # hrefs=[] — SPA React que não hidrata sem tempo. Aumentar o wait via
    # o próprio scrape_playwright_html não é possível daqui, então tentar
    # o URL com ?lbl= (parâmetro do site) que pré-renderiza no servidor.
    for u in ["https://www.villaskeyproperty.com/pt-pt/imoveis?lbl=30710",
              "https://www.villaskeyproperty.com/imoveis/?pag=1",
              "https://www.villaskeyproperty.com/imoveis/"]:
        try:
            its,_ = scrape_playwright_html("Villas Key", u,
                        "a[href*='/imovel/'], a[href*='/imoveis/'], "
                        "[class*='property'], [class*='listing'], article",
                        "Algarve", p)
            res.extend(its)
            if len(res) >= 5: break
        except Exception as e:
            log.error(f"  Villas Key: {e}")
    return res, 1

def scrape_dils(p):
    """Dils Portugal — Playwright (HTTP 200 sem conteúdo via proxy)."""
    return scrape_playwright_html(
        "Dils Portugal",
        "https://www.dils.pt/imoveis?tipo=venda&zona=algarve",
        "[class*='property'],[class*='listing'],article,.card",
        "Algarve", p
    )

def scrape_buyme(p):
    urls=[f"https://www.buymeproperty.pt/comprar?preco_max={p['preco_max']}&quartos_min={p['quartos_min']}"]
    return scrape_generico("BuyMe Property", urls, p)

def scrape_algarveproperty(p):
    urls=[f"https://www.algarveproperty.com/properties-for-sale?max_price={p['preco_max']}&bedrooms={p['quartos_min']}"]
    return scrape_generico("Algarve Property", urls, p)

def scrape_nurisimo(p):
    urls=["https://www.nurisimo.com/properties"]
    return scrape_generico("Nurisimo", urls, p)

def scrape_goldenproperties(p):
    urls=["https://www.goldenproperties.pt/imoveis?tipo=venda"]
    return scrape_generico("Golden Properties", urls, p)

def scrape_algarverealestate(p):
    urls=[f"https://www.algarverealestate.com/properties-for-sale?max_price={p['preco_max']}&bedrooms={p['quartos_min']}"]
    return scrape_generico("Algarve Real Estate", urls, p)

def scrape_espacosalgarve(p):
    urls=[f"https://www.espacos-algarve.com/comprar?preco_max={p['preco_max']}"]
    return scrape_generico("Espaços Algarve", urls, p)

def scrape_redereal(p):
    urls=[f"https://www.redereal.com/imoveis?tipo=venda&zona=algarve&preco_max={p['preco_max']}"]
    return scrape_generico("Rede Real", urls, p)

def scrape_vaprealestate(p):
    urls=[f"https://www.vaprealestate.com/properties?max_price={p['preco_max']}&bedrooms={p['quartos_min']}"]
    return scrape_generico("VAP Real Estate", urls, p)

def scrape_tripalgarve(p):
    urls=["https://tripalgarve.com/properties"]
    return scrape_generico("Tripalgarve", urls, p)

def scrape_algarvedream(p):
    """Algarve Dream Property — Playwright direto para /imoveis."""
    return scrape_playwright_html(
        "Algarve Dream Property",
        "https://www.algarvedreamproperty.com/imoveis",
        "[class*='property'],[class*='listing'],article",
        "Algarve", p
    )

# ── BARLAVENTO ───────────────────────────────────────────

def scrape_mimosa(p):
    """Mimosa (Lagos): DIAG 21/07 mostrou o path real 'imoveis-mimosaproperties'
    entre os hrefs mais frequentes. Os URLs antes usados não existiam."""
    # DIAG confirmou que 'imoveis-mimosaproperties' aparece nos hrefs (menu)
    res=[]
    for u in ["https://www.mimosaproperties.com/imoveis-mimosaproperties",
              "https://www.mimosaproperties.com/imoveis"]:
        its,_ = scrape_playwright_html("Mimosa Properties", u,
                    "a[href*='/imovel/'], a[href*='/property/'], a[href*='/properties/']",
                    "Barlavento", p)
        res.extend(its)
        if len(res) >= 5: break
    return res, 1


def scrape_vernon(p):
    urls=[f"https://www.vernonalgarve.com/for-sale?max_price={p['preco_max']}&bedrooms={p['quartos_min']}"]
    return scrape_generico("Vernon Algarve", urls, p)

def scrape_a1algarve(p):
    urls=[f"https://www.a1-algarve.com/properties?max_price={p['preco_max']}&bedrooms={p['quartos_min']}"]
    return scrape_generico("A1 Algarve", urls, p)

# ── SCRAPERS COM PLAYWRIGHT (sites que bloqueiam requests mas não browser) ──

# Semáforo para limitar Playwright a 2 instâncias simultâneas
_playwright_semaphore = threading.Semaphore(1)  # max 1 browser simultâneo

def _pw_open_page_browserless(nome, url, sel):
    """Usa Browserless.io (browser remoto) — zero RAM no Railway."""
    if not BROWSERLESS_TOKEN:
        return ""
    try:
        from playwright.sync_api import sync_playwright as _spw_bl
        with _spw_bl() as p:
            browser = p.chromium.connect_over_cdp(
                f"wss://chrome.browserless.io?token={BROWSERLESS_TOKEN}")
            ctx = browser.new_context(
                user_agent="Mozilla/5.0 Chrome/120",
                viewport={"width":1920,"height":1080}, locale="pt-PT")
            try:
                page = ctx.new_page()
                page.goto(url, wait_until="domcontentloaded", timeout=60000)
                try: page.wait_for_selector(sel, timeout=8000)
                except: pass
                page.wait_for_timeout(2000)
                html = page.content()
                log.info(f"    {nome} (browserless): {len(html):,} chars")
                return html
            finally:
                try: ctx.close()
                except: pass
                browser.close()
    except Exception as e:
        log.error(f"  {nome} browserless: {e}")
        return ""

def _pw_open_page_sync(nome, url, sel):
    """Versão síncrona pura de _pw_open_page — corre em thread sem asyncio."""
    try:
        # Usa playwright (não patchright) para evitar conflito asyncio
        from playwright.sync_api import sync_playwright as _spw_clean
    except ImportError:
        return ""
    try:
        with _spw_clean() as _p:
            b = _p.chromium.launch(headless=True, executable_path="/usr/bin/chromium",
                args=["--no-sandbox","--disable-dev-shm-usage","--disable-gpu",
                      "--single-process","--disable-blink-features=AutomationControlled"])
            ctx = b.new_context(user_agent="Mozilla/5.0 Chrome/120",
                viewport={"width":1920,"height":1080}, locale="pt-PT")
            try:
                page = ctx.new_page()
                page.goto(url, wait_until="domcontentloaded", timeout=60000)
                try: page.wait_for_selector(sel, timeout=8000)
                except: pass
                page.wait_for_timeout(2000)
                # Scroll adaptativo: EGO e outros carregam cards por lazy-load;
                # 3 passos fixos deixavam só o 1º lote (12). Continua enquanto
                # o nº de elementos do seletor crescer (máx 8 ciclos).
                try:
                    _n_ant = -1
                    for _ci in range(8):
                        page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
                        page.wait_for_timeout(900)
                        try: _n = page.locator(sel).count()
                        except Exception: _n = -2
                        if _n <= _n_ant: break
                        _n_ant = _n
                except Exception: pass
                html = page.content()
                log.info(f"    {nome}: {len(html):,} chars HTML")
                return html
            finally:
                try: ctx.close()
                except: pass
            b.close()
    except Exception as e:
        log.error(f"  {nome} _pw_open_page_sync: {e}")
        return ""

def _pw_open_page(nome, url, sel):
    """Abre página: browser partilhado → Browserless → thread limpa.
    O browser partilhado tem prioridade: dentro do thread PW já há um
    sync_playwright activo, e criar outro na mesma thread (Browserless)
    dá 'Sync API inside asyncio loop'."""
    _b = getattr(scrape_playwright_html, '_shared_browser', None)
    # 1. Browserless SÓ quando não há browser partilhado activo
    if BROWSERLESS_TOKEN and not (_b and _b.is_connected()):
        html = _pw_open_page_browserless(nome, url, sel)
        if html and len(html) > 3000:
            return html
    import asyncio
    # Detecta se estamos dentro de um asyncio loop (Gunicorn gthread)
    # Se sim, usa thread separada para correr sync_playwright
    try:
        loop = asyncio.get_running_loop()
        in_asyncio = loop is not None
    except RuntimeError:
        in_asyncio = False

    if in_asyncio:
        # Corre em thread separada para evitar conflito com asyncio
        import concurrent.futures
        with concurrent.futures.ThreadPoolExecutor(max_workers=1) as ex:
            future = ex.submit(_pw_open_page_sync, nome, url, sel)
            try:
                return future.result(timeout=90)
            except Exception as e:
                log.error(f"  {nome} PW thread: {e}")
                return ""

    _b = getattr(scrape_playwright_html, '_shared_browser', None)
    ctx = None
    try:
        if _b and _b.is_connected():
            ctx = _b.new_context(
                user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120",
                viewport={"width":1920,"height":1080}, locale="pt-PT")
            page = ctx.new_page()
        else:
            # Fallback: launch temporário (não deve acontecer em regime normal)
            log.warning(f"  {nome}: browser partilhado não disponível — a criar temporário")
            raise RuntimeError("no_shared_browser")
        page.goto(url, wait_until="domcontentloaded", timeout=60000)
        try: page.wait_for_selector(sel, timeout=8000)
        except: pass
        page.wait_for_timeout(2000)
        # Scroll adaptativo (lazy-load) — igual ao branch partilhado
        try:
            _n_ant = -1
            for _ci in range(8):
                page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
                page.wait_for_timeout(900)
                try: _n = page.locator(sel).count()
                except Exception: _n = -2
                if _n <= _n_ant: break
                _n_ant = _n
        except Exception: pass
        html = page.content()
        log.info(f"    {nome}: {len(html):,} chars HTML")
        if len(html) < 5000:
            log.warning(f"    {nome}: HTML pequeno ({len(html)} chars) — SPA?")
        return html
    except RuntimeError:
        # Fallback com browser próprio
        with sync_playwright() as _p:
            _b2 = _p.chromium.launch(headless=True, executable_path="/usr/bin/chromium",
                args=["--no-sandbox","--disable-dev-shm-usage","--disable-gpu","--single-process"])
            _ctx = _b2.new_context(user_agent="Mozilla/5.0 Chrome/120", locale="pt-PT")
            _pg  = _ctx.new_page()
            _pg.goto(url, wait_until="domcontentloaded", timeout=60000)
            _pg.wait_for_timeout(3000)
            html = _pg.content()
            _ctx.close(); _b2.close()
        return html
    except Exception as e:
        log.error(f"  {nome} _pw_open_page: {e}")
        return ""
    finally:
        if ctx:
            try: ctx.close()
            except: pass

def scrape_playwright_html(nome, url, sel, zona, perfil):
    """Scraper Playwright — usa browser partilhado (1 launch por ronda)."""
    if not PLAYWRIGHT_AVAILABLE:
        log.warning(f"{nome}: Playwright não disponível")
        return [], 0
    items = []
    with _playwright_semaphore:
        try:
            html = _pw_open_page(nome, url, sel)
            soup = safe_soup(html, nome)
            if not soup:
                return [], 0
            cards = soup.select(sel)
            log.info(f"    Playwright {nome}: {len(cards)} cards")
            if not cards:
                _classes = sorted({c for el in soup.select("[class]")[:200]
                                   for c in (el.get("class") or [])
                                   if any(k in c.lower() for k in ("prop","card","list","item","result"))})[:15]
                from collections import Counter as _Ct
                from urllib.parse import urlparse as _up2
                _segs = _Ct()
                for _a in soup.select("a[href]")[:400]:
                    _h = (_a.get("href") or "").strip()
                    if not _h or len(_h) < 2: continue
                    _pt = _up2(_h).path if _h.startswith("http") else _h
                    _ps = [s for s in _pt.split("/") if s]
                    if _ps: _segs[_ps[0][:25]] += 1
                log.info(f"    [DIAG {nome}] seletor '{sel}' → 0. Classes: {_classes} | hrefs: {_segs.most_common(6)}")
            
            # Primary: extract from cards with links
            found_from_cards = 0
            _debug_card = cards[0] if cards else None
            for card in cards:
                # Link pode estar dentro do card, ser o card, ou envolvê-lo.
                # Sunpoint/Algarve Dream: 12-136 cards, 0 itens, por só
                # procurarem <a> cá dentro.
                lt = _apanhar_link(card)
                if not lt: continue
                href = lt.get("href","")
                # Skip social media and non-property links
                if any(x in href.lower() for x in ["facebook","instagram","linkedin","twitter","youtube","mailto","tel:"]):
                    continue
                link = href if href.startswith("http") else urljoin(url, href)
                pt  = card.select_one("[class*='price'],[class*='preco'],[class*='valor'],[class*='Price']")
                tt  = card.select_one("h1,h2,h3,h4,[class*='title'],[class*='nome'],[class*='Title']")
                img = card.select_one("img[src]")
                # NÃO usar `nome` (a fonte) como fallback: validar() rejeita
                # titulo==fonte. Deixar vazio -> _melhor_titulo() deriva do slug do URL.
                # EGO (Sunpoint/A.Unique) traz o título completo no atributo
                # title do anchor — usar antes de cair no slug
                _t_el = tt.get_text(strip=True) if tt else ""
                if _t_el.lower() in ("usado","novo","reservado","vendido","em construção","new","used"):
                    _t_el = ""   # badge de estado apanhado por [class*='title'] — não é título
                titulo = _t_el \
                         or (lt.get("title") or "").strip() \
                         or (card.get("title") or "").strip()
                preco  = pt.get_text(strip=True) if pt else "N/D"
                imagem = img.get("src") or img.get("data-src","") if img else None
                item = fazer_item(link, titulo, preco, nome, zona, imagem)
                if validar(item, perfil):
                    items.append(item)
                    found_from_cards += 1
            # ── Último recurso: URLs embebidos em payloads JS/JSON ──
            # Sites como o Tripalgarve têm 0 âncoras <a href> (navegação por
            # onclick), mas os slugs dos anúncios estão no HTML em strings.
            if not items:
                _PAYLOAD_RE = re.compile(
                    r"""["'](/(?:imovel|imoveis|property|properties|propriedade|detalhe)/[a-z0-9\-_/]{8,120})["']""", re.I)
                _seen_pl = set()
                for _m in _PAYLOAD_RE.finditer(html):
                    _u = urljoin(url, _m.group(1))
                    if _u in _seen_pl: continue
                    _seen_pl.add(_u)
                    _it = fazer_item(_u, "", "N/D", nome, zona, None)
                    if validar(_it, perfil):
                        items.append(_it)
                    if len(_seen_pl) >= 60: break
                if items:
                    log.info(f"    {nome}: {len(items)} itens de payloads JS (regex)")

            if cards and found_from_cards == 0 and _debug_card is not None:
                # Autópsia do card[0]: repete os passos e loga onde morreu
                _c = _debug_card
                _lt = _apanhar_link(_c)
                if _lt is None:
                    log.info(f"    [DIAG {nome}] card[0] morreu em: _apanhar_link=None | {str(_c)[:250]}")
                else:
                    _h = (_lt.get("href") or "")
                    if any(x in _h.lower() for x in ["facebook","instagram","linkedin","twitter","youtube","mailto","tel:"]):
                        log.info(f"    [DIAG {nome}] card[0] morreu em: filtro social | href={_h[:80]}")
                    else:
                        _l = _h if _h.startswith("http") else urljoin(url, _h)
                        _tt = _c.select_one("h1,h2,h3,h4,[class*='title'],[class*='nome'],[class*='Title']")
                        _tit = (_tt.get_text(strip=True) if _tt else "") or (_lt.get("title") or "").strip()
                        # PREÇO REAL do card — a versão com "N/D" mascarava a
                        # verdadeira causa (Sunpoint: 12 destaques todos >250k)
                        _pt = _c.select_one("[class*='price'],[class*='preco'],[class*='valor'],[class*='Price']")
                        _pr = _pt.get_text(strip=True) if _pt else "N/D"
                        _it = fazer_item(_l, _tit, _pr, nome, zona, None)
                        _v  = validar(_it, perfil, _log_descarte=True)   # loga DESCARTADO [motivo]
                        log.info(f"    [DIAG {nome}] card[0]: validar={_v} preco='{_pr[:20]}' "
                                 f"titulo='{_it['titulo'][:40]}' link={_l[:65]}")

            # Fallback: if no items found from cards, try link-based extraction
            if not found_from_cards:
                keywords = ["imovel","property","detalhes","for-sale","venda","casa","villa","apartamento","moradia"]
                seen_links = set()
                for a in soup.select("a[href]"):
                    href = a.get("href","")
                    if not href or not any(k in href.lower() for k in keywords): continue
                    if len(href) < 15: continue
                    link = href if href.startswith("http") else urljoin(url, href)
                    if link in seen_links: continue
                    seen_links.add(link)
                    # Find parent container for price/title
                    parent = a.parent or a
                    for _ in range(3):  # go up max 3 levels
                        if parent and len(parent.get_text(strip=True)) > 20:
                            break
                        if parent: parent = parent.parent
                    pt = parent.select_one("[class*='price'],[class*='preco'],[class*='valor']") if parent else None
                    tt = parent.select_one("h1,h2,h3,h4,[class*='title']") if parent else None
                    # titulo=nome era rejeitado por validar (titulo==fonte);
                    # vazio → _melhor_titulo deriva do slug do URL
                    titulo = tt.get_text(strip=True) if tt else ""
                    preco  = pt.get_text(strip=True) if pt else "N/D"
                    item = fazer_item(link, titulo, preco, nome, zona, None)
                    if validar(item, perfil): items.append(item)
        except Exception as e:
            log.error(f"  Playwright {nome}: {e}")
    return items, 1

def scrape_boto(p):
    return scrape_playwright_html(
        "Boto Properties",
        "https://www.botoproperties.com/imoveis",
        "a[href*='/imovel/']", "Barlavento", p
    )

def scrape_sunpoint(p):
    # EGO: os destaques do site são de luxo (>1M€). Aplica o teto na origem
    # via query string suportada pela plataforma. Evidência 21/07: sem filtro
    # os 12 cards vinham a 1M+ e o validar rejeitava tudo.
    # A EGO tem 2 páginas: /propriedades (destaques) e /imoveis-<slug> (listagem
    # paginada). Testar ambos os slugs prováveis; o validar filtra por preço.
    # Não uso query-params de preço porque não confirmei o schema da EGO
    # (?p_max e ?vmax podem ser ignorados pela plataforma).
    res=[]
    for u in ["https://www.sunpoint.pt/imoveis-sunpoint",
              "https://www.sunpoint.pt/imoveis-sunpointproperties",
              "https://www.sunpoint.pt/propriedades"]:
        its,_=scrape_playwright_html("Sunpoint Properties", u,
                                     "a[href*='/imovel/']", "Barlavento", p)
        res.extend(its)
        if len(res) >= 5: break
    return res,1

def scrape_algarveuniqueproperties(p):
    return scrape_playwright_html(
        "Algarve Unique Properties",
        "https://www.algarveuniqueproperties.com/imoveis",
        "a[href*='/imovel/']", "Algarve", p
    )

def scrape_garvetur(p):
    return scrape_playwright_html(
        "Garvetur",
        "https://www.garveturproperties.com/",
        "a[href*='/detalhes-do-imovel/'],a[href*='/imovel/']", "Algarve", p
    )

def scrape_barraprime(p):
    return scrape_playwright_html(
        "Barra Prime",
        "https://www.barraprime.pt/imoveis-para-venda",
        "li[class]", "Triângulo Dourado", p
    )

def scrape_mimosaproperties(p):
    return scrape_playwright_html(
        "Mimosa Properties",
        "https://www.mimosaproperties.com/procuro-imovel-mimosaproperties",
        "a[href*='/imovel/']", "Barlavento", p
    )

def scrape_casasdobarlavento(p):
    """Casas do Barlavento (Lagos) — adicionada 22/07 quando o utilizador
    pediu recolha SEM filtro de valor (a app filtra). URLs candidatos;
    o DIAG afina na 1ª ronda."""
    return scrape_generico("Casas do Barlavento",
        ["https://www.casasdobarlavento.pt/imoveis/",
         "https://www.casasdobarlavento.pt/properties/"], p)

def scrape_marcela(p):
    """Marcela Properties (Barlavento histórico) — 22/07."""
    return scrape_generico("Marcela Properties",
        ["https://www.marcelaproperties.com/properties/",
         "https://www.marcelaproperties.com/imoveis/"], p)

def scrape_lagoshomes(p):
    """Lagos Homes (Lagos/Praia da Luz) — 22/07."""
    return scrape_generico("Lagos Homes",
        ["https://www.lagoshomes.com/properties/",
         "https://www.lagoshomes.com/for-sale/"], p)

def scrape_mapro(p):
    """Mapro Real Estate (Quinta do Lago/Vale do Lobo — luxo) — 22/07.
    Entra agora que não há filtro de valor no scraper."""
    return scrape_generico("Mapro Real Estate",
        ["https://www.maprorealestate.com/properties/",
         "https://www.maprorealestate.com/for-sale/"], p)

def scrape_vipalgarve(p):
    """VIP Algarve Property (Quarteira/Vilamoura) — 22/07."""
    return scrape_generico("VIP Algarve",
        ["https://www.vipalgarveproperty.com/properties/",
         "https://www.vipalgarveproperty.com/imoveis/"], p)

def scrape_landhouses(p):
    """Land & Houses Algarve (= Yellow Homes, mesma empresa) — Tavira.

    Verificado por pesquisa 22/07: site ASP.NET clássico (server-rendered),
    detalhes em /<slug>-<id>.aspx, listagens por concelho:
    /Eastern_Algarve/<Localidade>/All_Property_Types/price-0.aspx
    100% Sotavento (Eastern Algarve). NOTA: yellowhomes.com é a outra marca
    da mesma agência — um scraper cobre as duas; não adicionar em separado."""
    urls = [
        "https://www.landandhousesalgarve.com/Eastern_Algarve/Tavira/All_Property_Types/price-0.aspx",
        "https://www.landandhousesalgarve.com/Eastern_Algarve/Cabanas_de_Tavira/All_Property_Types/price-0.aspx",
    ]
    return scrape_generico("Land & Houses Algarve", urls, p,
                           seletores_extra=["a[href*='.aspx']"])

def scrape_togofor(p):
    """Togofor-Homes (AMI 6902, desde 2005) — escritórios em Vilamoura, Lagos
    e TAVIRA. Verificado por pesquisa 22/07; listagens em /en/...-for-sale/."""
    urls = [
        "https://www.togofor-homes.com/en/Tavira-villas-for-sale/",
        "https://www.togofor-homes.com/en/Properties-for-sale-in-the-Algarve-Portugal/",
    ]
    return scrape_generico("Togofor-Homes", urls, p)

def scrape_janela(p):
    """Janela Imobiliária (Faro) — 230+ imóveis, Sotavento.

    Plataforma eGO Real Estate — os mesmos seletores da Sunpoint/A.Unique/Mimosa
    funcionam aqui (a[href*='/imovel/']). URL de listagem confirmado por
    pesquisa web 22/07: /imoveis/venda/.

    NOTA: NÃO confundir com janelagrande.com (outra empresa, Alentejo, AMI
    12740). Esta é a Janela Algarvia (AMI 6110), sede em Faro."""
    res = []
    for u in ["https://www.janela-imobiliaria.com/imoveis/venda/",
              "https://www.janela-imobiliaria.com/comprarcasa"]:
        try:
            its,_ = scrape_playwright_html(
                "Janela Imobiliária", u,
                "a[href*='/imovel/'], a[href*='/imoveis/']",
                "Faro", p)
            res.extend(its)
            if len(res) >= 5: break
        except Exception as e:
            log.error(f"  Janela Imobiliária: {e}")
    return res, 1

def scrape_sortami(p):
    return scrape_playwright_html(
        "Sortami",
        "https://www.sortami.pt/imoveis",
        "[class*='item']", "Algarve", p
    )

def scrape_vernonalgarve(p):
    return scrape_playwright_html(
        "Vernon Algarve",
        "https://www.vernonalgarve.com/imoveis-para-venda",
        "li[class]", "Barlavento", p
    )

def scrape_dalmaportuguesa(p):
    """D'Alma Portuguesa — Playwright direto."""
    return scrape_playwright_html(
        "D'Alma Portuguesa",
        "https://www.dalmaportuguesa.com/imoveis",
        "[class*='property'],[class*='listing'],article,[class*='card']",
        "Algarve", p
    )

# ── TRIÂNGULO DOURADO ────────────────────────────────────

def scrape_qpsavills(p):
    urls=["https://www.quintaproperty.com/en/buy"]
    return scrape_generico("QP Savills", urls, p)

def scrape_jppproperties(p):
    """JPP (Vilamoura): descobrir_sites 21/07 → /properties/ deu 21 raw / 20
    válidos (185KB), e /en/properties_status/buy/ funciona também."""
    res=[]
    for u in ["https://jppproperties.com/properties/",
              "https://jppproperties.com/en/properties_status/buy/"]:
        try:
            its = scrape_generico("JPP Properties", [u], p,
                                  seletores_extra=["article",".property-item",".listing-item"])
            if isinstance(its, tuple): its = its[0]
            res.extend(its)
            if res: break
        except Exception as e:
            log.error(f"JPP: {e}")
    return res, 1


def scrape_yourluxury(p):
    """Your Luxury Property — Playwright com URL correta."""
    return scrape_playwright_html(
        "Your Luxury Property",
        "https://www.yourluxuryproperty.pt/imoveis",
        "[class*='property'],[class*='listing'],li[class],article",
        "Triângulo Dourado", p
    )

def scrape_insidevillas(p):
    urls=[f"https://www.inside-villas.com/for-sale?max_price={p['preco_max']}&bedrooms={p['quartos_min']}"]
    return scrape_generico("Inside-Villas", urls, p)

def scrape_cluttons(p):
    urls=[f"https://www.cluttons.com/algarve/properties-for-sale?max_price={p['preco_max']}&bedrooms={p['quartos_min']}"]
    return scrape_generico("Cluttons Algarve", urls, p)

def scrape_chestertons(p):
    urls=["https://www.chestertons.com/algarve/properties-for-sale"]
    return scrape_generico("Chestertons Algarve", urls, p)

# ── SOTAVENTO ────────────────────────────────────────────

def scrape_algarvemanta(p):
    urls=[f"https://casa.sapo.pt/comprar-apartamentos/tavira/?precomax={p['preco_max']}",
          f"https://casa.sapo.pt/comprar-moradias/tavira/?precomax={p['preco_max']}"]
    return scrape_generico("Algarve Manta Properties", urls, p)

# ── SITES DESATIVADOS (falham consistentemente) ────────────
# Timeout persistente — URLs provavelmente errados ou bloqueio total
SCRAPERS_DESATIVADOS = {
    # Após ronda de prova 21/07 com todas as capacidades novas:
    #
    # Vernon Algarve, D'Alma Portuguesa
    #   DIAG: "487 chars HTML | hrefs: [('mailto:help@moonshapes.pt', 1)]"
    #   Ambos os sites estão OFFLINE — página só devolve email de contacto
    #   do host (Moonshapes). Não é bloqueio; o negócio não tem site activo.
    "Vernon Algarve",
    "D'Alma Portuguesa",
    #
    # Barra Prime
    #   DIAG: "4,426 chars | Classes: [] | hrefs: []"
    #   Página completamente vazia (só header/footer). Site abandonado.
    "Barra Prime",
    #
    # Your Luxury Property
    #   DIAG: "2,046 chars | Classes: [] | hrefs: []"
    #   Site off ou permanentemente bloqueado (mesmo via Browserless).
    "Your Luxury Property",
    #
    # Tripalgarve
    #   DIAG: "44KB | Padrões de href: []"
    #   44KB de HTML renderizado com zero âncoras — SPA navega inteira por
    #   onclick. Payload-regex também não bateu; sem forma de extrair URLs.
    "Tripalgarve",
    #
}

# NOTA: Garvetur, Dils Portugal, Sortami e Boto Properties estavam nesta lista
# mas FUNCIONAM (7, 1, 1 e 16 resultados em 15/07) — foram retirados.
# Os que devolvem cards mas 0 items (Sunpoint 12→0, Mimosa, Algarve Unique 12→0,
# Algarve Dream 136→0) ficam ACTIVOS: o problema é o parsing, não o site.


# Sites que continuam bloqueados após fix_urls.py (HTTP 0 ou sem conteúdo em todos os URLs testados)
# Todos os sites foram migrados para Playwright ou tiveram URLs corrigidos
SCRAPERS_TEMPORARIAMENTE_INATIVOS = set()  # vazio — todos activos

SCRAPERS=[
    # ── PORTAIS (✅ todos funcionam) ─────────────────────
    ("Idealista",scrape_idealista),
    ("Imovirtual",scrape_imovirtual),
    ("Casa SAPO",scrape_casasapo),
    ("SuperCasa",scrape_supercasa),
    # ── SOTAVENTO (✅ todos funcionam) ───────────────────
    ("Casas do Sotavento",scrape_casasdosotavento),
    ("AlgarVila",scrape_algarvila),
    ("Villas Tavira",scrape_villastavira),
    ("Imocusto",scrape_imocusto),           # ✅ URL corrigido: /venda
    ("Algarve Manta Properties",scrape_algarvemanta),
    # ── REDES NACIONAIS (✅/🔧 corrigidos) ──────────────
    ("Engel & Völkers",scrape_engelvoelkers),
    ("ERA Imobiliária",scrape_era),
    ("RE/MAX",scrape_remax),
    ("KW Portugal",scrape_kwportugal),
    ("Coldwell Banker",scrape_coldwell),    # ✅ URL corrigido
    ("IAD Portugal",scrape_iad),            # 🔧 domínio correto: iadportugal.pt
    ("Fine & Country",scrape_fineandcountry),
    ("Century 21",scrape_century21),
    ("Arcada Imobiliária",scrape_arcada),
    # ── ALGARVE REGIÃO (✅/🔧 corrigidos) ───────────────
    ("BuyMe Property",scrape_buyme),
    ("Algarve Property",scrape_algarveproperty),
    ("Nurisimo",scrape_nurisimo),           # ✅ URL corrigido: /properties
    ("Golden Properties",scrape_goldenproperties), # ✅ URL corrigido
    ("Algarve Real Estate",scrape_algarverealestate),
    ("Espaços Algarve",scrape_espacosalgarve),
    ("Rede Real",scrape_redereal),
    ("VAP Real Estate",scrape_vaprealestate),
    ("Tripalgarve",scrape_tripalgarve),     # ✅ URL corrigido: sem www
    # ── BARLAVENTO ───────────────────────────────────────
    ("A1 Algarve",scrape_a1algarve),
    ("Boto Properties",scrape_boto),         # ✅ Playwright
    ("Sunpoint Properties",scrape_sunpoint), # ✅ Playwright
    ("Mimosa Properties",scrape_mimosaproperties), # ✅ Playwright
    ("Vernon Algarve",scrape_vernonalgarve), # ✅ Playwright
    # ── ALGARVE REGIÃO (Playwright) ──────────────────────
    ("Garvetur",scrape_garvetur),            # ✅ Playwright (garveturproperties.com)
    ("Algarve Unique Properties",scrape_algarveuniqueproperties), # ✅ Playwright
    ("Janela Imobiliária",scrape_janela),   # ✅ eGO — Faro (adicionado 22/07)
    ("Land & Houses Algarve",scrape_landhouses),  # ✅ ASP.NET — Tavira/Sotavento (22/07)
    ("Casas do Barlavento",scrape_casasdobarlavento),  # 22/07 sem filtro de valor
    ("Marcela Properties",scrape_marcela),   # 22/07
    ("Lagos Homes",scrape_lagoshomes),       # 22/07
    ("Mapro Real Estate",scrape_mapro),      # 22/07 luxo QDL
    ("VIP Algarve",scrape_vipalgarve),       # 22/07
    ("Togofor-Homes",scrape_togofor),        # ✅ HTTP — Tavira/Vilamoura (22/07)
    ("Sortami",scrape_sortami),              # ✅ Playwright
    ("D'Alma Portuguesa",scrape_dalmaportuguesa), # ✅ Playwright
    # ── TRIÂNGULO DOURADO (✅/🔧 corrigidos) ────────────
    ("QP Savills",scrape_qpsavills),        # ✅ URL corrigido: /en/buy
    ("JPP Properties",scrape_jppproperties),
    ("Inside-Villas",scrape_insidevillas),
    ("Cluttons Algarve",scrape_cluttons),
    ("Chestertons Algarve",scrape_chestertons), # ✅ URL corrigido
    ("Barra Prime",scrape_barraprime),       # ✅ Playwright
    # ── ANTERIORMENTE EM FALTA — agora com Playwright ────
    ("LNHouse",scrape_lnhouse),              # ✅ URL: /venda
    ("Algarve Dream Property",scrape_algarvedream), # ✅ Playwright
    ("Your Luxury Property",scrape_yourluxury), # ✅ Playwright
    ("Dils Portugal",scrape_dils),           # ✅ Playwright
    ("Sotheby's",scrape_sothebys),
    # ("Chave Nova",scrape_chavanova),        # ❌ ERR_NAME_NOT_RESOLVED - domínio não existe
    ("Villas Key",scrape_villaskey),
]

# ============================================================
# TELEGRAM
# ============================================================

def enviar_telegram(chat_id, texto):
    if not TELEGRAM_BOT_TOKEN or not chat_id: return
    try:
        requests.post(f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage",
            json={"chat_id":chat_id,"text":texto,"parse_mode":"HTML"},timeout=10)
    except Exception as e: log.warning(f"Telegram: {e}")

# ============================================================
# EMAIL
# ============================================================

def card_email(im, badge="NOVO", cor="#16a34a"):
    img_html=(f'<img src="{im["imagem_url"]}" style="width:100%;height:180px;object-fit:cover;">'
              if im.get("imagem_url") else
              '<div style="height:80px;background:#e5e7eb;text-align:center;line-height:80px;font-size:28px;">🏠</div>')
    extras=[]
    if im.get("area_m2"):  extras.append(f"📐 {im['area_m2']}m²")
    if im.get("quartos"):  extras.append(f"🛏 T{im['quartos']}")
    if im.get("preco_m2"): extras.append(f"💶 {im['preco_m2']:,}€/m²")
    if im.get("ano_construcao"): extras.append(f"🏗 {im['ano_construcao']}")
    score_html=f'<span style="float:right;background:#fef3c7;color:#92400e;padding:1px 6px;border-radius:4px;font-size:10px;">⭐{im.get("score",0)}</span>' if im.get("score") else ""
    preco_ant=(f'<span style="color:#aaa;text-decoration:line-through;font-size:12px;margin-right:4px;">{im["preco_antigo"]}</span>'
               if im.get("preco_antigo") else "")
    return f"""<div style="border:1px solid #e5e7eb;border-radius:8px;overflow:hidden;margin-bottom:14px;">
      <a href="{im['link']}">{img_html}</a>
      <div style="padding:12px;">
        {score_html}
        <span style="background:{cor};color:white;font-size:10px;font-weight:700;padding:2px 7px;border-radius:12px;">{badge}</span>
        <div style="margin-top:7px;"><a href="{im['link']}" style="color:#1e3a5f;font-weight:700;text-decoration:none;">{im['titulo']}</a></div>
        <div style="margin-top:5px;">{preco_ant}<span style="color:#16a34a;font-weight:700;font-size:16px;">{im['preco']}</span></div>
        {f'<div style="margin-top:4px;font-size:12px;color:#6b7280;">{" · ".join(extras)}</div>' if extras else ""}
        <div style="margin-top:4px;font-size:11px;color:#9ca3af;">📍 {im["zona"]} · 🏢 {im["fonte"]}</div>
        {f'<div style="margin-top:6px;font-size:12px;color:#4b5563;border-left:3px solid #e5e7eb;padding-left:8px;">{im["descricao"][:150]}...</div>' if im.get("descricao") else ""}
      </div>
    </div>"""

def _send_via_resend(assunto, html, destinatario):
    """Envia via Resend API — não usa SMTP (Railway bloqueia portas 465/587)."""
    key = os.getenv("RESEND_API_KEY","")
    # O Resend recusa remetentes de domínios não verificados (gmail.com, etc).
    # Sem domínio próprio verificado, usa-se o domínio de teste do Resend.
    _rem = os.getenv("EMAIL_REMETENTE","").strip()
    _dominio_proprio = _rem and not any(
        d in _rem.lower() for d in ("@gmail.", "@hotmail.", "@outlook.", "@yahoo.", "@sapo.", "@live."))
    remetente = _rem if _dominio_proprio else "Monitor Imóveis <onboarding@resend.dev>"
    if not key:
        return False, "RESEND_API_KEY não configurada"
    try:
        r = requests.post("https://api.resend.com/emails",
            headers={"Authorization": f"Bearer {key}",
                     "Content-Type": "application/json"},
            json={"from": remetente, "to": [destinatario],
                  "subject": assunto, "html": html},
            timeout=15)
        if r.status_code in (200,201,202):
            return True, "ok"
        return False, f"HTTP {r.status_code}: {r.text[:80]}"
    except Exception as e:
        return False, str(e)

def enviar_email(perfil, novos, baixas, reativados):
    if not novos and not baixas and not reativados: return
    partes=[]
    if novos: partes.append(f"{len(novos)} novo(s)")
    if baixas: partes.append(f"{len(baixas)} baixa(s)")
    if reativados: partes.append(f"{len(reativados)} reat.")
    assunto=f"🏠 {' · '.join(partes)} — {perfil['nome']}"
    todos=novos+baixas+reativados
    zonas_str=" · ".join(sorted(set(i["zona"] for i in todos))) if todos else ""
    cards_n="".join(card_email(i) for i in sorted(novos,key=lambda x:-x.get("score",0)))
    cards_b="".join(card_email(i,"📉 BAIXA","#dc2626") for i in baixas)
    cards_r="".join(card_email(i,"🔄 REATIVADO","#7c3aed") for i in reativados)
    s_b=f'<h3 style="color:#dc2626;margin-top:24px;">📉 Baixas ({len(baixas)})</h3>{cards_b}' if baixas else ""
    s_r=f'<h3 style="color:#7c3aed;margin-top:24px;">🔄 Reativados ({len(reativados)})</h3>{cards_r}' if reativados else ""
    html=f"""<html><body style="font-family:Arial,sans-serif;max-width:600px;margin:auto;background:#f9fafb;padding:20px;">
      <div style="background:white;border-radius:10px;padding:22px;box-shadow:0 1px 4px rgba(0,0,0,.08);">
        <h2 style="color:#1e3a5f;margin:0 0 12px;">🏠 {perfil['nome']}</h2>
        <p style="background:#f0f4ff;padding:10px;border-radius:6px;color:#555;font-size:13px;">
          🛏 T{perfil['quartos_min']}+ · 💰 Até {perfil['preco_max']:,}€ · 📍 {zonas_str}
        </p>
        <h3 style="color:#1e3a5f;margin-top:20px;">✨ Novos ({len(novos)}) — por relevância</h3>
        {cards_n}{s_b}{s_r}
        <p style="color:#aaa;font-size:11px;margin-top:20px;text-align:center;">
          {datetime.now().strftime('%d/%m/%Y %H:%M')}
        </p>
      </div></body></html>"""
    ok, motivo = _send_via_resend(assunto, html, perfil["email"])
    if ok:
        log.info(f"✉  Email → {perfil['email']}")
    else:
        log.error(f"Email alerta falha: {motivo}")

def enviar_resumo_semanal():
    for perfil in PERFIS:
        try:
            with get_db() as conn:
                with conn.cursor(row_factory=psycopg_rows.dict_row) as cur:
                    cur.execute("SELECT * FROM imoveis WHERE perfil_nome=%s AND criado_em>NOW()-INTERVAL '7 days' ORDER BY score DESC",(perfil["nome"],))
                    imoveis=[dict(r) for r in cur.fetchall()]
            if not imoveis: continue
            enviar_email(perfil,imoveis,[],[])
        except Exception as e: log.error(f"Resumo: {e}")

# ============================================================
# LOOP PRINCIPAL
# ============================================================

# Set global das fontes que correram sem excepção nesta ronda — usado por
# marcar_removidos para NÃO remover inventário de fontes que crasharam.
# Uma fonte que corre e devolve 0 items com sucesso É válida (n_raw>0 ou n_raw==0
# sem erro): pode ser um dia calmo. Só se excluem as que rebentaram.
_scraper_ok_lock = threading.Lock()
_scrapers_correram_ok = set()

def correr_scraper(args):
    """Corre um scraper individual — usado em paralelo."""
    nome, fn, perfil = args
    log.info(f"  → {nome}...")
    erros = ""; total = 0; pags = 1; excepcao = False
    t0 = time.time()
    try:
        r = fn(perfil)
        anuncios, pags = r if isinstance(r, tuple) else (r, 1)
        n_raw = len(anuncios or [])
        anuncios = [a for a in (anuncios or []) if validar(a, perfil, _log_descarte=True)]
        total = len(anuncios)
        n_desc = n_raw - total
        tempo = round((time.time()-t0)*1000)
        if total == 0:
            extra = f" [raw={n_raw} descartados={n_desc}]" if n_raw else " — scraper não encontrou cards"
            log.warning(f"    ⚠️ {nome}: 0 resultados ({tempo}ms){extra}")
        else:
            log.info(f"    ✅ {nome}: {total} resultados ({tempo}ms)")
    except Exception as e:
        erros = str(e)
        excepcao = True
        tempo = round((time.time()-t0)*1000)
        log.error(f"    ❌ {nome}: {e} ({tempo}ms)")
        anuncios = []
    if not excepcao:
        with _scraper_ok_lock:
            _scrapers_correram_ok.add(nome)
    registar_log_scraper(nome, perfil["nome"], total, 0, pags, erros, tempo)
    try:
        _update_scraper_stat(nome, total)
    except Exception: pass
    return anuncios

def verificar_perfil(perfil):
    _preflight_providers()
    log.info(f"▶ {perfil['nome']} — a correr {len(SCRAPERS)} scrapers em paralelo")
    todos_raw = []

    # Corre scrapers em paralelo com ThreadPoolExecutor
    from concurrent.futures import ThreadPoolExecutor, as_completed
    # SCRAPERS_DESATIVADOS era definida mas NUNCA usada — os sites mortos
    # continuavam a correr, ~13s cada por ronda.
    args = [(nome, fn, perfil) for nome, fn in SCRAPERS
             if nome not in _broken_scrapers
             and nome not in SCRAPERS_DESATIVADOS]
    if len(args) < len(SCRAPERS):
        n_off = len([1 for nome, _ in SCRAPERS if nome in SCRAPERS_DESATIVADOS])
        n_bad = len(SCRAPERS) - len(args) - n_off
        log.info(f"  ⏭ A saltar {len(SCRAPERS)-len(args)} scrapers "
                 f"({n_off} desactivados, {n_bad} em pausa automática)")
    
    _PW_NOMES = {
        "Boto Properties","Sunpoint Properties","Algarve Unique Properties",
        "Garvetur","Barra Prime","Mimosa Properties","Sortami","Vernon Algarve",
        "Janela Imobiliária",
        "Algarve Dream Property","Your Luxury Property","Dils Portugal",
        "D'Alma Portuguesa","Sotheby's",
        "KW Portugal","LNHouse",   # SPAs — reescritos para Playwright em 17/07
    }
    args_http = [a for a in args if a[0] not in _PW_NOMES]
    args_pw   = [a for a in args if a[0] in _PW_NOMES]
    log.info(f"  HTTP: {len(args_http)} scrapers | Playwright: {len(args_pw)} scrapers")

    with ThreadPoolExecutor(max_workers=3) as executor:
        futures = {executor.submit(correr_scraper, a): a[0] for a in args_http}
        for future in as_completed(futures):
            nome = futures[future]
            try:
                resultado = future.result(timeout=120)
                todos_raw.extend(resultado)
            except Exception as e:
                log.error(f"    {nome} thread error: {e}")

    if args_pw and PLAYWRIGHT_AVAILABLE:
        log.info(f"  🌐 Browser Playwright — {len(args_pw)} scrapers sequenciais")

        def _run_pw_scrapers():
            """Corre todos os scrapers Playwright num thread sem asyncio."""
            # Usa playwright (não patchright) para sync API estável
            try:
                from playwright.sync_api import sync_playwright as _spw_sync
            except ImportError:
                log.error("  playwright não instalado")
                return []
            resultados_pw = []
            try:
                with _spw_sync() as _spw:
                    _browser = None
                    # Preferência: Browserless (remoto, zero RAM local, 1 só ligação)
                    if BROWSERLESS_TOKEN:
                        try:
                            _browser = _spw.chromium.connect_over_cdp(
                                f"wss://chrome.browserless.io?token={BROWSERLESS_TOKEN}",
                                timeout=15000)
                            log.info("  🌐 Ligado ao Browserless (browser remoto)")
                        except Exception as _e:
                            log.warning(f"  Browserless indisponível ({_e}) — Chromium local")
                            _browser = None
                    if _browser is None:
                        _browser = _spw.chromium.launch(
                            headless=True, executable_path="/usr/bin/chromium",
                            args=["--no-sandbox","--disable-dev-shm-usage","--disable-gpu",
                                  "--single-process","--disable-blink-features=AutomationControlled"])
                    scrape_playwright_html._shared_browser = _browser
                    for a in args_pw:
                        try:
                            resultado = correr_scraper(a)
                            resultados_pw.extend(resultado)
                        except Exception as e:
                            log.error(f"    {a[0]} PW error: {e}")
                    scrape_playwright_html._shared_browser = None
                    _browser.close()
                    log.info("  🌐 Browser fechado")
            except Exception as e:
                log.error(f"  Playwright browser error: {e}")
            return resultados_pw

        # Lança em thread separada para garantir que não há asyncio loop
        import concurrent.futures
        with concurrent.futures.ThreadPoolExecutor(max_workers=1) as ex:
            future = ex.submit(_run_pw_scrapers)
            try:
                pw_results = future.result(timeout=300)
                todos_raw.extend(pw_results)
            except Exception as e:
                log.error(f"  PW thread error: {e}")

    # Enriquece novos itens com detalhes completos (amostra de 10 por ronda para não sobrecarregar)
    todos=deduplicar(todos_raw)
    novos_para_enriquecer=[i for i in todos if imovel_existe(i["id"])[0] is None][:10]
    for item in novos_para_enriquecer:
        item = enriquecer_com_detalhes(item)
        time.sleep(random.uniform(0.5,1.5))

    ids_ronda=[i["id"] for i in todos]
    novos=[]; baixas=[]; reativados=[]

    for item in todos:
        score=calcular_score(item,perfil); item["score"]=score
        p_ant,disp_ant=imovel_existe(item["id"])
        if p_ant is None:
            try:
                guardar_imovel(item,perfil["nome"],score); novos.append(item)
            except Exception as e:
                log.error(f"  guardar falhou [{item.get('fonte','?')}] {str(item.get('id',''))[:60]}: {e}")
        else:
            try:
                guardar_imovel(item,perfil["nome"],score)
            except Exception as e:
                log.error(f"  guardar falhou [{item.get('fonte','?')}] {str(item.get('id',''))[:60]}: {e}")
            pv_n=extrair_preco_valor(item["preco"]); pv_a=extrair_preco_valor(p_ant)
            if pv_n and pv_a and pv_n<pv_a:
                registar_mudanca_preco(item["id"],p_ant,item["preco"])
                item["preco_antigo"]=p_ant; baixas.append(item)
            if disp_ant==False:
                marcar_reativado(item["id"]); reativados.append(item)

    # Fontes que CORRERAM (mesmo com 0 items). Um site com nav só (0 items)
    # ainda existe — não é razão para marcar o inventário histórico como
    # removido. Só quando a fonte rebenta é que o inventário é poupado.
    # Uniao com fontes que devolveram items: se algo devolveu items mas a
    # thread rebentou depois, o inventário fica poupado na mesma.
    _fontes_com_items = {t.get("fonte") for t in todos if t.get("fonte")}
    with _scraper_ok_lock:
        _fontes_correram = set(_scrapers_correram_ok)
    _fontes_ok = _fontes_com_items | _fontes_correram
    removidos=marcar_removidos(ids_ronda,perfil["nome"],total_encontrados=len(todos),
                               fontes_ok=_fontes_ok)
    atualizar_snapshot_mercado(perfil["nome"])
    log.info(f"  N:{len(novos)} B:{len(baixas)} R:{len(reativados)} Rem:{len(removidos)}")

    if novos or baixas or reativados:
        enviar_email(perfil,novos,baixas,reativados)
        if perfil.get("telegram_chat"):
            msg=f"🏠 <b>{perfil['nome']}</b>\n"
            if novos: msg+=f"{len(novos)} novo(s)\n"
            if baixas: msg+=f"📉 {len(baixas)} baixa(s)\n"
            enviar_telegram(perfil["telegram_chat"],msg)

def verificar_creditos_scraperapi():
    """
    Verifica créditos disponíveis na ScraperAPI.
    Devolve (usados, limite, percentagem) ou None se não conseguir verificar.
    """
    if not SCRAPERAPI_KEY:
        return None
    try:
        r = requests.get(
            f"https://api.scraperapi.com/account?api_key={_get_active_key('scraperapi') or SCRAPERAPI_KEY}",
            timeout=10)
        data = r.json()
        used  = data.get("requestCount", 0)
        limit = data.get("requestLimit", 1000)
        pct   = int((used / limit) * 100) if limit else 100
        return used, limit, pct
    except Exception as e:
        log.warning(f"Não foi possível verificar créditos ScraperAPI: {e}")
        return None

_preflight_done = False
_preflight_lock = threading.Lock()

def _preflight_providers():
    """Testa cada provider ANTES de lançar scrapers em paralelo.
    Evita race condition onde 49 threads tentam ZenRows em simultâneo."""
    global _preflight_done
    with _preflight_lock:
        if _preflight_done: return
        # Marca como done APENAS APÓS os testes (não antes)
    TEST_URL = "https://www.example.com"
    log.info("  Preflight providers...")
    for prov, key, test_fn in [
        ("zenrows",    ZENROWS_KEY,     lambda: requests.get("https://api.zenrows.com/v1/", timeout=8,
            params={"url":TEST_URL,"apikey":ZENROWS_KEY,"js_render":"false"})),
        ("scrapingbee",SCRAPINGBEE_KEY, lambda: requests.get("https://app.scrapingbee.com/api/v1/", timeout=8,
            params={"api_key":_get_active_key("scrapingbee") or SCRAPINGBEE_KEY,"url":TEST_URL,"render_js":"false"})),
        ("scrapedo",   SCRAPEDO_KEY,    lambda: requests.get(
            f"https://api.scrape.do?token={_get_active_key('scrapedo') or SCRAPEDO_KEY}&url={requests.utils.quote(TEST_URL)}", timeout=8)),
        ("crawlbase",  CRAWLBASE_KEY,   lambda: requests.get(
            f"https://api.crawlbase.com/?token={CRAWLBASE_KEY}&url={requests.utils.quote(TEST_URL)}", timeout=8)),
        ("scrapingant",SCRAPINGANT_KEY, lambda: requests.get(
            "https://api.scrapingant.com/v2/general", timeout=8,
            params={"url": TEST_URL, "x-api-key": SCRAPINGANT_KEY, "browser": "false"},
            headers={"x-api-key": SCRAPINGANT_KEY})),
    ]:
        if not key or _is_exhausted(prov):
            log.info(f"    {prov}: ⏭ ignorado (sem key ou já esgotado)")
            continue
        try:
            r = test_fn()
            if len(r.text) < 500:
                msg = r.text[:150]
                for p, patterns in EXHAUSTED_PATTERNS:
                    if p == prov and any(pat.lower() in msg.lower() for pat in patterns):
                        _mark_exhausted(prov, msg)
                        break
                    # Pequena resposta = provider bloqueado/esgotado
                    log.warning(f"    {prov}: ⚠️ resposta suspeita ({len(r.text)} chars) — a marcar esgotado")
                    _mark_exhausted(prov, f"preflight suspeito ({len(r.text)} chars)")
            else:
                log.info(f"    {prov}: ✅ OK ({r.status_code})")
        except Exception as e:
            log.debug(f"    {prov} preflight: {e}")
    with _preflight_lock:
        _preflight_done = True  # marca como concluído

def verificar():
    log.info("="*55)
    log.info(f"Verificação: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    log.info("="*55)
    _load_provider_state()  # restaura providers esgotados após restart
    with _scraper_ok_lock:
        _scrapers_correram_ok.clear()  # ronda nova = estado limpo
    log.info(f"  Providers excluídos desta sessão: {sorted(_session_exhausted)}")

    # Verifica créditos — só para se ScraperAPI esgotado E sem outros proxies
    creditos = verificar_creditos_scraperapi()
    n_proxies = proxies_disponiveis()
    
    if creditos:
        used, limit, pct = creditos
        log.info(f"ScraperAPI: {used}/{limit} créditos usados ({pct}%)")

        # Key activa esgotada mas há outra? Roda e a ronda CONTINUA
        # (sem return — este bloco está inline em verificar(); um return
        #  aqui cancelava a ronda inteira)
        if pct >= 100 and _rotate_key("scraperapi"):
            log.info("  ScraperAPI: key seguinte activa — a continuar com ela")
            pct = 0  # key nova = créditos frescos; salta os ramos de esgotado

        if pct >= 100:
            if n_proxies <= 1:
                # Só tem ScraperAPI e está esgotado — para
                log.warning("⛔ ScraperAPI sem créditos e sem proxies alternativos! A saltar ronda.")
                try:
                    msg = MIMEMultipart("alternative")
                    msg["Subject"] = "⛔ Todos os proxies esgotados — Monitor pausado"
                    msg["From"] = EMAIL_REMETENTE
                    msg["To"]   = EMAIL_DESTINATARIO
                    msg.attach(MIMEText(f"""
                    <html><body style="font-family:Arial,sans-serif;padding:20px">
                    <h2 style="color:#dc2626">⛔ Todos os proxies esgotados!</h2>
                    <p>ScraperAPI: <b>{used}/{limit}</b> ({pct}%)</p>
                    <p>Sem ZenRows nem ScrapingBee configurados.</p>
                    <p>O monitor está <b>pausado</b> até ao reset mensal (dia 1).</p>
                    </body></html>""", "html"))
                    _assunto = msg["Subject"]
                    _html = "".join(p.get_payload(decode=True).decode("utf-8","replace")
                                    for p in msg.get_payload() if hasattr(p,"get_payload"))
                    ok, motivo = _send_via_resend(_assunto, _html, EMAIL_DESTINATARIO)
                    if ok: log.info("✉  Email enviado via Resend")
                    else: log.error(f"Email alerta falha: {motivo}")
                except Exception as e:
                    log.error(f"Email aviso créditos: {e}")
                return
            else:
                # Tem outros proxies — continua com ZenRows/ScrapingBee
                log.warning(f"⚠️ ScraperAPI esgotado — a usar ZenRows/ScrapingBee ({proxies_disponiveis()-1} proxy(s) alternativos)")

        elif pct >= 80:
            log.warning(f"⚠️ ScraperAPI a {pct}% do limite!")
            try:
                msg = MIMEMultipart("alternative")
                msg["Subject"] = f"⚠️ ScraperAPI a {pct}% do limite — Monitor Imóveis"
                msg["From"] = EMAIL_REMETENTE
                msg["To"]   = EMAIL_DESTINATARIO
                msg.attach(MIMEText(f"""
                <html><body style="font-family:Arial,sans-serif;padding:20px">
                <h2 style="color:#f59e0b">⚠️ ScraperAPI quase no limite!</h2>
                <p>Usados: <b>{used}/{limit}</b> ({pct}%)</p>
                <p>A continuar com ZenRows e ScrapingBee como backup.</p>
                </body></html>""", "html"))
                _assunto = msg["Subject"]
                _html = "".join(p.get_payload(decode=True).decode("utf-8","replace")
                                for p in msg.get_payload() if hasattr(p,"get_payload"))
                ok, motivo = _send_via_resend(_assunto, _html, EMAIL_DESTINATARIO)
                if ok: log.info("✉  Email enviado via Resend")
                else: log.error(f"Email alerta falha: {motivo}")
                log.info("Alerta 80% enviado por email.")
            except Exception as e:
                log.error(f"Email alerta 80%: {e}")

    perfis_ativos = get_perfis_db()
    log.info(f"{len(perfis_ativos)} perfil(is) ativo(s)")
    for p in perfis_ativos: verificar_perfil(p)
    verificar_scrapers_com_falha()
    log.info("Ronda concluída."); quit_driver()

# ============================================================
# AUTENTICAÇÃO
# ============================================================

def login_required(f):
    @functools.wraps(f)
    def decorated(*args,**kwargs):
        if not session.get("logged_in"):
            return redirect(url_for("login"))
        return f(*args,**kwargs)
    return decorated

# ============================================================
# IMT / CUSTOS
# ============================================================

def calcular_imt(preco, habitacao_propria=True):
    """Calcula IMT para habitação própria permanente em Portugal."""
    escaloes_hpp = [
        (97064, 0, 0),
        (132774, 0.02, 1941.28),
        (181034, 0.05, 5944.53),
        (301688, 0.07, 9552.23),
        (603289, 0.08, 12568.23),
        (1050400, 0.06, 0),  # taxa única acima de 603k
        (float("inf"), 0.075, 0),
    ]
    escaloes_outros = [
        (97064, 1, 0),
        (132774, 2, 970.64),
        (181034, 5, 4942.22),
        (301688, 7, 8571.86),
        (578598, 8, 11588.50),
        (float("inf"), 6, 0),
    ]
    escaloes = escaloes_hpp if habitacao_propria else escaloes_outros
    for limite, taxa, deducao in escaloes:
        if preco <= limite:
            if taxa == 0: return 0.0
            return round(preco * (taxa/100) - deducao, 2)
    return round(preco * 0.075, 2)

def calcular_custos_totais(preco, habitacao_propria=True):
    imt      = calcular_imt(preco, habitacao_propria)
    is_val   = round(preco * 0.008, 2)      # Imposto de Selo 0.8%
    registo  = round(250 + preco*0.001, 2)  # estimativa registo
    notario  = round(300 + preco*0.001, 2)  # estimativa escritura
    total    = round(imt+is_val+registo+notario, 2)
    return {"preco":preco,"imt":imt,"imposto_selo":is_val,
            "registo":registo,"notario":notario,"total_encargos":total,
            "total_aquisicao":round(preco+total,2)}

# ============================================================
# FLASK APP
# ============================================================

LOGIN_HTML = """<!DOCTYPE html>
<html lang="pt">
<head>
  <meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
  <title>Monitor Imóveis — Login</title>
  <style>
    *{box-sizing:border-box;margin:0;padding:0}
    body{font-family:Arial,sans-serif;background:linear-gradient(135deg,#0f2942,#1a6b8a);
         min-height:100vh;display:flex;align-items:center;justify-content:center;}
    .box{background:white;border-radius:16px;padding:40px;width:100%;max-width:380px;
         box-shadow:0 20px 60px rgba(0,0,0,.3);}
    .logo{text-align:center;margin-bottom:28px;}
    .logo .icon{font-size:48px;}
    .logo h1{font-size:22px;color:#0f2942;margin-top:8px;}
    .logo p{font-size:13px;color:#9ca3af;margin-top:4px;}
    label{display:block;font-size:13px;font-weight:600;color:#374151;margin-bottom:5px;}
    input{width:100%;padding:11px 14px;border:1.5px solid #e5e7eb;border-radius:8px;
          font-size:14px;margin-bottom:16px;outline:none;}
    input:focus{border-color:#1a6b8a;}
    button{width:100%;padding:12px;background:#0f2942;color:white;border:none;
           border-radius:8px;font-size:15px;font-weight:600;cursor:pointer;}
    button:hover{background:#1a6b8a;}
    .error{background:#fee2e2;color:#991b1b;padding:10px;border-radius:6px;
           font-size:13px;margin-bottom:14px;}
  </style>
</head>
<body>
  <div class="box">
    <div class="logo">
      <div class="icon">🏠</div>
      <h1>Monitor Imóveis</h1>
      <p>Algarve Sotavento</p>
    </div>
    {% if error %}<div class="error">{{ error }}</div>{% endif %}
    <form method="POST">
      <label>Utilizador</label>
      <input name="username" type="text" autocomplete="username" required>
      <label>Password</label>
      <input name="password" type="password" autocomplete="current-password" required>
      <button type="submit">Entrar →</button>
    </form>
  </div>
</body>
</html>"""

_dashboard_path = os.path.join(os.path.dirname(__file__), "dashboard.html")
DASHBOARD_HTML = ""
if os.path.exists(_dashboard_path):
    with open(_dashboard_path) as _f:
        DASHBOARD_HTML = _f.read()
else:
    DASHBOARD_HTML = "<!-- dashboard.html not found -->"

app = Flask(__name__)
app.secret_key = SECRET_KEY

@app.route("/login", methods=["GET","POST"])
def login():
    error = None
    if request.method == "POST":
        u = request.form.get("username","")
        p = request.form.get("password","")
        if u==DASHBOARD_USERNAME and p==DASHBOARD_PASSWORD:
            session["logged_in"] = True
            return redirect(url_for("dashboard"))
        error = "Utilizador ou password incorretos."
    return render_template_string(LOGIN_HTML, error=error)

@app.route("/logout")
def logout():
    session.clear(); return redirect(url_for("login"))

@app.route("/")
@login_required
def dashboard():
    return render_template_string(DASHBOARD_HTML)

@app.route("/api/imoveis")
@login_required
def api_imoveis():
    try:
        return jsonify(get_imoveis())
    except Exception as e:
        import traceback
        log.error(f"/api/imoveis falhou: {e}\n{traceback.format_exc()}")
        return jsonify({"erro": str(e)}), 500

@app.route("/api/stats")
@login_required
def api_stats():
    try: return jsonify(get_stats())
    except Exception as e: return jsonify({"erro":str(e)}),500

@app.route("/api/mercado")
@login_required
def api_mercado():
    try:
        dados = get_dados_mercado()
        return jsonify(dados)
    except Exception as e:
        log.error(f"api_mercado: {e}")
        return jsonify({"por_zona":[],"evolucao":[],"distribuicao":{}}), 200

@app.route("/api/imovel/favorito", methods=["POST"])
@login_required
def api_favorito():
    data=request.get_json(); iid=data.get("id")
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("UPDATE imoveis SET favorito=NOT favorito WHERE id=%s RETURNING favorito",(iid,))
            row=cur.fetchone()
        conn.commit()
    return jsonify({"favorito":row[0] if row else False})

@app.route("/api/imovel/estado", methods=["POST"])
@login_required
def api_estado():
    data=request.get_json()
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("UPDATE imoveis SET estado=%s WHERE id=%s",(data.get("estado"),data.get("id")))
        conn.commit()
    return jsonify({"ok":True})

@app.route("/api/imovel/nota", methods=["POST"])
@login_required
def api_nota():
    data=request.get_json()
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("UPDATE imoveis SET nota=%s WHERE id=%s",(data.get("nota",""),data.get("id")))
        conn.commit()
    return jsonify({"ok":True})

@app.route("/api/imovel/<iid>")
@login_required
def api_imovel_detalhe(iid):
    with get_db() as conn:
        with conn.cursor(row_factory=psycopg_rows.dict_row) as cur:
            cur.execute("SELECT * FROM imoveis WHERE id=%s",(iid,))
            row=cur.fetchone()
    if not row: return jsonify({"erro":"não encontrado"}),404
    d=dict(row)
    for k in ["criado_em","atualizado_em","removido_em","reativado_em"]:
        if d.get(k): d[k]=d[k].isoformat()
    return jsonify(d)

@app.route("/api/comparar", methods=["POST"])
@login_required
def api_comparar():
    ids=request.get_json().get("ids",[])
    if len(ids)<2 or len(ids)>3: return jsonify({"erro":"Seleciona 2 ou 3 imóveis"}),400
    imoveis=[]
    with get_db() as conn:
        with conn.cursor(row_factory=psycopg_rows.dict_row) as cur:
            for iid in ids:
                cur.execute("SELECT * FROM imoveis WHERE id=%s",(iid,))
                row=cur.fetchone()
                if row:
                    d=dict(row)
                    for k in ["criado_em","atualizado_em"]:
                        if d.get(k): d[k]=d[k].isoformat()
                    imoveis.append(d)
    custos=[calcular_custos_totais(i["preco_valor"]) for i in imoveis if i.get("preco_valor")]
    return jsonify({"imoveis":imoveis,"custos":custos})

@app.route("/api/custos")
@login_required
def api_custos():
    try:
        preco=int(request.args.get("preco",0))
        hpp=request.args.get("hpp","true").lower()=="true"
        return jsonify(calcular_custos_totais(preco,hpp))
    except Exception as e: return jsonify({"erro":str(e)}),400

@app.route("/api/visitas", methods=["GET","POST"])
@login_required
def api_visitas():
    if request.method=="GET":
        iid=request.args.get("imovel_id")
        return jsonify(get_visitas(iid))
    data=request.get_json()
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                INSERT INTO visitas(imovel_id,data_visita,nota,avaliacao)
                VALUES(%s,%s,%s,%s)
            """, (data["imovel_id"],data["data_visita"],
                  data.get("nota",""),data.get("avaliacao")))
        conn.commit()
    return jsonify({"ok":True})

@app.route("/api/exportar/excel")
@login_required
def api_exportar_excel():
    """Exporta favoritos para Excel."""
    try:
        import io
        try: import openpyxl
        except ImportError: return jsonify({"erro":"openpyxl não instalado"}),500
        with get_db() as conn:
            with conn.cursor(row_factory=psycopg_rows.dict_row) as cur:
                cur.execute("""
                    SELECT titulo,preco,area_m2,quartos,preco_m2,ano_construcao,
                           zona,fonte,estado,nota,score,link,criado_em
                    FROM imoveis WHERE favorito=TRUE ORDER BY score DESC
                """)
                rows=[dict(r) for r in cur.fetchall()]
        wb=openpyxl.Workbook(); ws=wb.active; ws.title="Favoritos"
        headers=["Título","Preço","Área m²","Quartos","€/m²","Ano","Zona","Fonte","Estado","Notas","Score","Link","Data"]
        ws.append(headers)
        from openpyxl.styles import Font, PatternFill
        for cell in ws[1]:
            cell.font=Font(bold=True,color="FFFFFF")
            cell.fill=PatternFill(fill_type="solid",fgColor="0F2942")
        for r in rows:
            ws.append([r.get("titulo"),r.get("preco"),r.get("area_m2"),r.get("quartos"),
                       r.get("preco_m2"),r.get("ano_construcao"),r.get("zona"),r.get("fonte"),
                       r.get("estado"),r.get("nota"),r.get("score"),r.get("link"),
                       str(r.get("criado_em",""))[:10]])
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = max(len(str(cell.value or "")) for cell in col)+4
        buf=io.BytesIO(); wb.save(buf); buf.seek(0)
        resp=make_response(buf.read())
        resp.headers["Content-Type"]="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        resp.headers["Content-Disposition"]=f"attachment; filename=favoritos_algarve_{datetime.now().strftime('%Y%m%d')}.xlsx"
        return resp
    except Exception as e: return jsonify({"erro":str(e)}),500

@app.route("/api/mapa")
@login_required
def api_mapa():
    """Imóveis com coordenadas GPS para o mapa."""
    with get_db() as conn:
        with conn.cursor(row_factory=psycopg_rows.dict_row) as cur:
            cur.execute("""
                SELECT id,titulo,preco,preco_valor,zona,fonte,score,
                       lat,lng,imagem_url,estado,favorito
                FROM imoveis WHERE disponivel=TRUE AND lat IS NOT NULL AND lng IS NOT NULL
                ORDER BY score DESC LIMIT 500
            """)
            return jsonify([dict(r) for r in cur.fetchall()])

@app.route("/api/push/subscribe", methods=["POST"])
@login_required
def api_push_subscribe():
    data=request.get_json()
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("INSERT INTO push_subscriptions(endpoint,p256dh,auth) VALUES(%s,%s,%s) ON CONFLICT(endpoint) DO NOTHING",
                    (data["endpoint"],data["keys"]["p256dh"],data["keys"]["auth"]))
            conn.commit()
        return jsonify({"ok":True})
    except Exception as e: return jsonify({"erro":str(e)}),500

@app.route("/api/perfis", methods=["GET"])
@login_required
def api_perfis_get():
    try:
        perfis = get_todos_perfis_db()
        for p in perfis:
            for k in ["criado_em","atualizado_em"]:
                if p.get(k): p[k] = str(p[k])
        return jsonify(perfis)
    except Exception as e:
        return jsonify({"erro":str(e)}),500

@app.route("/api/perfis", methods=["POST"])
@login_required
def api_perfis_criar():
    try:
        d = request.get_json()
        pid = criar_perfil_db(
            d["nome"], d["email"],
            int(d.get("preco_max",200000)), int(d.get("quartos_min",2)),
            d.get("tipos",["apartamentos","moradias-e-vivendas"]),
            d.get("zonas",["faro","tavira","olhao","vila-real-de-santo-antonio","castro-marim"])
        )
        return jsonify({"ok":True,"id":pid})
    except Exception as e:
        return jsonify({"erro":str(e)}),500

@app.route("/api/perfis/<int:pid>", methods=["PUT"])
@login_required
def api_perfis_atualizar(pid):
    try:
        d = request.get_json()
        atualizar_perfil_db(
            pid, d["nome"], d["email"],
            int(d.get("preco_max",200000)), int(d.get("quartos_min",2)),
            d.get("tipos",["apartamentos","moradias-e-vivendas"]),
            d.get("zonas",["faro","tavira","olhao"]),
            d.get("ativo",True)
        )
        return jsonify({"ok":True})
    except Exception as e:
        return jsonify({"erro":str(e)}),500

@app.route("/api/perfis/<int:pid>", methods=["DELETE"])
@login_required
def api_perfis_apagar(pid):
    try:
        apagar_perfil_db(pid)
        return jsonify({"ok":True})
    except Exception as e:
        return jsonify({"erro":str(e)}),500

@app.route("/api/perfis/teste-email", methods=["POST"])
@login_required
def api_teste_email():
    try:
        d = request.get_json()
        email_dest = d.get("email", EMAIL_DESTINATARIO)
        msg = MIMEMultipart("alternative")
        msg["Subject"] = "✅ Teste — Monitor Imóveis Algarve"
        msg["From"] = EMAIL_REMETENTE
        msg["To"] = email_dest
        msg.attach(MIMEText(f"""
        <html><body style="font-family:Arial,sans-serif;padding:20px">
        <h2 style="color:#0f2942">✅ Email de teste enviado com sucesso!</h2>
        <p>O Monitor de Imóveis está configurado corretamente.</p>
        <p style="color:#888;font-size:12px">{datetime.now().strftime('%d/%m/%Y %H:%M')}</p>
        </body></html>""", "html"))
        _assunto = msg["Subject"]
        _html_body = "".join(p.get_payload(decode=True).decode("utf-8","replace")
                             for p in msg.get_payload() if hasattr(p,"get_payload"))
        ok, motivo = _send_via_resend(_assunto, _html_body, email_dest)
        if ok:
            log.info("✉  Email enviado via Resend")
        else:
            log.error(f"Email alerta falha: {motivo}")
        return jsonify({"ok":True})
    except Exception as e:
        return jsonify({"erro":str(e)}),500

def verificar_limite_scraperapi():
    """Verifica uso de todos os proxies e envia alerta se algum >80%."""
    # Log estatísticas de rotação
    stats = get_proxy_stats()
    if stats: log.info(f"Proxy stats: {stats}")
    
    if not SCRAPERAPI_KEY: return
    try:
        r = requests.get(
            f"https://api.scraperapi.com/account?api_key={_get_active_key('scraperapi') or SCRAPERAPI_KEY}",
            timeout=10)
        data = r.json()
        used    = data.get("requestCount", 0)
        limit   = data.get("requestLimit", 1000)
        pct     = int((used / limit) * 100) if limit else 0
        log.info(f"ScraperAPI: {used}/{limit} pedidos ({pct}%)")
        if pct >= 80:
            msg = MIMEMultipart("alternative")
            msg["Subject"] = f"⚠️ ScraperAPI a {pct}% do limite ({used}/{limit} pedidos)"
            msg["From"] = EMAIL_REMETENTE
            msg["To"]   = EMAIL_DESTINATARIO
            msg.attach(MIMEText(f"""
            <html><body style="font-family:Arial,sans-serif;padding:20px">
            <h2 style="color:#dc2626">⚠️ ScraperAPI quase no limite!</h2>
            <p>Usados: <b>{used}/{limit}</b> pedidos ({pct}%)</p>
            <p>Considera fazer upgrade em <a href="https://scraperapi.com">scraperapi.com</a></p>
            </body></html>""", "html"))
            _assunto = msg["Subject"]
            _html_body = "".join(p.get_payload(decode=True).decode("utf-8","replace")
                                 for p in msg.get_payload() if hasattr(p,"get_payload"))
            ok, motivo = _send_via_resend(_assunto, _html_body, EMAIL_DESTINATARIO)
            if ok: log.info("✉  Email enviado via Resend")
            else: log.error(f"Email alerta falha: {motivo}")
            log.warning(f"⚠️ Alerta ScraperAPI enviado ({pct}%)")
        return {"used": used, "limit": limit, "pct": pct}
    except Exception as e:
        log.warning(f"verificar_limite_scraperapi: {e}")
        return {}

@app.route("/api/scraperapi/stats")
@login_required
def api_scraperapi_stats():
    stats = verificar_limite_scraperapi()
    proxy_stats = get_proxy_stats()
    return jsonify({"scraperapi": stats or {}, "rotacao": proxy_stats})

@app.route("/api/historico_precos")
@login_required
def api_historico_precos():
    imovel_id = request.args.get("imovel_id","")
    try:
        with get_db() as conn:
            with conn.cursor(row_factory=psycopg_rows.dict_row) as cur:
                cur.execute("""
                    SELECT preco_antigo, preco_novo, alterado_em::TEXT
                    FROM historico_precos WHERE imovel_id=%s
                    ORDER BY alterado_em DESC
                """, (imovel_id,))
                return jsonify([dict(r) for r in cur.fetchall()])
    except Exception as e:
        return jsonify([])

@app.route("/api/imovel/apagar", methods=["DELETE"])
@login_required
def api_apagar_imovel():
    try:
        data = request.get_json()
        iid = data.get("id")
        if not iid:
            return jsonify({"erro":"ID não fornecido"}), 400
        with get_db() as conn:
            with conn.cursor() as cur:
                # Apaga também histórico de preços e visitas
                cur.execute("DELETE FROM historico_precos WHERE imovel_id=%s", (iid,))
                cur.execute("DELETE FROM visitas WHERE imovel_id=%s", (iid,))
                cur.execute("DELETE FROM imoveis WHERE id=%s", (iid,))
                apagados = cur.rowcount
            conn.commit()
        if apagados:
            log.info(f"Imóvel apagado: {iid[:50]}")
            return jsonify({"ok": True})
        else:
            return jsonify({"erro": "Imóvel não encontrado"}), 404
    except Exception as e:
        log.error(f"api_apagar_imovel: {e}")
        return jsonify({"erro": str(e)}), 500

@app.route("/api/provider_status")
@login_required
def api_provider_status():
    """Estado completo dos providers com métricas e circuit breaker."""
    result = provider_status_dict()
    # Add circuit breaker state
    for p, d in result.items():
        pd = _provider_data.get(p, {})
        d["circuit_open"]    = pd.get("circuit_open", False)
        d["consecutive_failures"] = pd.get("consecutive_failures", 0)
        if pd.get("circuit_open_until"):
            remaining = max(0, int((pd["circuit_open_until"] - _dt.now().timestamp())/60))
            d["circuit_open_minutes_remaining"] = remaining
    return jsonify(result)

@app.route("/api/price_history/<path:imovel_id>")
@login_required
def api_price_history(imovel_id):
    """Histórico de preços de um imóvel."""
    try:
        with get_db() as conn:
            rows = conn.execute("""
                SELECT preco, registado_em FROM price_history
                WHERE imovel_id=%s ORDER BY registado_em
            """, (imovel_id,)).fetchall()
            hist = [{"preco": r[0], "data": r[1].strftime("%d/%m/%Y")} for r in rows]
            # Current price
            cur = conn.execute(
                "SELECT preco_valor, preco_inicial, first_seen, dias_online FROM imoveis WHERE id=%s",
                (imovel_id,)).fetchone()
            return jsonify({
                "historico": hist,
                "preco_atual": cur[0] if cur else None,
                "preco_inicial": cur[1] if cur else None,
                "first_seen": cur[2].strftime("%d/%m/%Y") if cur and cur[2] else None,
                "dias_online": cur[3] if cur else None,
            })
    except Exception as e:
        return jsonify({"erro": str(e)}), 500

@app.route("/api/domain_stats")
@login_required
def api_domain_stats():
    """Estatísticas por domínio × provider."""
    out = {}
    for domain, providers in _domain_stats.items():
        out[domain] = {}
        for provider, stats in providers.items():
            total = stats["total"]
            if total == 0: continue
            suc  = stats["success"]
            lats = stats["latencies"]
            conf = min(1.0, total / 50)
            out[domain][provider] = {
                "success_rate": round(suc/total*100, 1),
                "total":        total,
                "confidence":   round(conf, 2),
                "avg_latency_ms": round(sum(lats)/len(lats)) if lats else None,
            }
        # Sort by success_rate
        out[domain] = dict(sorted(out[domain].items(),
            key=lambda x: x[1]["success_rate"], reverse=True))
    return jsonify(out)

@app.route("/api/logs")
def api_logs():
    """Endpoint para consulta remota de logs — protegido por API key."""
    key = request.headers.get("X-API-Key") or request.args.get("key", "")
    if key != os.getenv("LOGS_API_KEY", "algarve2026logs"):
        return jsonify({"erro": "chave inválida"}), 401
    n     = int(request.args.get("n", 100))
    level = request.args.get("level", "").upper()  # INFO, WARNING, ERROR
    filtro= request.args.get("q", "").lower()       # texto livre
    logs  = list(_log_buffer)[-n:]
    if level:
        logs = [l for l in logs if l["level"] == level]
    if filtro:
        logs = [l for l in logs if filtro in l["msg"].lower()]
    return jsonify({
        "total": len(_log_buffer),
        "retornados": len(logs),
        "logs": logs
    })

@app.route("/api/system_health")
@login_required
def api_system_health():
    """Health check geral do sistema — ideal para dashboards e alertas."""
    metrics = _scraper_metrics
    total   = len(SCRAPERS)
    healthy = sum(1 for m in metrics.values() if m.get("status")=="ok")
    warning = sum(1 for m in metrics.values() if m.get("status")=="vazio")
    error   = sum(1 for m in metrics.values() if m.get("status")=="erro")
    providers = provider_status_dict()
    active_p  = sum(1 for p in providers.values() if not p.get("exhausted") and not p.get("circuit_open"))
    cooldown_p= sum(1 for p in providers.values() if p.get("exhausted") or p.get("circuit_open"))
    # Last run info from DB
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT COUNT(*) FROM imoveis WHERE disponivel=TRUE")
                total_props = cur.fetchone()[0]
                cur.execute("SELECT COUNT(*) FROM imoveis WHERE criado_em > NOW()-INTERVAL '24 hours'")
                new_24h = cur.fetchone()[0]
    except: total_props = new_24h = 0
    total_ms  = sum(m.get("ms",0) for m in metrics.values())
    total_items = sum(m.get("items",0) for m in metrics.values())
    return jsonify({
        "scrapers": {
            "total": total, "healthy": healthy,
            "warning": warning, "error": error,
            "ran": len(metrics),
        },
        "providers": {
            "active": active_p, "cooldown": cooldown_p,
            "detail": providers,
        },
        "last_run": {
            "duration_seconds": round(total_ms/1000),
            "properties_found": total_items,
            "new_properties": new_24h,
            "total_in_db": total_props,
        },
        "notifications": {
            "telegram": bool(os.getenv("TELEGRAM_TOKEN")),
            "discord":  bool(os.getenv("DISCORD_WEBHOOK")),
            "email":    bool(os.getenv("RESEND_API_KEY") or os.getenv("EMAIL_REMETENTE")),
            "webhook":  bool(os.getenv("WEBHOOK_URL")),
        }
    })

@app.route("/api/saude")
@login_required
def api_saude():
    """Dashboard de saúde — estado de todos os scrapers."""
    try:
        with get_db() as conn:
            with conn.cursor(row_factory=psycopg_rows.dict_row) as cur:
                cur.execute("""
                    SELECT
                        sl.fonte, sl.perfil_nome,
                        MAX(sl.executado_em)::TEXT ultima_exec,
                        SUM(sl.total) total_imoveis,
                        COUNT(*) total_rondas,
                        COUNT(*) FILTER (WHERE sl.erros != '' AND sl.erros IS NOT NULL) rondas_com_erro,
                        0::INTEGER avg_ms,
                        ROUND(100.0 * COUNT(*) FILTER (WHERE sl.total>0) / NULLIF(COUNT(*),0)) taxa_sucesso
                    FROM scraper_logs sl
                    -- (scraper_stats joined removed: schema mismatch)
                    WHERE sl.executado_em > NOW() - INTERVAL '7 days'
                    GROUP BY sl.fonte, sl.perfil_nome
                    ORDER BY sl.fonte
                """)
                scrapers = [dict(r) for r in cur.fetchall()]
        # Add proxy stats
        proxy_info = get_proxy_stats()
        # Calculate health score for each scraper
        for s in scrapers:
            taxa = float(s.get("taxa_sucesso") or 0)
            tempo = float(s.get("avg_ms") or 0)
            media = float(s.get("total_imoveis") or 0) / max(float(s.get("total_rondas") or 1), 1)
            # Days since last error
            ultima_falha = None
            if s.get("ultima_exec"):
                try:
                    from datetime import datetime
                    dt = datetime.fromisoformat(str(s["ultima_exec"])[:19])
                    ultima_falha = (datetime.now() - dt).days
                except: pass
            score, label, emoji = calcular_saude_scraper(taxa, tempo, media, ultima_falha)
            s["health_score"] = score
            s["health_label"] = label
            s["health_emoji"] = emoji
        return jsonify({"scrapers": scrapers, "proxies": proxy_info})
    except Exception as e:
        return jsonify({"erro": str(e)}), 500

@app.route("/api/search")
@login_required
def api_search():
    """Pesquisa semântica simples por texto livre."""
    q = request.args.get("q","").lower().strip()
    if not q: return jsonify([])
    try:
        imoveis = get_imoveis(limite=500)
        # Parse query
        resultado = []
        for im in imoveis:
            texto = f"{im.get('titulo','')} {im.get('descricao','')} {im.get('zona','')}".lower()
            # Check all words in query
            palavras = q.split()
            if all(p in texto for p in palavras):
                resultado.append(im)
        return jsonify(sorted(resultado, key=lambda x: -x.get("score",0))[:50])
    except Exception as e:
        return jsonify({"erro": str(e)}), 500

@app.route("/api/new")
@login_required
def api_new():
    """Imóveis das últimas 24h."""
    try:
        with get_db() as conn:
            with conn.cursor(row_factory=psycopg_rows.dict_row) as cur:
                cur.execute("""
                    SELECT * FROM imoveis
                    WHERE criado_em > NOW() - INTERVAL '24 hours'
                    ORDER BY score DESC LIMIT 50
                """)
                rows = [dict(r) for r in cur.fetchall()]
                for r in rows:
                    for k in ["criado_em","atualizado_em"]:
                        if r.get(k): r[k] = r[k].isoformat()
                return jsonify(rows)
    except Exception as e:
        return jsonify({"erro": str(e)}), 500

@app.route("/api/removed")
@login_required
def api_removed():
    """Imóveis removidos recentemente."""
    try:
        with get_db() as conn:
            with conn.cursor(row_factory=psycopg_rows.dict_row) as cur:
                cur.execute("""
                    SELECT * FROM imoveis
                    WHERE disponivel=FALSE AND removido_em > NOW() - INTERVAL '7 days'
                    ORDER BY removido_em DESC LIMIT 50
                """)
                rows = [dict(r) for r in cur.fetchall()]
                for r in rows:
                    for k in ["criado_em","removido_em"]:
                        if r.get(k): r[k] = r[k].isoformat()
                return jsonify(rows)
    except Exception as e:
        return jsonify({"erro": str(e)}), 500

@app.route("/api/favorites")
@login_required
def api_favorites():
    """Todos os favoritos."""
    try:
        with get_db() as conn:
            with conn.cursor(row_factory=psycopg_rows.dict_row) as cur:
                cur.execute("""
                    SELECT * FROM imoveis WHERE favorito=TRUE ORDER BY score DESC
                """)
                rows = [dict(r) for r in cur.fetchall()]
                for r in rows:
                    for k in ["criado_em","atualizado_em"]:
                        if r.get(k): r[k] = r[k].isoformat()
                return jsonify(rows)
    except Exception as e:
        return jsonify({"erro": str(e)}), 500

@app.route("/api/status_scraper")
@login_required
def api_status_scraper():
    """Estado do scraper para a barra de estado do dashboard."""
    try:
        with get_db() as conn:
            with conn.cursor(row_factory=psycopg_rows.dict_row) as cur:
                # Última execução
                cur.execute("""
                    SELECT MAX(executado_em)::TEXT as ultima,
                           COUNT(DISTINCT fonte) as total_fontes
                    FROM scraper_logs
                    WHERE executado_em > NOW() - INTERVAL '48 hours'
                """)
                row = cur.fetchone()
                ultima = row["ultima"] if row else None
                total = row["total_fontes"] or 0

                # Scrapers OK vs falhas na última ronda
                cur.execute("""
                    SELECT fonte,
                           MAX(executado_em)::TEXT ultima_exec,
                           SUM(total) total_imoveis,
                           MAX(erros) erro
                    FROM scraper_logs
                    WHERE executado_em > NOW() - INTERVAL '48 hours'
                    GROUP BY fonte
                """)
                scrapers = cur.fetchall()
                ok = [s for s in scrapers if (s["total_imoveis"]or 0)>0]
                fail = [s for s in scrapers if (s["total_imoveis"]or 0)==0]

        # Calcular próxima execução (24h após última)
        proxima = None
        if ultima:
            try:
                dt_ultima = datetime.fromisoformat(ultima[:19])
                dt_proxima = dt_ultima + timedelta(hours=24)
                proxima = dt_proxima.isoformat()
            except: pass

        return jsonify({
            "ultima_execucao": ultima,
            "proxima_execucao": proxima,
            "total_scrapers": total,
            "scrapers_ok": len(ok),
            "scrapers_fail": len(fail),
            "proxies_ativos": proxies_disponiveis(),
            "falhas": [{"fonte":s["fonte"],"erro":s["erro"],"ultima_exec":s["ultima_exec"]}
                       for s in fail]
        })
    except Exception as e:
        return jsonify({"erro": str(e)}), 500

@app.route("/health")
def health(): return jsonify({"status":"ok","ts":datetime.now().isoformat(),"proxies":proxies_disponiveis()})

# ============================================================
# ARRANQUE
# ============================================================

if __name__=="__main__":
    log.info("╔══════════════════════════════════════════════════╗")
    log.info("║   Monitor de Imóveis — Algarve  v4.0             ║")
    log.info("╚══════════════════════════════════════════════════╝")
    if DATABASE_URL:
        init_db()
        sincronizar_perfis_iniciais()
        # Geocodifica imóveis existentes em background
        threading.Thread(target=geocodificar_existentes, daemon=True).start()
    else: log.warning("DATABASE_URL não definida!")
    log.info(f"Login configurado: {DASHBOARD_USERNAME} (password definida)")
    log.info(f"{len(SCRAPERS)} fontes | {len(PERFIS)} perfil(is) | cada {INTERVALO_HORAS}h")
    def run_server():
        try:
            # Use gunicorn in production if available
            import subprocess, sys
            cmd = [sys.executable, "-m", "gunicorn",
                   f"--bind=0.0.0.0:{PORT}",
                   "--workers=1", "--threads=4",
                   "--worker-class=gthread",
                   "--timeout=120",
                   "algarve_monitor:app"]
            _gunicorn_proc = subprocess.Popen(cmd)
            log.info(f"Dashboard (gunicorn) em http://localhost:{PORT} (pid={_gunicorn_proc.pid})")
        except Exception:
            # Fallback to Flask dev server
            app.run(host="0.0.0.0", port=PORT, use_reloader=False)
            log.info(f"Dashboard (flask) em http://localhost:{PORT}")
    threading.Thread(target=run_server, daemon=True).start()
    log.info(f"Dashboard em http://localhost:{PORT}")
    verificar()
    # Corre uma vez por dia às 08:00 — poupa pedidos ScraperAPI
    schedule.every(24).hours.do(verificar)  # uma vez por dia
    schedule.every().monday.at("08:00").do(enviar_resumo_semanal)
    schedule.every().day.at("09:00").do(verificar_scrapers_com_falha)
    schedule.every().day.at("20:00").do(verificar_limite_scraperapi)
    schedule.every(24).hours.do(geocodificar_existentes)
    log.info("A monitorizar. Ctrl+C para parar.\n")
    while True: schedule.run_pending(); time.sleep(60)
# v4.1
